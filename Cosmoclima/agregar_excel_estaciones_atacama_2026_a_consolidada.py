#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
agregar_excel_estaciones_atacama_2026_a_consolidada.py -- suma a la tabla
consolidada los datos diarios reales de 17 estaciones de la Región de
Atacama que Alexis mandó en 'datos/Excel Estaciones/' (09-ago-2026),
formato "Informe Anual de Agua Caída" (DMC/FDF/INIA), cobertura real
2019-2026 (algunas desde 2005-2012) -- el tramo que faltaba para densificar
el norte del gradiente latitudinal.

Formato de cada archivo: hoja de metadata (ANTECEDENTE/ANTECEDENTES/nombre
propio) con Código Nacional, Nombre, Propietario, Latitud, Longitud (ambos
×100000 sin punto decimal), Altura(Mts.) -- ESTA es la fuente de elevación
real por estación, no satelital, verificada contra el catastro DMC/DGA
(coincide exacto donde se cruzó). Luego una hoja por año (grilla Día×Mes,
13 columnas), más una hoja "AGOSTO-2026"-style que es la grilla del año en
curso completa hasta el mes más reciente (NO una hoja separada por mes --
el archivo de Caldera trae además JULIO2026/JUNIO2026/MAYO2026 como
reportes redundantes de un solo mes, con OTRO layout; se ignoran a
propósito porque AGOSTO-2026 ya los incluye).

Convención de datos (verificada 09-ago-2026 cruzando "Total Mensual" contra
"N° de días con Agua>=0.1" para un mes con huecos): "s/p" = SIN
PRECIPITACIÓN, dato real confirmado = 0.0mm. "." = SIN OBSERVACIÓN ese día,
se descarta -- igual que LLUVIA_DIARIA_1966_2017, nunca se inventa.

Convención de nombre de localidad: igual que
agregar_estaciones_julio2026_a_consolidada.py (el precedente directo de
este script) -- f"{nombre} (DMC/INIA {codigo})". Para estaciones que ya
tienen un registro histórico largo bajo OTRO nombre de localidad (ej.
"Desierto de Atacama, Caldera (DMC)" 2005-2025 vs esta "...Ad. (DMC/INIA
270008)"), NO se tocan ni se fusionan los nombres viejos -- se declara el
solape como conocido (ver SOLAPES_CONOCIDOS al final) en vez de arriesgar
un merge no pedido de series con linaje de fuente distinto.
"""
import csv
import glob
import os
import re
import sqlite3

import openpyxl

CARPETA_EXCEL = "/Users/alexis/Desktop/RMD/Cosmolab/Cosmoclima/datos/Excel Estaciones"
DB = "/Users/alexis/Desktop/RMD/Cosmolab/Cosmoclima/investigacion/fuentes/pluviosidad_diaria_consolidada.sqlite"
CSV_OUT = "/Users/alexis/Desktop/RMD/Cosmolab/Cosmoclima/investigacion/fuentes/pluviosidad_diaria_estaciones_atacama_excel_2019_2026.csv"

# hoja de metadata por archivo -- nombres distintos entre archivos (ver
# revisión manual del 09-ago-2026)
HOJA_METADATA = {
    "ALTO.DEL.CARMEN-ATACAMA.xlsx": "ANTECEDENTE",
    "ALTO.DEL.CARMEN_LOS SAUCES-ATACAMA-SOLO2026.xlsx": "ANTECEDENTE",
    "AMOLANA-ATACAMA.xlsx": "EST.AMOLANA-ATACAMA",
    "CE_HUASCO-ATACAMA.xlsx": "CE-HUASCO-ATACAMA",
    "COPIAPO-ATACAMA.xlsx": "ANTECEDENTE",
    "COPIAPO.BODEGA-ATACAMA.xlsx": "ANTECEDENTE",
    "COPIAPO.UNIVERSIDAD.ATACAMA-ATACAMA.xlsx": "ANTECEDENTE",
    "DESIERTO.ATACAMA-CALDERA-ATACAMA.xlsx": "DETALLE ESTACION",
    "FALDA.VERDE-ATACAMA.xlsx": "ANTECEDENTE",
    "FREIRINA-ATACAMA.xlsx": "ANTECEDENTE",
    "FREIRINA.NICOLASA-ATACAMA-HASTA2024.xlsx": "ANTECEDENTES",
    "FREIRINA.VALLENAR-ATACAMA.xlsx": "ANTECEDENTE",
    "LA.COPA-ATACAMA.xlsx": "ANTECEDENTE",
    "TIERR.AMARI_HORNITOS-ATACAMA.xlsx": "ANTECEDENTES",
    "TIERR.AMARI_IGLES.COLORADA-ATACAMA.xlsx": "ANTECEDENTE",
    "TIERR.AMARI_JOTABECHE-ATACAMA.xlsx": "ANTECEDENTE",
    "TIERR.AMARI_TRANQUE.LAUTARO-ATACAMA.xlsx": "ANTECEDENTE",
    "VALLENAR.HUASCO-ATACAMA-SOLO2026.xlsx": "ANTECEDENTE",
}

MESES_COLUMNA = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
                  "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# único typo real detectado a mano (09-ago-2026): hoja "20219" en
# ALTO.DEL.CARMEN-ATACAMA.xlsx -- mismo layout de grilla que "2019" en
# todos los demás archivos, un dígito de más pegado al final.
TYPOS_HOJA_ANIO = {
    ("ALTO.DEL.CARMEN-ATACAMA.xlsx", "20219"): 2019,
}


def leer_metadata(wb, hoja):
    ws = wb[hoja]
    for fila in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if fila[0] is not None and isinstance(fila[0], (int, float)):
            codigo, nombre, propietario, lat_raw, lon_raw, altura = fila[0], fila[1], fila[2], fila[3], fila[4], fila[5]
            lat = float(lat_raw) / 100000 if abs(lat_raw) > 1000 else float(lat_raw)
            lon = float(lon_raw) / 100000 if abs(lon_raw) > 1000 else float(lon_raw)
            return int(codigo), nombre, propietario, lat, lon, float(altura)
    raise SystemExit(f"No se encontró fila de metadata en hoja {hoja!r}")


def clasificar_hoja(archivo, nombre_hoja, hoja_metadata):
    """Devuelve el año que representa la hoja si es una grilla Día×Mes
    utilizable, o None si hay que saltarla (metadata, o reporte mensual
    redundante de un solo mes como JULIO2026 en el archivo de Caldera)."""
    if nombre_hoja == hoja_metadata:
        return None
    clave = (archivo, nombre_hoja)
    if clave in TYPOS_HOJA_ANIO:
        return TYPOS_HOJA_ANIO[clave]
    if re.fullmatch(r"(19|20)\d{2}", nombre_hoja):
        return int(nombre_hoja)
    m = re.fullmatch(r"AÑO(\d{4})", nombre_hoja)
    if m:
        return int(m.group(1))
    m = re.fullmatch(r"[A-ZÁÉÍÓÚÑ]+-(\d{4})", nombre_hoja)
    if m:
        return int(m.group(1))
    # cualquier otra cosa (ej. JULIO2026/JUNIO2026/MAYO2026, sin guión,
    # layout distinto de horas UTC) -- reporte redundante, se salta
    return None


def parsear_grilla_anio(ws, anio):
    """Grilla Día×Mes (13 columnas: Dia + 12 meses). Verifica el header de
    meses antes de confiar en las columnas, para no leer mal si algún
    archivo trae un orden distinto."""
    filas = list(ws.iter_rows(values_only=True))
    header_meses = filas[1][1:13]
    if list(header_meses) != MESES_COLUMNA:
        raise SystemExit(f"Header de meses inesperado en hoja del año {anio}: {header_meses}")

    dias = []
    for fila in filas[2:]:
        dia = fila[0]
        if not isinstance(dia, int) or not (1 <= dia <= 31):
            continue
        for mes_idx, valor in enumerate(fila[1:13], start=1):
            if valor is None or valor == "." or valor == "":
                continue
            if valor == "s/p":
                mm = 0.0
            else:
                try:
                    mm = float(str(valor).replace(",", "."))
                except ValueError:
                    continue
            try:
                fecha = f"{anio:04d}-{mes_idx:02d}-{dia:02d}"
                # validar fecha real (descarta ej. 31 de febrero si algún
                # archivo lo trajera relleno por error de plantilla)
                from datetime import date
                date(anio, mes_idx, dia)
            except ValueError:
                continue
            dias.append((fecha, mm))
    return dias


def main():
    filas_totales = []  # (fecha, localidad, lat, lon, mm, fuente, codigo)
    metadatas = []  # (archivo, codigo, nombre, propietario, lat, lon, altura, n_dias, fecha_min, fecha_max)

    for archivo, hoja_meta in sorted(HOJA_METADATA.items()):
        ruta = os.path.join(CARPETA_EXCEL, archivo)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        codigo, nombre, propietario, lat, lon, altura = leer_metadata(wb, hoja_meta)
        localidad = f"{nombre} (DMC/INIA {codigo})"
        fuente = f"Excel estación real Atacama (Alexis, 09-ago-2026), código {codigo}, propietario {propietario}"

        dias_estacion = []
        for nombre_hoja in wb.sheetnames:
            anio = clasificar_hoja(archivo, nombre_hoja, hoja_meta)
            if anio is None:
                continue
            dias_estacion.extend(parsear_grilla_anio(wb[nombre_hoja], anio))

        if not dias_estacion:
            print(f"AVISO: {archivo} ({localidad}) no aportó ningún día real -- revisar a mano.")
            continue

        for fecha, mm in dias_estacion:
            filas_totales.append((fecha, localidad, lat, lon, mm, fuente, codigo))

        fechas = sorted(f for f, _ in dias_estacion)
        metadatas.append((archivo, codigo, nombre, propietario, lat, lon, altura,
                           len(dias_estacion), fechas[0], fechas[-1]))
        print(f"{localidad:55s} {len(dias_estacion):5d} días reales, {fechas[0]} a {fechas[-1]}, altura={altura}m")

    # CSV de trazabilidad
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["fecha", "localidad", "lat", "lon", "lluvia_mm", "fuente"])
        w.writerows((f, l, la, lo, mm, fu) for (f, l, la, lo, mm, fu, _cod) in filas_totales)
    print(f"\nCSV guardado: {CSV_OUT} ({len(filas_totales)} filas)")

    # inserción idempotente: por localidad, borrar el rango de fechas que
    # este script está a punto de insertar, y reinsertar -- mismo patrón
    # que agregar_estaciones_julio2026_a_consolidada.py
    con = sqlite3.connect(DB)
    por_localidad = {}
    for fila in filas_totales:
        por_localidad.setdefault(fila[1], []).append(fila)

    for localidad, filas in por_localidad.items():
        fecha_min = min(f[0] for f in filas)
        fecha_max = max(f[0] for f in filas)
        con.execute(
            "DELETE FROM pluviosidad_diaria WHERE localidad = ? AND fecha BETWEEN ? AND ?",
            (localidad, fecha_min, fecha_max),
        )
    con.executemany(
        "INSERT INTO pluviosidad_diaria (fecha, localidad, lat, lon, lluvia_mm, fuente, tipo_fuente) VALUES (?,?,?,?,?,?,?)",
        [(f, l, la, lo, mm, fu, "estacion_real") for (f, l, la, lo, mm, fu, _cod) in filas_totales],
    )
    con.commit()
    total = con.execute("SELECT COUNT(*) FROM pluviosidad_diaria").fetchone()[0]
    print(f"Total filas en la tabla consolidada tras sumar Excel Atacama: {total}")
    con.close()

    print("\n=== metadata por estación (para catálogo de elevación) ===")
    for m in metadatas:
        print(m)

    print("""
=== SOLAPES CONOCIDOS, no fusionados a propósito ===
Las siguientes estaciones nuevas coinciden geográficamente (<2km) con un
localidad YA EXISTENTE de registro histórico largo, bajo OTRO nombre. Se
insertaron como series separadas (mismo criterio que ya usaba
agregar_estaciones_julio2026_a_consolidada.py) -- no se fusionaron los
nombres porque mezclar series de fuente distinta sin revisar valor a valor
no fue pedido. Quedan dos "voces" reales para la misma estación física en
2019-2025 (más historia una, más reciente + altura real la otra):
  - Desierto de Atacama, Caldera (DMC) [2005-2025] vs
    Desierto de Atacama, Caldera Ad. (DMC/INIA 270008) [2019-2026, este script]
  - Copiapo [1971-2018] vs Copiapo (DMC/INIA 270016) [2019-2026, este script]
  - Copiapó Universidad de Atacama (DMC/INIA 270009) ya existía con solo
    julio-2026 (31 días) -- este script la extiende a 2019-2026 completo.
""")


if __name__ == "__main__":
    main()
