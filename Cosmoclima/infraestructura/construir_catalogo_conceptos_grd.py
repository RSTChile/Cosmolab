"""
Convierte el «Glosario Gestión del Riesgo de Desastres» (ONEMI, 2021) en un
catálogo de conceptos del proyecto, y lo alinea con el RMD 2.0.

POR QUÉ
-------
El proyecto tiene que hablar el idioma del sistema chileno de emergencias, no
uno propio. El glosario es el vocabulario oficial: si la matriz dice
«interrupción de servicio» donde el país dice «aislamiento», el instrumento no
le sirve a quien lo tiene que usar.

NOTA INSTITUCIONAL (importante, no borrar)
------------------------------------------
El glosario lo publicó ONEMI (Oficina Nacional de Emergencia) en mayo de 2021.
ONEMI fue reemplazada por SENAPRED (Servicio Nacional de Prevención y Respuesta
ante Desastres). Los conceptos siguen vigentes; el organismo emisor ya no
existe con ese nombre. Cada ficha del catálogo lo deja anotado para que nadie
cite a ONEMI como organismo vigente.

QUÉ HACE
--------
1. Extrae el texto del PDF respetando la estructura de secciones (terminología
   general + las nueve variables de riesgo).
2. Separa cada entrada «Término: definición».
3. Marca la relevancia de cada concepto para este proyecto y, donde existe, su
   equivalente en el RMD 2.0.

No modifica el PDF ni ningún archivo del RMD: sólo lee y escribe su propio Excel.
"""

import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

AQUI = Path(__file__).parent
PDF = AQUI / "fuentes" / "Glosario_GRD_Onemi.pdf"
SALIDA = AQUI / "Catalogo_Conceptos_GRD_Proyecto.xlsx"

# Secciones del glosario, por página de inicio (del índice del propio documento)
SECCIONES = [
    (8, "1 · Terminología general de GRD"),
    (35, "2.1 · Variable de riesgo — Sísmico"),
    (42, "2.2 · Variable de riesgo — Tsunami"),
    (48, "2.3 · Variable de riesgo — Volcánico"),
    (55, "2.4 · Variable de riesgo — Remoción en masa"),
    (59, "2.5 · Variable de riesgo — Meteorológico"),
    (68, "2.6 · Variable de riesgo — Incendio forestal"),
    (71, "2.7 · Variable de riesgo — Materiales peligrosos"),
    (76, "2.8 · Variable de riesgo — Biológico"),
    (78, "2.9 · Variable de riesgo — Depósitos o tranques de relaves"),
    (82, "3 · Anexos"),
]

# Ruido de encabezado/pie que se repite en todas las páginas
RUIDO = re.compile(
    r"^\s*(\d+\s*)?(GLOSARIO|GESTIÓN DEL\s*$|GESTIÓN DEL RIESGO DE DESASTRES|ONEMI|"
    r"RIESGO DE DESASTRES|CONTENIDO)\s*$", re.I)

# ─────────────────────────────────────────────────────────────────────────────
# ALINEACIÓN CON EL RMD. Sólo se mapea lo que de verdad mapea; el resto queda
# catalogado sin forzar equivalencia. Clave = término normalizado (sin tildes,
# minúsculas). Valor = (relevancia, equivalente RMD, nota de alineación).
# ─────────────────────────────────────────────────────────────────────────────
ALINEACION = {
    "aislamiento": ("★ Crítica", "NGF-L (MCSGS) · IDSE_X #318",
        "★★ EL HALLAZGO. El glosario oficial define aislamiento como «aquella "
        "condición en que el acceso normal se encuentra interrumpido y NO SE "
        "CUENTA TAMPOCO CON UN ACCESO ALTERNATIVO». Eso es, palabra por "
        "palabra, la definición de nodo de flujo del MCSGS: paso obligado sin "
        "alternativa. Chile ya tenía el concepto en su vocabulario oficial de "
        "emergencias; el RMD lo tenía en su teoría. Nadie los había juntado."),
    "afectados": ("★ Crítica", "P_Afect en IDCE #250, TD #242, IVC #241",
        "La definición oficial cierra diciendo «especialmente aplicable en "
        "casos de cortes de energía eléctrica, teléfono, agua y aislamiento» — "
        "o sea, la unidad de medida humana del país ya está definida por corte "
        "de servicio, que es exactamente lo que mide este proyecto."),
    "amenaza": ("★ Crítica", "InEvExtre #309 · el forzante climático",
        "Es el término oficial para lo que MACLIMA llama forzante. Alinear "
        "vocabulario: en los entregables al COGRID se dice AMENAZA, no "
        "«forzante»."),
    "vulnerabilidad": ("★ Crítica", "FVT · VT (MICR)",
        "Concepto compartido con la Ley 21.542 (sub-criterio de criticidad) y "
        "con la MICR. Verificar que las tres definiciones coincidan antes de "
        "usarlas como si fueran la misma."),
    "exposicion": ("★ Crítica", "FEX en InEvExtre #309",
        "El canon de MACLIMA ya usa un Factor de Exposición (FEX 0,7-1,5) y "
        "dice que se ajusta por «criticidad MICR local». Es el mismo concepto."),
    "riesgo": ("★ Crítica", "PF · Pen (MICR)",
        "El país define riesgo como función de amenaza, exposición y "
        "vulnerabilidad. La MICR calcula PF = IB × FVT, que NO incluye "
        "amenaza. Ahí está, dicho en vocabulario oficial, el hueco que este "
        "proyecto viene a llenar: falta la amenaza, y la amenaza es el clima."),
    "resiliencia": ("★ Crítica", "FRC (MCSGS) · IRL #210",
        "Exigido por la Ley 21.542 como sub-criterio de criticidad y ausente "
        "de la MICR (hallazgo H-12)."),
    "infraestructura critica": ("★ Crítica", "MICR completa",
        "Contrastar con la definición del art. 32 N°21 de la Constitución "
        "(Ley 21.542): si el glosario y la ley difieren, manda la ley."),
    "remocion en masa": ("★ Crítica", "EAL (subíndice de InEvExtre #309)",
        "★ El nombre oficial chileno de lo que MACLIMA llama EAL "
        "(aluvión/inundación). Es EL evento que corta caminos. El glosario le "
        "dedica una sección entera (2.4), con tipología propia."),
    "sistema de alerta temprana": ("★ Crítica", "Destino del instrumento",
        "Es el producto que el COGRID espera. Un FEN dinámico es, en la "
        "práctica, un insumo de alerta temprana."),
    "centro de alerta temprana": ("★ Crítica", "Receptor de la salida",
        "El CAT Nacional es quien opera el SAE. La matriz le entrega insumo a "
        "ÉL; no emite nada por su cuenta."),
    "sistema de alerta de emergencia-sae": ("★ Crítica", "INTERFAZ DE SALIDA del proyecto",
        "★★ REQUISITO DE ARQUITECTURA (instrucción del 15-ago-2026). El SAE "
        "envía aviso GEORREFERENCIADO por Plataforma IADC desde el Centro de "
        "Alerta Temprana Nacional, con tecnología CBS (cell broadcast) a los "
        "celulares de la zona. La matriz debe poder ENTREGAR DATOS a ese "
        "sistema como insumo para que la autoridad decida si emite alerta. "
        "LÍMITE DURO: la matriz NO emite alertas ni las dispara "
        "automáticamente — produce insumo, la decisión y la emisión son de la "
        "autoridad. Que el aviso sea georreferenciado es lo que obliga a que "
        "nuestra salida también lo sea, hasta nivel comunal."),
    "alerta verde": ("★ Crítica", "Tramo bajo del ICSGS",
        "Primer peldaño de la escala oficial: vigilancia permanente. El "
        "instrumento debe expresarse en ESTA escala, no inventar una propia. "
        "Incluye la «Alerta Temprana Preventiva», que es exactamente el "
        "producto de una matriz predictiva."),
    "alerta amarilla": ("★ Crítica", "Tramo medio del ICSGS",
        "«Cuando una amenaza crece en extensión y severidad y se evalúa que no "
        "podrá ser controlada con los recursos locales habituales». Nótese: el "
        "criterio oficial es capacidad de respuesta local superada — o sea, "
        "resiliencia (FRC del MCSGS), no magnitud del daño."),
    "alerta roja": ("★ Crítica", "Tramo alto del ICSGS",
        "Movilización de todos los recursos. Mapear contra los tramos del "
        "ICSGS para que ambos hablen el mismo idioma."),
    "amplitud de una alerta": ("Alta", "US (Unidad de Sistema) del MCSGS",
        "La «amplitud» oficial es el mismo concepto que la US del MCSGS: hasta "
        "dónde alcanza. Confirma que declarar el nivel es práctica del país."),
    "emergencia": ("Alta", "Fase de emergencia (Ley 21.542)",
        "Delimita cuándo el instrumento pasa de planificación a respuesta."),
    "desastre": ("Alta", "ICSGS 46-80% (MCSGS)",
        "Contrastar la escala oficial de desastre con los tramos del ICSGS "
        "para que los umbrales hablen el mismo idioma."),
    "catastrofe": ("Alta", "ICSGS 81-100% (MCSGS)",
        "Idem: alinear el techo del ICSGS con la definición oficial."),
    "gestion del riesgo de desastres": ("Alta", "Marco general del proyecto",
        "Es el marco al que el instrumento debe servir."),
    "sistema nacional de proteccion civil": ("Alta", "SINAPRED · SENAPRED",
        "Reemplazado por el SINAPRED. Verificar vigencia del término antes de "
        "usarlo en un entregable."),
    "damnificados": ("Alta", "P_Afect · TD #242",
        "Categoría más severa que «afectados»; el país las distingue y el "
        "instrumento debería también."),
    "capacidad de respuesta": ("Alta", "FRC (MCSGS)", "Alimenta la resiliencia."),
    "mitigacion": ("Media", "Fase de normalidad (Ley 21.542)",
        "La fase donde Castillo y Saldaña sitúan la falla del estatuto."),
    "prevencion": ("Media", "Fase de normalidad (Ley 21.542)", "Idem."),
    "preparacion": ("Media", "Fase de normalidad (Ley 21.542)", "Idem."),
    "amenaza hidrometeorologica": ("★ Crítica", "ANPrecip #307 · InEvExtre #309",
        "La familia de amenaza que este proyecto modela."),
    "aluvion": ("★ Crítica", "EAL (subíndice de InEvExtre #309)",
        "El evento del piloto: Copiapó, 24-25 de marzo de 2015."),
    "inundacion": ("★ Crítica", "EOP/EAL (subíndices de InEvExtre #309)",
        "Modo de falla directo de una subestación."),
    "sequia": ("Alta", "ANPrecip #307 (lado déficit) · EstHidric #308",
        "★ Es la razón por la que ANPrecip NO puede usar valor absoluto "
        "(hallazgo H-07): el país distingue sequía de inundación, y con "
        "|anomalía| las dos dan el mismo número."),
    "microzonificacion": ("Alta", "Nivel US-Comunal",
        "El país ya tiene metodología para bajar a nivel local (AIDEP). El "
        "instrumento debería producir a esa escala."),
}


def normalizar(texto):
    """Minúsculas y sin tildes, para poder cruzar términos sin pelear con la
    ortografía."""
    t = unicodedata.normalize("NFD", texto.lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").strip()


def limpiar_pagina(texto):
    """Saca encabezados repetidos y notas al pie, y rearma las palabras que el
    PDF partió con guion al final de línea."""
    lineas = []
    for linea in texto.split("\n"):
        if RUIDO.match(linea):
            continue
        # nota al pie: empieza con número y sigue con una fuente citada
        if re.match(r"^\s*\d{1,3}\s+[A-ZÁÉÍÓÚÑ(]", linea) and (
                "http" in linea or re.search(r"\(\d{4}\)|Ibid|Ley N°", linea)):
            continue
        if re.match(r"^\s*\d{1,3}\s*$", linea):      # número de página suelto
            continue
        lineas.append(linea)
    t = "\n".join(lineas)
    t = re.sub(r"(\w)\s*-\s*\n\s*(\w)", r"\1\2", t)   # palabra partida por guion
    t = re.sub(r"\n(?![A-ZÁÉÍÓÚÑ])", " ", t)          # une líneas de un párrafo
    t = re.sub(r"[ \t]{2,}", " ", t)
    return t


def seccion_de(pagina):
    actual = SECCIONES[0][1]
    for inicio, nombre in SECCIONES:
        if pagina >= inicio:
            actual = nombre
    return actual


def extraer_entradas():
    """Devuelve [(término, definición, sección, página)] del glosario."""
    lector = PdfReader(PDF)
    entradas = []
    # Un término es una línea que arranca en mayúscula y trae ':' temprano.
    patron = re.compile(
        r"(?:^|\n)\s*([A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúñüÜ /()\-,\.]{2,70}?)\s*:\s+"
        r"(?=[A-ZÁÉÍÓÚÑa-z¿«])")
    for n_pag, pagina in enumerate(lector.pages, start=1):
        if n_pag < 8 or n_pag > 86:      # antes: portada e índice; después: referencias
            continue
        texto = limpiar_pagina(pagina.extract_text() or "")
        cortes = list(patron.finditer(texto))
        for i, m in enumerate(cortes):
            fin = cortes[i + 1].start() if i + 1 < len(cortes) else len(texto)
            termino = re.sub(r"\s+", " ", m.group(1)).strip(" .,")
            definicion = re.sub(r"\s+", " ", texto[m.end():fin]).strip()
            definicion = re.sub(r"\s*\d{1,3}\s*$", "", definicion)  # nº de nota al pie
            if len(termino) < 3 or len(definicion) < 25:
                continue
            if termino.isupper() and len(termino.split()) > 4:      # títulos de sección
                continue
            entradas.append((termino, definicion, seccion_de(n_pag), n_pag))
    return entradas


def main():
    entradas = extraer_entradas()
    # dedupe por término normalizado, quedándome con la definición más larga
    mejor = {}
    for t, d, s, p in entradas:
        k = normalizar(t)
        if k not in mejor or len(d) > len(mejor[k][1]):
            mejor[k] = (t, d, s, p)
    filas = sorted(mejor.values(), key=lambda x: (x[2], normalizar(x[0])))

    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo de conceptos"
    enc = ["Término", "Definición oficial (ONEMI 2021)", "Sección del glosario",
           "Pág.", "Relevancia para el proyecto", "Equivalente en el RMD 2.0",
           "Nota de alineación"]
    ws.append(enc)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"

    n_alineados = 0
    for t, d, s, p in filas:
        rel, equiv, nota = ALINEACION.get(normalizar(t), ("", "", ""))
        if rel:
            n_alineados += 1
        ws.append([t, d, s, p, rel, equiv, nota])

    for i, w in enumerate([38, 96, 34, 7, 18, 40, 86], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    destacado = PatternFill("solid", fgColor="FFF2CC")
    for fila in ws.iter_rows(min_row=2):
        for c in fila:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = borde
        if str(fila[4].value).startswith("★"):
            fila[0].fill = destacado
            fila[4].font = Font(bold=True, color="C00000")

    # ── hoja de contexto ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("LÉEME")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 118
    for k, v in [
        ("CATÁLOGO DE CONCEPTOS · GESTIÓN DEL RIESGO DE DESASTRES", ""),
        ("", ""),
        ("Qué es",
         "El vocabulario oficial chileno de gestión del riesgo de desastres, "
         "convertido en catálogo y alineado con el RMD 2.0. El proyecto tiene "
         "que hablar el idioma del sistema que lo va a usar, no uno propio."),
        ("Fuente",
         "«Glosario – Gestión del Riesgo de Desastres», ONEMI, 1ª edición, mayo "
         "de 2021. División de Protección Civil y Academia de Protección Civil, "
         "Subdirección de Gestión del Riesgo. 90 páginas."),
        ("ATENCIÓN institucional",
         "El glosario lo publicó ONEMI (Oficina Nacional de Emergencia), que YA "
         "NO EXISTE: fue reemplazada por SENAPRED (Servicio Nacional de "
         "Prevención y Respuesta ante Desastres). Los conceptos siguen "
         "vigentes, el organismo emisor no. En cualquier entregable hay que "
         "citar el glosario como fuente histórica de ONEMI, nunca presentar a "
         "ONEMI como organismo actual. Ojo también con los términos que "
         "nombran estructuras derogadas (p. ej. «Sistema Nacional de "
         "Protección Civil», hoy SINAPRED)."),
        ("Cómo leerlo",
         "Las dos primeras columnas son textuales del glosario y no se tocan. "
         "Las tres últimas son del proyecto: qué tan relevante es el concepto, "
         "con qué elemento del RMD se corresponde, y qué hay que cuidar al "
         "alinearlos. Sólo se mapeó lo que de verdad mapea; el resto queda "
         "catalogado sin forzar equivalencias."),
        ("★ El hallazgo del glosario",
         "«AISLAMIENTO: aquella condición en que el acceso normal (terrestre, "
         "marítimo o aéreo) se encuentra interrumpido y NO SE CUENTA TAMPOCO "
         "CON UN ACCESO ALTERNATIVO». Eso es, palabra por palabra, la "
         "definición de nodo de flujo del MCSGS: punto de paso obligado sin "
         "alternativa de corto plazo. Chile ya tenía el concepto en su "
         "vocabulario oficial de emergencias desde 2021; el RMD lo tenía en su "
         "teoría. Nadie los había juntado. Y «AFECTADOS» se define como las "
         "personas que ven perturbado su quehacer «especialmente en casos de "
         "cortes de energía eléctrica, teléfono, agua y aislamiento»: la unidad "
         "de medida humana del país ya está definida por corte de servicio."),
        ("★ Y una confirmación",
         "El país define RIESGO en función de amenaza, exposición y "
         "vulnerabilidad. La MICR calcula PF = IB × FVT, donde no entra la "
         "amenaza por ninguna parte. Dicho en vocabulario oficial: a la matriz "
         "le falta el término «amenaza» de la ecuación de riesgo — y la "
         "amenaza, acá, es el clima. Es el mismo agujero del FEN estático, "
         "visto desde la doctrina chilena en vez de desde el dato."),
        ("Generado",
         "15-ago-2026 · script «construir_catalogo_conceptos_grd.py», en esta "
         "misma carpeta. Reproducible: se vuelve a correr y reconstruye el "
         "archivo desde el PDF, sin tocarlo."),
    ]:
        ws2.append([k, v])
    ws2["A1"].font = Font(bold=True, size=15, color="1F3864")
    for fila in ws2.iter_rows(min_row=3):
        fila[0].font = Font(bold=True, color="1F3864")
        for c in fila:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.move_sheet("LÉEME", offset=-1)
    wb.save(SALIDA)

    from collections import Counter
    print(f"Escrito: {SALIDA}")
    print(f"  · {len(filas)} conceptos extraídos del glosario")
    print(f"  · {n_alineados} alineados con el RMD "
          f"({len(ALINEACION)-n_alineados} del mapa no aparecieron — revisar)")
    for s, n in Counter(f[2] for f in filas).most_common():
        print(f"      {n:4d}  {s}")


if __name__ == "__main__":
    main()
