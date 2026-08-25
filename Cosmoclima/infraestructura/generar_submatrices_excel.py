"""
GENERAR LAS SUB-MATRICES COMO ARCHIVOS EXCEL, LISTOS PARA SUBIR A SHAREPOINT
=============================================================================

INSTRUCCIÓN (Alexis, 21-ago-2026)
-----------------------------------
«Consolida las nuevas matrices como archivos excel, y yo los subo directamente.
Nombra los archivos con el número de cada elemento de la Matriz MICR seguido de
su descriptor, "XXX Hospitales" por ejemplo.»

Antes se había decidido: **28 sub-matrices, sólo la fuente autoritativa** de cada
familia — se dejan fuera los respaldos que duplican el mismo activo desde otro
organismo, y las familias que no tienen fila en la Matriz.

EL ESTILO QUE SE REPLICA
------------------------
Leído de las dos listas que ya existen en SharePoint —`120` (Subestaciones) y
`centrales`—, que comparten esquema:

    Título · Número · Ítem · Sector · Elemento · Región · Provincia ·
    Dirección · Latitud · Longitud · Responsable · Teléfono · MICR

`MICR` es una **columna de búsqueda** («Vínculo a la Matriz de Infraestructura
Crítica») que apunta a la columna Título de la lista `mic`. Una columna de
búsqueda **no se puede crear importando un Excel**: SharePoint infiere columnas
de texto y número, nunca vínculos. Por eso el Excel trae todo menos esa columna,
y hay dos formas de resolverlo al subir — están explicadas en el informe.

DOS DESVIACIONES DEL ESTILO, DECLARADAS
-----------------------------------------
1. **Se agrega la columna `Comuna`**, que el esquema de SharePoint no tiene. La
   comuna es el nivel al que trabaja el Comité para la Gestión del Riesgo de
   Desastres (COGRID) y está en el marco normativo del proyecto; perderla al
   subir sería perder el nivel administrativo más operativo. Si se prefiere
   fidelidad estricta, se borra la columna y ya.
2. **Se agregan `Latitud decimal` y `Longitud decimal`** al final. Las listas
   existentes guardan sólo grados-minutos-segundos, que es formato de lectura;
   todo el cálculo del proyecto se hace en decimal. Conservar las dos evita que
   la conversión de ida y vuelta introduzca error en el registro oficial.

QUÉ HAY EN CADA COLUMNA
-----------------------
  Título      el nombre del activo tal como lo llama el organismo que lo publica
  Número      correlativo dentro del archivo, como en `centrales`
  Ítem        el número de fila de la MICR al que pertenece el activo
  Sector      copiado de la MICR para ese ítem, no escrito a mano
  Elemento    copiado de la MICR para ese ítem, no escrito a mano
  Responsable el operador declarado por la fuente
  Teléfono    vacío: ninguna de las fuentes públicas lo entrega

USO
---
    ../.venv-esa/bin/python generar_submatrices_excel.py
"""

import csv
import sys
from pathlib import Path

AQUI = Path(__file__).resolve().parent
DATOS = AQUI / "datos"
SALIDA = AQUI / "submatrices_excel"
MICR = DATOS / "micr_sharepoint.csv"

COLUMNAS = ["Título", "Número", "Ítem", "Sector", "Elemento", "Región",
            "Provincia", "Comuna", "Dirección", "Latitud", "Longitud",
            "Responsable", "Teléfono", "Latitud decimal", "Longitud decimal"]


def dms(valor, eje):
    """Decimal → grados-minutos-segundos en el formato exacto de la lista 120.

    Ejemplo real de esa lista: -18° 28' 40.80" S · -70° 17' 49.20" W
    Nótese que el formato del RMD lleva el signo Y la letra del hemisferio, que
    es redundante pero es el que está en uso: se replica tal cual.
    """
    if valor in (None, ""):
        return ""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return ""
    hemi = ("S" if v < 0 else "N") if eje == "lat" else ("W" if v < 0 else "E")
    a = abs(v)
    g = int(a)
    m_dec = (a - g) * 60
    m = int(m_dec)
    s = (m_dec - m) * 60
    if round(s, 2) >= 60:                 # arrastre por redondeo
        s = 0.0
        m += 1
    if m >= 60:
        m = 0
        g += 1
    signo = "-" if v < 0 else ""
    return f"{signo}{g}° {m:02d}' {s:05.2f}\" {hemi}"


# ── el mapeo: cada familia con su ítem, su fuente y de dónde sale cada campo ──
# `campos` traduce del CSV del proyecto a las columnas de SharePoint. Lo que no
# tiene equivalente queda vacío: no se inventa.
COMUN = dict(titulo="nombre", responsable="operador", region="region",
             provincia="provincia", comuna="comuna", lat="lat", lon="lon",
             direccion=None)

# ★ Ítem por fila. `centrales` demuestra que una sub-matriz puede cubrir varios
# ítems: cada fila declara el suyo. Donde la fuente permite distinguir, se
# distingue; donde no, todas las filas llevan el ítem principal.
def item_educacion_superior(f):
    """Universidades (443) o Centros de Formación Técnica (446), por tipo."""
    t = (f.get("src_tipo") or "").lower()
    return 443 if "universidad" in t else 446

SUBMATRICES = [
 (616, "Red Vial", "Tramo Vial", "mop_tramos.csv",
  dict(titulo="nombre_camino", responsable="tuicion", region="region",
       provincia=None, comuna=None, lat="lat_punto_medio",
       lon="lon_punto_medio", direccion=("Rol {rol} · km {km_i_metros} a "
                                         "{km_f_metros} · {clasificacion}"))),
 (618, "Puentes de Carreteras", "Puente", "mop_puentes.csv",
  dict(titulo="nombre_puente", responsable="tuicion", region="region",
       provincia="provincia", comuna="comuna", lat="lat", lon="lon",
       direccion="Camino {nombre_camino} · Rol {rol} · cauce {cauce_queb}")),
 (17, "Agua Potable Rural", "Sistema APR", "inventario_agua_potable_rural.csv",
  COMUN),
 (46, "Embalses", "Embalse", "inventario_embalses.csv", COMUN),
 (42, "Depositos de Relaves", "Depósito", "relaves_depositos.csv",
  dict(titulo="instalacion", responsable="empresa", region="region",
       provincia="provincia", comuna="comuna", lat="lat", lon="lon",
       direccion="Faena {faena} · {tipo_deposito} · estado {estado}")),
 (265, "Establecimientos de Salud", "Establecimiento", "inventario_salud.csv",
  COMUN),
 (441, "Establecimientos Educacionales", "Establecimiento",
  "inventario_educacion.csv", COMUN),
 (622, "Obras Portuarias", "Obra", "inventario_obras_portuarias.csv", COMUN),
 (624, "Aeropuertos y Aerodromos", "Aeródromo", "inventario_aeropuertos.csv",
  COMUN),
 (309, "Cuarteles de Bomberos", "Cuartel", "inventario_senapred_bomberos.csv",
  COMUN),
 (353, "Unidades de Carabineros", "Unidad",
  "inventario_senapred_carabineros.csv", COMUN),
 (353, "Unidades de la PDI", "Unidad", "inventario_senapred_pdi.csv", COMUN),
 (183, "Telecomunicaciones", "Instalación",
  "inventario_senapred_telefonia.csv", COMUN),
 (117, "Lineas de Transmision Electrica", "Tramo de línea",
  "inventario_senapred_energia_lineal.csv", COMUN),
 (120, "Instalaciones Electricas", "Instalación",
  "inventario_senapred_energia_puntual.csv", COMUN),
 (117, "Derivaciones de Linea", "Derivación", "inventario_taps_electricos.csv",
  COMUN),
 (33, "Infraestructura Sanitaria", "Instalación",
  "inventario_senapred_siss.csv", COMUN),
 (16, "Suministro Alternativo de Agua", "Punto",
  "inventario_senapred_suministro_alternativo_agua.csv", COMUN),
 (575, "Oficinas Municipales", "Municipio",
  "inventario_senapred_municipios.csv", COMUN),
 (572, "Sedes de Gobierno Provincial", "Sede",
  "inventario_senapred_gobernaciones.csv", COMUN),
 (572, "Sedes de Gobierno Regional", "Sede",
  "inventario_senapred_intendencias.csv", COMUN),
 (572, "Edificios Publicos", "Edificio",
  "inventario_senapred_edificios_publicos.csv", COMUN),
 (529, "Recintos Deportivos", "Recinto",
  "inventario_senapred_recintos_deportivos.csv", COMUN),
 # ★ Se usa `centros_publicos` y NO `supermercados`: comprobado que el segundo
 # está contenido ÍNTEGRAMENTE en el primero (cero filas exclusivas), y el
 # primero trae además 53 malls. Mismo catastro, más completo.
 (528, "Centros Comerciales", "Local",
  "inventario_senapred_centros_publicos.csv", COMUN),
 # ★ NO era huérfana: los ítems 443 y 446 ya existen. Se reparte por `src_tipo`.
 (443, "Educacion Superior", "Sede", "inventario_senapred_sedes_universitarias.csv",
  dict(COMUN, item_por_fila=item_educacion_superior)),
 # ★ Las dos que SÍ necesitan fila nueva en la Matriz (ver PROPUESTA_MICR).
 (836, "Establecimientos para Adultos Mayores", "Establecimiento",
  "inventario_senapred_senama.csv", COMUN),
 (837, "Establecimientos de Proteccion de la Infancia", "Establecimiento",
  "inventario_senapred_mejor_ninez.csv", COMUN),
 (355, "Pasos Fronterizos", "Paso",
  "inventario_senapred_pasos_fronterizos.csv", COMUN),
 (351, "Direcciones Regionales SENAPRED", "Dirección regional",
  "inventario_senapred_direcciones_regionales.csv", COMUN),
 (639, "Comunicacion Aerea", "Instalación",
  "inventario_senapred_comunicacion_aerea.csv", COMUN),
 (265, "Servicio Medico Legal", "Sede",
  "inventario_senapred_servicio_medico_legal.csv", COMUN),
]


def valor(fila, campo):
    """Lee un campo, o arma un texto con plantilla, o devuelve vacío."""
    if campo is None:
        return ""
    if "{" in campo:                       # plantilla tipo "Rol {rol} · km {km}"
        try:
            return campo.format(**{k: (v or "") for k, v in fila.items()}).strip()
        except KeyError:
            return ""
    return (fila.get(campo) or "").strip()


def main():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    micr = {int(x["n"]): (x["Sector"], x["elemento"], x["Pen"])
            for x in csv.DictReader(MICR.open(encoding="utf-8"))}
    # Las dos filas propuestas todavía no están en la Matriz; se declaran acá
    # para que las sub-matrices puedan generarse ya. Ver PROPUESTA_MICR_836_837.md
    micr.setdefault(836, ("Protección Social",
                          "Establecimientos para Adultos Mayores "
                          "(ELEAM y Centros Diurnos)", "Alta"))
    micr.setdefault(837, ("Protección Social",
                          "Establecimientos de Protección de la Infancia "
                          "(Residencias y Programas Ambulatorios)", "Alta"))
    SALIDA.mkdir(exist_ok=True)

    print("=" * 96)
    print("SUB-MATRICES → EXCEL, con el esquema de las listas 120 y centrales")
    print("=" * 96)
    print(f"\n  destino: {SALIDA}\n")
    print(f"    {'archivo':52s} {'filas':>7s} {'con coord':>10s}  ítem")
    print("    " + "-" * 88)

    total = con_coord = 0
    resumen = []
    for item, descriptor, titulo_col, fuente, campos in SUBMATRICES:
        ruta = DATOS / fuente
        if not ruta.exists():
            print(f"    ✗ falta {fuente}")
            continue
        filas = list(csv.DictReader(ruta.open(encoding="utf-8")))
        sector, elemento, pen = micr[item]
        por_fila = campos.get("item_por_fila")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{item} {descriptor}"[:31]
        encabezado = list(COLUMNAS)
        encabezado[0] = titulo_col          # como en `120`, cuyo Título se llama «Subestación»
        ws.append(encabezado)

        n_coord = 0
        for i, f in enumerate(filas, start=1):
            lat = valor(f, campos["lat"])
            lon = valor(f, campos["lon"])
            if lat and lon:
                n_coord += 1
            it = por_fila(f) if por_fila else item
            sec_f, ele_f, _ = micr[it]
            ws.append([
                valor(f, campos["titulo"]), i, it, sec_f, ele_f,
                valor(f, campos["region"]), valor(f, campos["provincia"]),
                valor(f, campos["comuna"]), valor(f, campos["direccion"]),
                dms(lat, "lat"), dms(lon, "lon"),
                valor(f, campos["responsable"]), "",
                float(lat) if lat else "", float(lon) if lon else "",
            ])

        for c in range(1, len(encabezado) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        anchos = [42, 8, 7, 16, 34, 24, 20, 20, 46, 22, 22, 28, 12, 16, 16]
        for c, a in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(c)].width = a
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezado))}{len(filas)+1}"

        nombre = f"{item} {descriptor}.xlsx"
        wb.save(SALIDA / nombre)
        total += len(filas)
        con_coord += n_coord
        resumen.append((nombre, len(filas), n_coord, elemento, pen))
        print(f"    {nombre:52s} {len(filas):7,d} {n_coord:10,d}  {pen}")

    print("    " + "-" * 88)
    print(f"    {'TOTAL':52s} {total:7,d} {con_coord:10,d}  {len(resumen)} archivos")

    # índice, para que se sepa qué se subió y contra qué ítem
    with (SALIDA / "00 INDICE.csv").open("w", newline="", encoding="utf8") as fh:
        w = csv.writer(fh)
        w.writerow(["archivo", "filas", "con_coordenada", "elemento_micr", "Pen"])
        w.writerows(resumen)
    print(f"\n  índice: {SALIDA / '00 INDICE.csv'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
