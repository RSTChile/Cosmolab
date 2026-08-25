"""
LECTOR DE FICHAS DE VARIABLES DEL WORD OFICIAL DEL RMD 2.0
===========================================================

QUÉ HACE
--------
Recorre el Word canónico de Variables y Métricas y extrae UNA FICHA por cada
variable definida, con su número, su sigla y su nombre. Después cruza el
resultado contra el Excel oficial del modelo y reporta:

  1. las SIGLAS DUPLICADAS del Word (dos variables distintas con la misma sigla),
  2. las variables que están en el Word y NO en el Excel,
  3. las que están en el Excel y NO en el Word.

POR QUÉ EXISTE
--------------
Regla de trabajo fijada por Alexis el 16-ago-2026: **manda el Word**. El Excel es
el traspaso operativo. Cuando hay siglas duplicadas hay que dilucidar cuál
fórmula aplica al modelo y desambiguar agregando una letra a la sigla. Para poder
hacer eso primero hay que SABER cuáles están duplicadas — que es lo que este
lector responde.

CÓMO ENCUENTRA LAS FICHAS (y por qué así)
-----------------------------------------
Los encabezados del Word vienen en al menos siete formatos distintos
("29 - Nombre - SIGLA", "10.- SIGLA: Nombre", "31. SIGLA — Nombre",
"7. Nombre (SIGLA)", y algunos sin sigla). Buscar por formato de encabezado
pierde fichas en silencio.

En cambio, TODAS las fichas tienen una línea "Fórmula Numérica/Algebraica:".
Esa línea se usa como ANCLA: el encabezado es el párrafo con texto que está
inmediatamente antes. Es un anclaje por contenido y no por formato, así que no
depende de cómo se tituló cada bloque.

Para las fichas cuyo encabezado no trae sigla (un bloque entero de variables
electorales las omite), la sigla se recupera cruzando el NOMBRE contra el Excel
oficial. Si tampoco así aparece, la ficha se reporta como «sigla sin resolver» —
nunca se inventa.

USO
---
    python leer_fichas_word.py              # informe a pantalla
    python leer_fichas_word.py --csv        # además escribe datos/fichas_word.csv
"""

import csv
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

# --- Rutas oficiales (regla de Alexis, 16-ago-2026: manda el Word) -----------

AQUI = Path(__file__).resolve().parent

WORD_OFICIAL = Path(
    "/Users/alexis/Desktop/Go en Conflictos/RMD 2.0/"
    "RMD_2_Variables_y_METRICAS_COMPLETAS-11-06-2026.docx"
)
EXCEL_OFICIAL = (
    AQUI / "fuentes" /
    "Variables-y-Metricas-318-06-03-2026-Tabla(Recuperado automáticamente).xlsx"
)

# La línea que toda ficha de variable tiene, y que usamos como ancla.
ANCLA = ("fórmula numérica", "formula numérica")

# Caracteres que puede tener una sigla del RMD. Incluye guion bajo y guion medio
# (ICD-NARCO, IAS_X, IIP_X_2), dígitos (ICI3D), y letras griegas (η).
SIGLA = r"[A-Za-zÁÉÍÓÚÑáéíóúñ0-9_\-/ηΜΦΚΙ]{1,16}"


# --- Lectura del Word --------------------------------------------------------

def parrafos_del_word(ruta):
    """Devuelve el texto plano de cada párrafo del .docx, en orden.

    Se lee `word/document.xml` a mano en vez de usar python-docx porque las
    fórmulas del anexo son ecuaciones OMML y los extractores comunes las saltan
    en silencio (ver leer_docx_con_formulas.py). Acá sólo necesitamos el texto
    corrido de los encabezados, así que basta con juntar los <w:t>.
    """
    with zipfile.ZipFile(ruta) as z:
        xml = z.read("word/document.xml").decode("utf8")
    salida = []
    for p in re.findall(r"<w:p[ >].*?</w:p>", xml, re.S):
        texto = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", p, re.S))
        salida.append(re.sub(r"<[^>]+>", "", texto).strip())
    return salida


def encabezados_de_ficha(parrafos):
    """Ubica cada ficha por su ancla y devuelve (índice_párrafo, texto_encabezado)."""
    encabezados = []
    for i, t in enumerate(parrafos):
        if not t.lower().startswith(ANCLA):
            continue
        j = i - 1
        while j > 0 and not parrafos[j]:      # saltar párrafos vacíos
            j -= 1
        encabezados.append((j, parrafos[j]))
    return encabezados


# --- Interpretación del encabezado ------------------------------------------

# Cada patrón captura (numero, sigla, nombre) en el orden que declara su comentario.
PATRONES = [
    # "10.- ICS_T: Índice de Conflictividad Social"
    (re.compile(rf"^(\d+)\s*[\.\-–]*\s*({SIGLA})\s*:\s*(.+)$"), "n,s,nom"),
    # "31. IDIM — Índice de Discriminación Inversa Migratoria"
    (re.compile(rf"^(\d+)\s*[\.\-–]*\s*({SIGLA})\s*[—–]\s*(.+)$"), "n,s,nom"),
    # "4.- Anomalía de Precipitación — ANPrecip"
    (re.compile(rf"^(\d+)\s*[\.\-–]*\s*(.+?)\s*[—–]\s*({SIGLA})$"), "n,nom,s"),
    # "29 - Índice de Vulnerabilidad de Infraestructura Crítica - IVIC"
    (re.compile(rf"^(\d+)\s*[\.\-–]+\s*(.+?)\s+-\s+({SIGLA})$"), "n,nom,s"),
    # "7. Índice de Impacto Psicológico (IIP_X_2)"
    (re.compile(rf"^(\d+)\s*[\.\-–]*\s*(.+?)\s*\(({SIGLA})\)$"), "n,nom,s"),
    # "1.- Índice de Aceptación Multicultural"  (sin sigla: se resuelve por nombre)
    (re.compile(r"^(\d+)\s*[\.\-–]+\s*(.+)$"), "n,nom"),
]


def interpretar(encabezado):
    """Devuelve (numero, sigla|None, nombre) o None si no parece ficha."""
    for patron, orden in PATRONES:
        m = patron.match(encabezado)
        if not m:
            continue
        if orden == "n,s,nom":
            return int(m.group(1)), m.group(2).strip(), m.group(3).strip()
        if orden == "n,nom,s":
            return int(m.group(1)), m.group(3).strip(), m.group(2).strip()
        return int(m.group(1)), None, m.group(2).strip()
    return None


# --- Cruce contra el Excel ---------------------------------------------------

def normalizar(s):
    """Minúsculas, sin tildes y sin puntuación: para comparar nombres."""
    s = unicodedata.normalize("NFD", str(s).lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]", " ", s).strip()


def leer_excel(ruta):
    """Devuelve {sigla: (numero, nombre, categoria)} y {nombre_normalizado: sigla}."""
    from openpyxl import load_workbook
    ws = load_workbook(ruta, data_only=True)["Hoja1"]
    por_sigla, por_nombre = {}, {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila[1]:
            continue
        sigla = str(fila[1]).strip()
        nombre = str(fila[2]).strip()
        por_sigla[sigla] = (fila[0], nombre, str(fila[4]).strip())
        por_nombre.setdefault(normalizar(nombre), sigla)
    return por_sigla, por_nombre


# --- Informe -----------------------------------------------------------------

def main():
    parrafos = parrafos_del_word(WORD_OFICIAL)
    encabezados = encabezados_de_ficha(parrafos)
    por_sigla, por_nombre = leer_excel(EXCEL_OFICIAL)

    fichas, descartadas = [], []
    for indice, texto in encabezados:
        leido = interpretar(texto)
        if not leido:
            descartadas.append((indice, texto))
            continue
        numero, sigla, nombre = leido
        origen = "encabezado"
        if not sigla:                                  # sigla ausente en el título
            sigla = por_nombre.get(normalizar(nombre))
            origen = "recuperada del Excel por nombre" if sigla else "SIN RESOLVER"
        fichas.append(dict(parrafo=indice, numero=numero, sigla=sigla,
                           nombre=nombre, origen=origen))

    print("=" * 78)
    print("FICHAS DE VARIABLES EN EL WORD OFICIAL")
    print("=" * 78)
    print(f"  fichas encontradas (ancla «Fórmula Numérica»): {len(encabezados)}")
    print(f"  interpretadas                                : {len(fichas)}")
    print(f"  descartadas (no parecen ficha)               : {len(descartadas)}")
    sin_resolver = [f for f in fichas if f["origen"] == "SIN RESOLVER"]
    recuperadas = [f for f in fichas if f["origen"].startswith("recuperada")]
    print(f"  sigla tomada del encabezado                  : "
          f"{len(fichas) - len(sin_resolver) - len(recuperadas)}")
    print(f"  sigla recuperada del Excel por nombre        : {len(recuperadas)}")
    print(f"  sigla SIN RESOLVER (no se inventa)           : {len(sin_resolver)}")

    print("\n" + "=" * 78)
    print("SIGLAS DUPLICADAS EN EL WORD  ← lo que hay que desambiguar")
    print("=" * 78)
    agrupadas = defaultdict(list)
    for f in fichas:
        if f["sigla"]:
            agrupadas[f["sigla"]].append(f)
    duplicadas = {s: v for s, v in agrupadas.items() if len(v) > 1}
    # Sólo cuentan como duplicado real si los NOMBRES difieren: una misma
    # variable repetida en dos capítulos no es colisión de sigla.
    colisiones, repetidas = {}, {}
    for s, v in duplicadas.items():
        nombres = {normalizar(x["nombre"]) for x in v}
        (colisiones if len(nombres) > 1 else repetidas)[s] = v

    print(f"\n  COLISIONES REALES (misma sigla, variables distintas): {len(colisiones)}")
    for s in sorted(colisiones):
        print(f"\n  ▸ {s}")
        for f in colisiones[s]:
            en_excel = "  [ES LA DEL EXCEL]" if por_sigla.get(s) and \
                normalizar(por_sigla[s][1]) == normalizar(f["nombre"]) else ""
            print(f"      #{f['numero']:>3}  {f['nombre'][:70]}{en_excel}")
        if s in por_sigla:
            print(f"      Excel dice: #{por_sigla[s][0]} {por_sigla[s][1]} "
                  f"[{por_sigla[s][2]}]")
        else:
            print("      Excel: la sigla NO está")

    print(f"\n  MISMA VARIABLE REPETIDA (no es colisión): {len(repetidas)}")
    for s in sorted(repetidas):
        print(f"      {s} ×{len(repetidas[s])} — {repetidas[s][0]['nombre'][:60]}")

    print("\n" + "=" * 78)
    print("WORD vs EXCEL")
    print("=" * 78)
    siglas_word = {f["sigla"] for f in fichas if f["sigla"]}
    solo_word = sorted(siglas_word - set(por_sigla))
    solo_excel = sorted(set(por_sigla) - siglas_word)
    print(f"\n  En el WORD y NO en el Excel ({len(solo_word)}):")
    for s in solo_word:
        nom = next(f["nombre"] for f in fichas if f["sigla"] == s)
        print(f"      {s:14s} {nom[:60]}")
    print(f"\n  En el EXCEL y NO en el Word ({len(solo_excel)}):")
    for s in solo_excel:
        print(f"      {s:14s} {por_sigla[s][1][:60]} [{por_sigla[s][2]}]")

    if sin_resolver:
        print("\n  Fichas con sigla SIN RESOLVER:")
        for f in sin_resolver:
            print(f"      [{f['parrafo']}] #{f['numero']} {f['nombre'][:70]}")

    if "--csv" in sys.argv:
        destino = AQUI / "datos" / "fichas_word.csv"
        with open(destino, "w", newline="", encoding="utf8") as fh:
            w = csv.DictWriter(fh, fieldnames=["parrafo", "numero", "sigla",
                                               "nombre", "origen"])
            w.writeheader()
            w.writerows(fichas)
        print(f"\n  escrito: {destino}")


if __name__ == "__main__":
    main()
