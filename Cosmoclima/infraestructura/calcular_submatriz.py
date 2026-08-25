"""
CALCULAR LA CADENA CLIMÁTICA SOBRE UNA SUB-MATRIZ REAL
=======================================================

INSTRUCCIÓN (Alexis, 21-ago-2026)
-----------------------------------
«Ya tenemos el corpus básico de la MICR, ahora sí puedes calcular contra los
listados reales. Usa un listado pequeño primero para probar.»

Hasta ahora la cadena `PelPre → CClimP → FENef` sólo se había corrido sobre las
39 subestaciones del piloto y los 91 puntos del registro de deslizamientos. Este
script la corre sobre **cualquiera de las 31 sub-matrices** que ya están en
SharePoint, con activos reales y coordenada verificada.

    PelPre   peligro de precipitación del lugar y el mes
    CClimP   la perilla: 1,0 · 1,2 · 1,4 · 1,6
    FENef    la fragilidad del activo, en ESTE lugar y ESTE mes

★ EL PROBLEMA QUE HAY QUE RESOLVER ANTES DE CALCULAR NADA
-----------------------------------------------------------
`PelPre` es un **percentil nacional**: mide el evento contra la distribución de
todo el país. Y esa distribución `adaptadores/era5.py` la construye **con los
puntos que se le carguen**.

Si se corriera sobre 16 puntos, «percentil nacional» pasaría a significar
«comparado con estos 16», que es otra cosa y no se notaría: los números saldrían
igual de plausibles. Por eso este script **siempre agrupa la sub-matriz con los
130 puntos ya establecidos** (39 subestaciones + 91 del registro de
deslizamientos) y sólo después lee los resultados de los activos nuevos.

★★ Y esto destapa algo que hay que decir: **`PelPre` tiene el MISMO defecto de no
estacionariedad que `Pen`.** Su referencia crece cada vez que se agregan puntos,
así que el `PelPre` de un activo cambia según a quién más se haya medido. Con 130
puntos de base, agregar 16 lo mueve poco; agregar 16.669 lo movería mucho. **La
referencia nacional debería congelarse y declararse**, igual que el divisor de
las tres columnas de prioridad. Queda anotado, no resuelto.

USO
---
    ../.venv-esa/bin/python calcular_submatriz.py "351 Direcciones Regionales SENAPRED"
    ../.venv-esa/bin/python calcular_submatriz.py --listar
"""

import csv
import json
import re
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))
sys.path.insert(0, str(AQUI / "adaptadores"))
import normalizar                                    # noqa: E402
import cclimp                                        # noqa: E402
from adaptadores import era5                         # noqa: E402

ORIGEN = AQUI / "submatrices_excel"
DATOS = AQUI / "datos"
MICR = DATOS / "micr_sharepoint.csv"
# Los 130 puntos que fijan la referencia nacional establecida.
BASE = [(DATOS / "clima_diario_subestaciones_erA5.csv",
         DATOS / "subestaciones_puntos.csv"),
        (DATOS / "clima_diario_reterm_era5_corregido.csv",
         DATOS / "reterm_puntos.csv")]

DESDE, HASTA = "1990-01-01", "2026-08-20"
HILOS, PAUSA_S = 2, 1.5        # el servicio es gratuito: apurarlo devuelve 429
URL = ("https://archive-api.open-meteo.com/v1/archive?latitude={lat}"
       "&longitude={lon}&start_date=" + DESDE + "&end_date=" + HASTA +
       "&daily=precipitation_sum,temperature_2m_max,temperature_2m_min"
       "&timezone=America%2FSantiago")


def dms_a_decimal(texto):
    """«-27° 22' 45.47" S» → -27.37930. El signo puede venir por el número o
    por la letra del hemisferio; se respeta cualquiera de los dos."""
    m = re.match(r"\s*(-?)(\d+)°\s*(\d+)'\s*([\d.]+)\"?\s*([NSEW])?", str(texto))
    if not m:
        return None
    signo, g, mi, s, hemi = m.groups()
    v = int(g) + int(mi) / 60 + float(s) / 3600
    return round(-v if (signo == "-" or hemi in ("S", "W")) else v, 5)


def leer_submatriz(nombre):
    """Los activos de una sub-matriz, con su coordenada decimal."""
    ruta = ORIGEN / f"{nombre}.xlsx"
    if not ruta.exists():
        print("no existe:", ruta)
        raise SystemExit(1)
    ws = openpyxl.load_workbook(ruta, read_only=True).active
    filas = list(ws.iter_rows(values_only=True))
    cab = {c: i for i, c in enumerate(filas[0])}
    activos = []
    for f in filas[1:]:
        # el Excel trae decimal al final; si faltara, se deriva del formato DMS
        lat = f[cab["Latitud decimal"]] or dms_a_decimal(f[cab["Latitud"]])
        lon = f[cab["Longitud decimal"]] or dms_a_decimal(f[cab["Longitud"]])
        if lat is None or lon is None:
            continue
        activos.append({"id": f"{nombre[:3].strip()}·{f[cab['Título'] if 'Título' in cab else 0]}"[:120],
                        "nombre": f[0], "item": f[cab["Ítem"]],
                        "region": f[cab["Región"]], "comuna": f[cab["Comuna"]],
                        "lat": float(lat), "lon": float(lon)})
    return activos


def traer_punto(a, destino, reintentos=4):
    """Baja la serie diaria de un activo y la agrega al CSV. Reanudable."""
    for intento in range(reintentos):
        try:
            with urllib.request.urlopen(
                    URL.format(lat=a["lat"], lon=a["lon"]), timeout=120) as r:
                d = json.loads(r.read())["daily"]
            filas = list(zip(d["time"], d["precipitation_sum"],
                             d["temperature_2m_max"], d["temperature_2m_min"]))
            with destino.open("a", newline="", encoding="utf8") as fh:
                w = csv.writer(fh)
                for fecha, p, tx, tn in filas:
                    w.writerow([a["id"], fecha, p if p is not None else "",
                                tx if tx is not None else "",
                                tn if tn is not None else ""])
            return len(filas)
        except Exception as e:
            time.sleep(5 * (intento + 1))
            if intento == reintentos - 1:
                print(f"    ✗ {a['nombre'][:40]}: {e}")
                return 0


def main():
    if "--listar" in sys.argv:
        for p in sorted(ORIGEN.glob("*.xlsx")):
            n = openpyxl.load_workbook(p, read_only=True).active.max_row - 1
            print(f"   {n:7,d}  {p.stem}")
        return 0
    nombre = next((a for a in sys.argv[1:] if not a.startswith("-")), None)
    if not nombre:
        print(__doc__)
        return 1

    activos = leer_submatriz(nombre)
    micr = {int(x["n"]): (x["Sector"], x["elemento"], x["FEN"], x["Pen"])
            for x in csv.DictReader(MICR.open(encoding="utf-8"))}
    print("=" * 82)
    print(f"CADENA CLIMÁTICA SOBRE «{nombre}»")
    print("=" * 82)
    item = activos[0]["item"]
    sec, ele, fen, pen = micr[int(item)]
    print(f"\n  activos con coordenada : {len(activos)}")
    print(f"  ítem {int(item)} · {sec} · {ele}")
    print(f"  FEN de la Matriz: {fen}   ·   Pen: {pen}")

    # ── 1 · clima, reanudable ────────────────────────────────────────────────
    destino = DATOS / f"clima_diario_{nombre.split()[0]}.csv"
    ya = set()
    if destino.exists():
        with destino.open(encoding="utf-8") as fh:
            ya = {r[0] for r in csv.reader(fh) if r}
    else:
        with destino.open("w", newline="", encoding="utf8") as fh:
            csv.writer(fh).writerow(
                ["subestacion", "fecha", "precip_mm", "tmax_c", "tmin_c"])
    faltan = [a for a in activos if a["id"] not in ya]
    print(f"\n  serie climática: {len(ya)} ya bajados · {len(faltan)} por bajar")
    if faltan:
        t0 = time.time()
        with ThreadPoolExecutor(max_workers=HILOS) as ex:
            for i, n in enumerate(ex.map(lambda a: (time.sleep(PAUSA_S),
                                                    traer_punto(a, destino))[1],
                                         faltan), start=1):
                if i % 5 == 0 or i == len(faltan):
                    print(f"     {i}/{len(faltan)} · {time.time()-t0:.0f}s")

    # ── 2 · unir con los 130 puntos que fijan la referencia nacional ─────────
    combinado = DATOS / f"_combinado_{nombre.split()[0]}.csv"
    puntos = DATOS / f"_puntos_{nombre.split()[0]}.csv"
    with combinado.open("w", newline="", encoding="utf8") as fh:
        w = csv.writer(fh)
        w.writerow(["subestacion", "fecha", "precip_mm", "tmax_c", "tmin_c"])
        for clima, _ in BASE + [(destino, None)]:
            with clima.open(encoding="utf-8") as g:
                r = csv.reader(g)
                next(r, None)
                for fila in r:
                    w.writerow(fila)
    with puntos.open("w", newline="", encoding="utf8") as fh:
        w = csv.writer(fh)
        w.writerow(["subestacion", "lat", "lon"])
        for _, pfile in BASE:
            with pfile.open(encoding="utf-8") as g:
                for x in csv.DictReader(g):
                    w.writerow([x["subestacion"], x["lat"], x["lon"]])
        for a in activos:
            w.writerow([a["id"], a["lat"], a["lon"]])

    era5.CSV_DIARIO, era5.CSV_PUNTOS = combinado, puntos
    obs, problema = era5.traer()
    if problema:
        print("SIN DATO:", problema)
        return 1
    mios = {a["id"]: a for a in activos}
    serie = {}
    for o in obs:
        if o["variable"] != "peligro_precipitacion" or o["territorio_id"] not in mios:
            continue
        anio, mes = int(o["vigencia_inicio"][:4]), int(o["vigencia_inicio"][5:7])
        serie[(o["territorio_id"], anio, mes)] = (
            float(o["valor_normalizado"]), float(o["valor_original"]),
            float(o["confianza"]))
    print(f"  referencia nacional construida con "
          f"{len({o['territorio_id'] for o in obs})} puntos "
          f"({len(activos)} de esta sub-matriz + 130 de base)")
    print(f"  pares (activo, mes) de esta sub-matriz: {len(serie):,}")

    # ── 3 · la cadena ────────────────────────────────────────────────────────
    base_fen = cclimp.fen_base(fen, "tres")
    filas = []
    for (aid, anio, mes), (pel, mm, conf) in sorted(serie.items()):
        cc, _ = cclimp.coeficiente(pel, conf)
        filas.append(dict(activo=mios[aid]["nombre"], region=mios[aid]["region"],
                          anio=anio, mes=mes, mm_48h=round(mm, 1),
                          PelPre=round(pel, 4), CClimP=cc,
                          FENef_producto=round(cclimp.fenef(base_fen, cc, "producto"), 4),
                          FENef_potencia=round(cclimp.fenef(base_fen, cc, "potencia"), 4)))
    salida = DATOS / f"cadena_{nombre.split()[0]}.csv"
    with salida.open("w", newline="", encoding="utf8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(filas[0].keys()))
        w.writeheader()
        w.writerows(filas)

    from collections import Counter, defaultdict
    rep = Counter(f["CClimP"] for f in filas)
    print(f"\n  FEN base del tipo (escala de 3): {base_fen}")
    print("  reparto de la perilla sobre todos los pares activo-mes:")
    for v in (1.0, 1.2, 1.4, 1.6):
        print(f"      {v:.1f}  {rep[v]:7,d}  ({100*rep[v]/len(filas):5.1f} %)")

    print("\n  LOS DIEZ PARES MÁS PELIGROSOS DE LA SERIE COMPLETA:")
    for f in sorted(filas, key=lambda x: -x["PelPre"])[:10]:
        print(f"      {f['anio']}-{f['mes']:02d}  PelPre {f['PelPre']:.4f}  "
              f"{f['mm_48h']:6.1f} mm/48h  CClimP {f['CClimP']}  "
              f"{f['activo'][:44]}")

    print("\n  EXPOSICIÓN POR ACTIVO — % de meses con la perilla movida:")
    porac = defaultdict(list)
    for f in filas:
        porac[f["activo"]].append(f["CClimP"])
    orden = sorted(porac.items(),
                   key=lambda t: -sum(1 for v in t[1] if v > 1) / len(t[1]))
    for a, vs in orden:
        movidos = sum(1 for v in vs if v > 1)
        print(f"      {100*movidos/len(vs):5.1f} %  ({movidos:3d}/{len(vs)})  {a[:56]}")

    print(f"\n  escrito: {salida}  ({len(filas):,} filas)")
    combinado.unlink(); puntos.unlink()
    return 0


if __name__ == "__main__":
    sys.exit(main())
