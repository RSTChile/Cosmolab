"""
PILOTO · paso 1 — Trae la serie climática diaria de las 39 subestaciones.

Fuente: Open-Meteo Archive API (reanálisis ERA5), acceso anónimo, sin
credenciales, por latitud/longitud. Es la única fuente con cobertura nacional
homogénea disponible sin montar una red de estaciones.

LÍMITE DECLARADO — importante, no borrar:
    En la ronda 17 de Cosmoclima se comprobó que ERA5 en la zona de Illapel
    exagera los años secos (hasta 2,4x) y fabrica meses de lluvia que no
    ocurrieron. Por eso allá se abandonó el reanálisis y se usaron estaciones
    reales. Acá NO hay estaciones para 39 puntos repartidos en 35 grados de
    latitud, así que se usa ERA5 a sabiendas.
    En consecuencia: este piloto prueba si el MÉTODO funciona, no cuánto riesgo
    real corre cada subestación. Cualquier número de riesgo que salga de acá es
    provisional hasta contrastarlo con estaciones DMC/DGA.

Descarga día a día 1990-2026: precipitación, temperatura máxima y mínima.
No modifica nada; sólo escribe un CSV nuevo en datos/.
"""

import csv
import json
import re
import time
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import load_workbook

AQUI = Path(__file__).parent
FUENTE_XLSX = AQUI / "fuentes" / "Matriz 120-Subestaciones.xlsx"
SALIDA = AQUI / "datos" / "clima_diario_subestaciones_erA5.csv"
SALIDA_PUNTOS = AQUI / "datos" / "subestaciones_puntos.csv"

DESDE, HASTA = "1990-01-01", "2026-08-14"
# El servicio es público y gratuito: limita la tasa si se lo apura. Con 6 hilos
# devolvió HTTP 429 y sólo completó 16 de 39. Con 2 hilos y una pausa entre
# pedidos entra entero. Además el script es REANUDABLE: si el CSV ya trae una
# subestación completa, no la vuelve a pedir.
HILOS = 2
PAUSA_S = 1.5


def dms_a_decimal(texto):
    """Convierte «-27° 22' 45.47" S» a decimal. El signo puede venir por el
    número (-27) o por el hemisferio (S/W); se respeta cualquiera de los dos."""
    m = re.match(r"\s*(-?)(\d+)°\s*(\d+)'\s*([\d.]+)\"?\s*([NSEW])?", str(texto))
    if not m:
        return None
    signo, grados, minutos, segundos, hemisferio = m.groups()
    valor = int(grados) + int(minutos) / 60 + float(segundos) / 3600
    if signo == "-" or hemisferio in ("S", "W"):
        valor = -valor
    return round(valor, 5)


def leer_subestaciones():
    """Lee la sub-matriz y devuelve los 39 puntos con coordenada válida."""
    wb = load_workbook(FUENTE_XLSX, data_only=True, read_only=True)
    puntos = []
    for f in wb["Hoja1"].iter_rows(min_row=2, values_only=True):
        if not f[0]:
            continue
        lat, lon = dms_a_decimal(f[7]), dms_a_decimal(f[8])
        if lat is None or lon is None:
            print(f"  ¡OJO! sin coordenada legible: {f[0]}")
            continue
        # control de cordura: si cae fuera de Chile no se descarga nada
        if not (-56 <= lat <= -17 and -76 <= lon <= -66):
            print(f"  ¡OJO! coordenada fuera de Chile, se omite: {f[0]}")
            continue
        puntos.append({
            "subestacion": f[0],
            "tipo": "Rural" if "Rural" in str(f[0]) else "Urbana",
            "region": f[4], "provincia": f[5], "direccion": f[6],
            "responsable": f[9], "lat": lat, "lon": lon,
        })
    return puntos


def ya_bajadas():
    """Subestaciones que ya están completas en el CSV, para no volver a pedirlas."""
    if not SALIDA.exists():
        return {}
    conteo = {}
    with SALIDA.open(encoding="utf-8") as fh:
        for fila in csv.reader(fh):
            if fila and fila[0] != "subestacion":
                conteo[fila[0]] = conteo.get(fila[0], 0) + 1
    # ~13.375 días entre 1990-01-01 y 2026-08-14; exijo casi todos para darla por buena
    return {k: v for k, v in conteo.items() if v > 13000}


# Qué variables se piden. Open-Meteo cobra su cuota por PESO —número de días
# multiplicado por número de variables—, no por número de llamadas. Pidiendo las
# tres variables para 91 puntos nos cortó: 60 fallaron definitivamente tras 342
# reintentos. La medida de peligro sólo usa la lluvia, así que pedir sólo eso
# baja el peso a un tercio. La temperatura se pide únicamente cuando hace falta.
VARIABLES = "precipitation_sum,temperature_2m_max,temperature_2m_min"


def traer_punto(p, reintentos=5):
    """Pide la serie diaria de un punto. Reintenta con espera creciente."""
    time.sleep(PAUSA_S)
    url = ("https://archive-api.open-meteo.com/v1/archive"
           f"?latitude={p['lat']}&longitude={p['lon']}"
           f"&start_date={DESDE}&end_date={HASTA}"
           f"&daily={VARIABLES}"
           "&timezone=UTC")
    for intento in range(reintentos):
        try:
            with urllib.request.urlopen(url, timeout=180) as r:
                d = json.load(r)
            dia = d["daily"]
            # Si sólo se pidió lluvia, las columnas de temperatura no vienen en
            # la respuesta. Se rellenan vacías para que el CSV conserve siempre
            # la misma forma y nadie tenga que adivinar cuántas columnas trae.
            n = len(dia["time"])
            tmax = dia.get("temperature_2m_max") or [None] * n
            tmin = dia.get("temperature_2m_min") or [None] * n
            return [(p["subestacion"], fecha, pr, tx, tn)
                    for fecha, pr, tx, tn
                    in zip(dia["time"], dia["precipitation_sum"], tmax, tmin)]
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError,
                json.JSONDecodeError) as e:
            espera = 10 * (intento + 1)
            print(f"    reintento {intento+1}/{reintentos} en {p['subestacion']}"
                  f" ({type(e).__name__}); espero {espera}s")
            time.sleep(espera)
    print(f"  FALLÓ definitivamente: {p['subestacion']}")
    return []


def main():
    (AQUI / "datos").mkdir(exist_ok=True)
    puntos = leer_subestaciones()
    print(f"{len(puntos)} subestaciones con coordenada válida")

    with SALIDA_PUNTOS.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(puntos[0].keys()))
        w.writeheader()
        w.writerows(puntos)

    hechas = ya_bajadas()
    faltan = [p for p in puntos if p["subestacion"] not in hechas]
    print(f"ya estaban completas: {len(hechas)} · por bajar: {len(faltan)}")

    t0 = time.time()
    nuevas = []
    if faltan:
        with ThreadPoolExecutor(max_workers=HILOS) as ex:
            nuevas = [f for grupo in ex.map(traer_punto, faltan) for f in grupo]

    # reescribo el CSV entero: lo que ya estaba (sólo de puntos completos) + lo nuevo
    previas = []
    if SALIDA.exists():
        with SALIDA.open(encoding="utf-8") as fh:
            previas = [f for f in csv.reader(fh)
                       if f and f[0] != "subestacion" and f[0] in hechas]
    filas = previas + nuevas
    with SALIDA.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["subestacion", "fecha", "precip_mm", "tmax_c", "tmin_c"])
        w.writerows(filas)

    completos = len(ya_bajadas())
    print(f"{len(filas):,} filas de {completos}/{len(puntos)} subestaciones "
          f"en {time.time()-t0:.0f}s")
    print(f"escrito: {SALIDA}")
    if completos < len(puntos):
        print("ATENCIÓN: faltan subestaciones. Volver a correr antes de analizar.")


if __name__ == "__main__":
    main()
