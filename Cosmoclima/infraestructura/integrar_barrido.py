"""
INTEGRAR EL BARRIDO · de 77 carpetas de crudo al índice de activos
====================================================================

INSTRUCCIÓN (Alexis, 25-ago-2026): «integra todo, y para Industrial crea las
categorías generales que sean necesarias».

Los siete agentes bajaron 911.182 registros verificados. Este paso los normaliza
al formato del índice: un activo por fila, con su ítem de la Matriz, su
coordenada y su origen declarado.

★ EL MAPEO ESTÁ AQUÍ, A LA VISTA, Y ES LA PARTE DISCUTIBLE
------------------------------------------------------------
Cada entrada dice de qué archivo sale, a qué ítem va y por qué. Nadie tiene que
leer el código para discutir una asignación: están todas en `FUENTES`, con su
motivo. Donde el dato no identifica el ítem, la fuente NO se reparte — es la
regla que este proyecto viene aplicando desde el rubro «Producción de alimentos»
del RETC.

★★ LO QUE NO ENTRA, Y POR QUÉ
-------------------------------
    militares de OpenStreetMap    cartografía voluntaria de instalaciones de
                                  defensa; meterla le daría un respaldo que no
                                  tiene. Esos ítems son «no observables»
    componentes internos          aulas, baños, ascensores: poblarlos heredando
                                  la coordenada del edificio diría «hay un
                                  ascensor aquí» sin que nadie lo catastrara
    registros sin coordenada      SERVEL, delegaciones, hoteles de SERNATUR:
                                  geocodificar direcciones está prohibido
    térmicas sin combustible      las 21 que no se pudieron identificar no van
                                  a diésel «porque es lo más probable»

USO
---
    ../.venv-esa/bin/python integrar_barrido.py
    ../.venv-esa/bin/python integrar_barrido.py --lote glaciares
"""

import csv
import glob
import json
import math
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

AQUI = Path(__file__).resolve().parent
CRUDO = AQUI / "busqueda" / "crudo"
SALIDA = AQUI / "datos" / "barrido_por_item.csv"
INFORME = AQUI / "busqueda" / "INTEGRACION.md"

CAMPOS = ["item", "elemento", "nombre", "lat", "lon", "comuna", "region",
          "fuente", "confianza"]

# ★ Lo que se deja fuera se cuenta y se nombra. Un integrador que descarta en
#   silencio produce un total que parece completo y no lo es.
DESCARTES = defaultdict(list)


def dentro(la, lo):
    return -56 < la < -17 and -110 < lo < -66


def num(v):
    """Acepta 3,14 y 3.14, y el texto que algunos servicios devuelven en vez
    de número. Devuelve None en vez de reventar."""
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str) or not v.strip():
        return None
    try:
        return float(v.strip().replace(",", "."))
    except ValueError:
        return None


def sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn").lower().strip()


# ── lectores ────────────────────────────────────────────────────────────────

def leer_csv(ruta, **kw):
    with open(ruta, encoding="utf-8-sig", errors="replace", newline="") as fh:
        yield from csv.DictReader(fh, **kw)


def leer_xlsx(ruta, hoja=None):
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else "" for c in next(it)]
    for f in it:
        yield dict(zip(cab, f))


def leer_geojson(ruta):
    d = json.loads(Path(ruta).read_text(encoding="utf-8"))
    for f in d.get("features", d if isinstance(d, list) else []):
        yield f


def punto_de(geom):
    """★ El punto que representa una geometría. Para líneas y polígonos se toma
    el vértice CENTRAL de la traza y no el centroide: en un tendido o un canal
    que rodea un cerro, el centroide cae fuera del objeto."""
    if not geom:
        return None
    t = geom.get("type")
    c = geom.get("coordinates")
    if not c:
        return None
    if t == "Point":
        return c[0], c[1]
    pts = []
    pila = [c]
    while pila:
        x = pila.pop()
        if x and isinstance(x[0], (int, float)):
            pts.append(x)
        elif x:
            pila.extend(x)
    if not pts:
        return None
    p = pts[len(pts) // 2]
    return p[0], p[1]


def gms_concatenado(v):
    """★★ UNA TRAMPA QUE NO DA ERROR: las coordenadas del registro de
    radiodifusión de SUBTEL vienen en grados-minutos-segundos CONCATENADOS y sin
    separador. `LATPTA = 182857` no es 182.857: es **18° 28' 57"**.

    Leerlo como decimal no falla, no avisa, y deja la radio en mitad del océano
    Índico. Se valida el resultado contra el territorio, que es la única defensa.

    ⚠️ Además el datum declarado es PSAD 56, no WGS 84. La diferencia en Chile
    ronda los 200-300 m: suficiente para ubicar una planta transmisora en su
    cerro, insuficiente para nada que exija precisión métrica. Queda declarado.
    """
    # Los segundos pueden traer decimales: «220548.9» es 22° 05' 48,9".
    # Se recorta la fracción de segundo, que a esta escala es irrelevante
    # (0,1" ≈ 3 m) y de otro modo tiraba 37 concesiones válidas.
    bruto = str(v or "").strip()
    # Segundo formato presente en el mismo archivo: «34°24'47,26''». La fuente
    # mezcla los dos y sólo se ve al validar contra el territorio.
    if "°" in bruto:
        import re as _re
        n = _re.findall(r"\d+(?:[.,]\d+)?", bruto)
        if len(n) < 3:
            return None
        g, m, seg = (float(x.replace(",", ".")) for x in n[:3])
        return -(g + m / 60 + seg / 3600)
    s = bruto.split(".")[0]
    if not s.isdigit() or len(s) < 5 or int(s) == 0:
        return None
    s = s.zfill(6)
    g, m, seg = int(s[:-4]), int(s[-4:-2]), int(s[-2:])
    if m > 59 or seg > 59:
        return None
    return -(g + m / 60 + seg / 3600)


# ── el mapeo, que es la parte discutible ────────────────────────────────────

def f_termicas():
    """Las 192 con combustible identificado. Las 21 sin identificar NO entran."""
    m = {"carbón": ("89", "Plantas de Generación de Energía (Carbón)"),
         "carbon": ("89", "Plantas de Generación de Energía (Carbón)"),
         "gas natural": ("90", "Plantas de Generación de Energía (Gas Natural)"),
         "diesel": ("91", "Plantas de Generación de Energía (Diésel)"),
         "diésel": ("91", "Plantas de Generación de Energía (Diésel)"),
         "biomasa": ("106", "Plantas de Biomasa"),
         # El biogás sale de digestión anaerobia y no de quemar leña, pero la
         # Matriz no tiene ítem propio: entra en Biomasa y queda declarado.
         "biogas": ("106", "Plantas de Biomasa"),
         "licor negro": ("106", "Plantas de Biomasa")}
    p = CRUDO / "coordinador_combustible" / "2026-08-25" / "centrales_termicas_combustible.csv"
    if not p.exists():
        return
    for r in leer_csv(p):
        comb = sin_tildes(r.get("combustible_ficha_coordenador")
                          or r.get("combustible_ficha_coordinador")
                          or r.get("combustible_capa_cne_2018") or "")
        it = None
        for clave, destino in m.items():
            if sin_tildes(clave) in comb:
                it = destino
                break
        la, lo = num(r.get("latitud")), num(r.get("longitud"))
        if not it:
            # ⚠️ Se declara lo que NO se reparte. Una térmica sin combustible
            #    identificado no va a diésel «porque es lo más probable».
            DESCARTES["térmicas sin combustible identificable"].append(
                f"{r.get('central','?')} [{comb or 'vacío'}]")
            continue
        if la is None or lo is None or not dentro(la, lo):
            DESCARTES["térmicas sin coordenada válida"].append(r.get("central", "?"))
            continue
        yield {"item": it[0], "elemento": it[1], "nombre": r.get("central", ""),
               "lat": la, "lon": lo, "comuna": r.get("comuna", ""),
               "region": r.get("region", ""),
               "fuente": "Coordinador Eléctrico · fichas técnicas + CNE",
               "confianza": "consolidado"}


def f_glaciares():
    p = CRUDO / "dga_glaciares" / "2026-08-25" / "glaciares_centroides.csv"
    if not p.exists():
        return
    for r in leer_csv(p):
        la, lo = num(r.get("lat")), num(r.get("lon"))
        if la is None or lo is None or not dentro(la, lo):
            continue
        n = (r.get("nombre") or "").strip()
        yield {"item": "1", "elemento": "Glaciares",
               "nombre": "" if n.lower() in ("s/n", "sin nombre") else n,
               "lat": la, "lon": lo, "comuna": r.get("comuna", ""),
               "region": r.get("region", ""),
               "fuente": "DGA · Inventario Público de Glaciares 2022 v2",
               "confianza": "consolidado"}


def f_radios():
    """Planta transmisora de cada concesión de radiodifusión vigente."""
    p = (CRUDO / "subtel_radiodifusion" / "2026-08-25" /
         "Actualiza_julio_2026_web.xlsx")
    if not p.exists():
        return
    import openpyxl
    ws = openpyxl.load_workbook(p, read_only=True, data_only=True)["Listado"]
    it = ws.iter_rows(values_only=True)
    for _ in range(4):
        next(it)
    cab = [str(c).replace("\n", " ").strip() if c is not None else "" for c in next(it)]
    for f in it:
        r = dict(zip(cab, f))
        la, lo = gms_concatenado(r.get("LATPTA")), gms_concatenado(r.get("LONGPTA"))
        if la is None or lo is None or not dentro(la, lo):
            DESCARTES["radios con coordenada ilegible"].append(
                f"{(r.get('NOMBRERADIO') or '?')} [{r.get('LATPTA')}/{r.get('LONGPTA')}]")
            continue
        nom = (r.get("NOMBRERADIO") or "").strip()
        señal = (r.get("SEÑAL") or "").strip()
        yield {"item": "222", "elemento": "Estaciones de Radio",
               "nombre": f"{nom} {señal}".strip()[:70],
               "lat": round(la, 6), "lon": round(lo, 6),
               "comuna": r.get("COMUNA PLANTA", ""), "region": r.get("REG PLANTA", ""),
               "fuente": "SUBTEL · Concesiones de radiodifusión jul-2026 (PSAD 56)",
               "confianza": "consolidado"}


def f_sppc():
    """Sitios con potencial presencia de contaminantes. Se reparten sólo los
    que la propia actividad declarada identifica; el resto queda fuera."""
    m = {"estaciones de servicio de combustibles":
         ("113", "Almacenamiento de Combustibles (Tanques)"),
         "disposicion de residuos solidos":
         ("684", "Instalaciones de Tratamiento (Residuos)"),
         "mal manejo de sustancias peligrosas":
         ("675", "Materiales Químicos (Peligrosos)"),
         "formulacion o fabricacion de productos quimicos":
         ("676", "Materiales Químicos (Producción)"),
         "reciclaje y valorizacion de residuos":
         ("692", "Almacenes de Residuos Químicos"),
         "industria forestal":
         ("848", "Aserraderos y Plantas de Elaboración de Madera"),
         "fabricacion de muebles a nivel industrial":
         ("851", "Plantas Manufactureras (General)")}
    p = CRUDO / "mma_sppc_suelos_contaminantes" / "2026-08-25" / "sppc_puntos.csv"
    if not p.exists():
        return
    for r in leer_csv(p):
        it = m.get(sin_tildes(r.get("actividad_principal")))
        la, lo = num(r.get("lat")), num(r.get("lon"))
        if not it or la is None or lo is None or not dentro(la, lo):
            continue
        n = (r.get("nombre_sitio") or "").strip()
        emp = (r.get("nombre_empresa") or "").strip()
        yield {"item": it[0], "elemento": it[1],
               "nombre": (emp if emp and emp != "Sin Información" else n)[:70],
               "lat": la, "lon": lo, "comuna": r.get("comuna", ""),
               "region": r.get("region", ""),
               "fuente": "MMA · Sitios con Potencial Presencia de Contaminantes",
               "confianza": "consolidado"}


# CIIU rev.4 sección C (manufactura) → ítem. Las divisiones que NO son de
# Industrial van a su sector propio: alimentos al Alimentario, químicos al
# Químico, refinación a Energía. La frontera la pone la CIIU, no nosotros.
CIIU_A_ITEM = {
    "23": ("847", "Plantas de Materiales de Construcción (Cemento, Cal y Áridos)"),
    "16": ("848", "Aserraderos y Plantas de Elaboración de Madera"),
    "24": ("849", "Plantas Metalúrgicas y Metalmecánicas"),
    "25": ("849", "Plantas Metalúrgicas y Metalmecánicas"),
    "17": ("850", "Plantas de Papel y Celulosa"),
    "20": ("660", "Plantas Químicas"),
    "21": ("660", "Plantas Químicas"),
    "19": ("111", "Refinerías de Petróleo"),
    "11": ("409", "Plantas de Envasado (Alimentos)"),
}
for _d in ("13", "14", "15", "18", "22", "26", "27", "28", "29", "30", "31",
           "32", "33"):
    CIIU_A_ITEM[_d] = ("851", "Plantas Manufactureras (General)")


def f_manufactura():
    """★ Los 1.514 manufactureros del RETC, repartidos por división CIIU.

    ⚠️ La división 10 (elaboración de alimentos, 575 establecimientos) NO se
    reparte: dentro caben panaderías, plantas de carnes, de lácteos y de
    conservas, y la Matriz tiene un ítem distinto para cada una. Repartirlas por
    el nombre de la actividad sería inventar a qué se dedican — el mismo error
    que ya costó el ítem 660 en el primer intento con el RETC.
    """
    p = (CRUDO / "retc_fuentes_puntuales" / "2026-08-25" / "ckan_fp_2024_da.xlsx")
    if not p.exists():
        return
    vistos = {}
    for r in leer_xlsx(p, "Data"):
        idv = r.get("id_vu")
        if idv is not None:
            vistos[idv] = r
    for r in vistos.values():
        cid = str(r.get("ciiu4_id") or "")
        if not cid.startswith("C"):
            continue
        it = CIIU_A_ITEM.get(cid[1:3])
        la, lo = num(r.get("latitud")), num(r.get("longitud"))
        if not it or la is None or lo is None or not dentro(la, lo):
            continue
        yield {"item": it[0], "elemento": it[1],
               "nombre": str(r.get("nombre_establecimiento")
                             or r.get("razon_social") or "")[:70],
               "lat": la, "lon": lo, "comuna": r.get("comuna", ""),
               "region": r.get("region", ""),
               "fuente": f"RETC · fuentes puntuales 2024 (CIIU {cid})",
               "confianza": "consolidado"}


FUENTES = [
    ("termicas", f_termicas, "centrales térmicas con combustible identificado"),
    ("glaciares", f_glaciares, "inventario público de glaciares"),
    ("radios", f_radios, "plantas transmisoras de radiodifusión"),
    ("sppc", f_sppc, "sitios con potencial presencia de contaminantes"),
    ("manufactura", f_manufactura, "manufactureros del RETC por división CIIU"),
]


def main():
    solo = None
    if "--lote" in sys.argv:
        solo = sys.argv[sys.argv.index("--lote") + 1]

    filas, por_lote = [], {}
    for clave, fn, desc in FUENTES:
        if solo and clave != solo:
            continue
        print(f"  {clave:<14}{desc}…", flush=True)
        n = 0
        for fila in fn():
            filas.append(fila)
            n += 1
        por_lote[clave] = n
        print(f"  {'':<14}→ {n:,} activos")

    if not filas:
        print("\n  nada que integrar")
        return 1

    por_item = Counter((f["item"], f["elemento"]) for f in filas)
    print(f"\n  {'ítem':<6}{'elemento':<50}{'activos':>9}")
    print("  " + "-" * 68)
    for (n, el), c in sorted(por_item.items(), key=lambda t: -t[1]):
        print(f"  {n:<6}{el[:49]:<50}{c:>9,}")

    with SALIDA.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=CAMPOS, extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)
    print(f"\n  escrito: {SALIDA.name} · {len(filas):,} activos en "
          f"{len(por_item)} ítems ({SALIDA.stat().st_size/1e6:.1f} MB)")

    if DESCARTES:
        print(f"\n  ⚠️ NO INTEGRADO, y por qué:")
        for motivo, lista in sorted(DESCARTES.items(), key=lambda t: -len(t[1])):
            print(f"       {len(lista):>5}  {motivo}")
            for x in lista[:3]:
                print(f"              · {x[:66]}")
            if len(lista) > 3:
                print(f"              · …y {len(lista)-3} más")
    return 0




# ── LOTE 2 · el visor GRD de SENAPRED y otros catastros de punto ────────────

def esri(carpeta, archivo=None):
    """Los servicios de SENAPRED devuelven esriJSON: `attributes` + `geometry`
    con `x`/`y`, no GeoJSON."""
    d = CRUDO / carpeta / "2026-08-25"
    if not d.exists():
        return
    p = d / archivo if archivo else next(
        (x for x in d.glob("*.json") if "PROCEDENCIA" not in x.name), None)
    if not p or not p.exists():
        return
    o = json.loads(p.read_text(encoding="utf-8"))
    fs = o["features"] if isinstance(o, dict) and "features" in o else o
    for f in fs:
        a = f.get("attributes") or f.get("properties") or {}
        g = f.get("geometry") or {}
        la, lo = num(g.get("y")), num(g.get("x"))
        if la is None or lo is None:
            la, lo = num(a.get("latitud")), num(a.get("longitud"))
        if la is None or lo is None or not dentro(la, lo):
            continue
        yield a, la, lo


def fila(item, elemento, a, la, lo, fuente, nombre_campos, conf="consolidado"):
    nom = ""
    for k in nombre_campos:
        v = str(a.get(k) or "").strip()
        if v and v.lower() not in ("none", "s/i", "sin información"):
            nom = v
            break
    return {"item": item, "elemento": elemento, "nombre": nom[:70],
            "lat": round(la, 6), "lon": round(lo, 6),
            "comuna": a.get("comuna") or a.get("nom_com_es") or "",
            "region": a.get("region") or a.get("región") or "",
            "fuente": fuente, "confianza": conf}


# tipo declarado por SENAPRED → ítem de Salud. Lo que no calza no se reparte.
SALUD = {
    "hospital": ("265", "Hospitales Generales"),
    "posta de salud rural": ("266", "Clínicas Rurales"),
    "servicio de urgencia rural": ("266", "Clínicas Rurales"),
    "centro de salud familiar": ("267", "Centros de Atención Primaria"),
    "centro comunitario de salud familiar": ("267", "Centros de Atención Primaria"),
    "servicio de atención primaria de urgencia": ("267", "Centros de Atención Primaria"),
    "centro de salud": ("267", "Centros de Atención Primaria"),
    "laboratorio clínico": ("270", "Laboratorios Clínicos"),
}


def f_senapred_salud():
    for a, la, lo in esri("senapred_salud"):
        t = sin_tildes(a.get("tipo"))
        it = None
        for clave, destino in SALUD.items():
            if sin_tildes(clave) in t:
                it = destino
                break
        if not it:
            DESCARTES["salud: tipo sin ítem en la Matriz"].append(str(a.get("tipo"))[:40])
            continue
        yield fila(it[0], it[1], a, la, lo, "SENAPRED · visor GRD",
                   ("nombre", "nombre_del"))


def f_senapred_simple():
    """Capas del visor GRD donde toda la capa es un mismo ítem."""
    # ⚠️ TRES CAPAS BAJADAS Y VERIFICADAS QUE NO ENTRAN, PORQUE LA MATRIZ NO
    #    TIENE ÍTEM DONDE PONERLAS. No es un descuido: es el mismo hallazgo que
    #    ya apareció en Industrial, y conviene que quede contado.
    #
    #      82  centros penitenciarios  → no hay ítem de cárcel ni de recinto
    #                                    de detención en los 851
    #      47  servicio médico legal   → no hay ítem forense
    #   4.503  farmacias comunitarias  → el único ítem de farmacia es el 272,
    #                                    «Farmacias HOSPITALARIAS», que es otra
    #                                    cosa: meterlas ahí haría falso el ítem
    #  11.951  jardines infantiles     → ver f_parvularia()
    capas = [
        ("senapred_bomberos", "309", ("nombre", "compa__ia")),
        ("senapred_carabineros", "353", ("nombre_uni", "tipo_de_un")),
        ("senapred_pdi", "353", ("unidad", "inmueble")),
    ]
    for carpeta, cuantos, motivo in (
            ("senapred_centros_penitenciarios", 82,
             "no hay ítem de cárcel ni recinto de detención en la Matriz"),
            ("senapred_servicio_medico_legal", 47, "no hay ítem forense"),
            ("senapred_farmacias", 4503,
             "el único ítem de farmacia (272) es HOSPITALARIA, no comunitaria")):
        DESCARTES[motivo].append(f"{carpeta} ({cuantos:,} registros verificados)")
    m = json.loads((AQUI / "web" / "publico" / "datos" / "matriz.json")
                   .read_text(encoding="utf-8"))
    valido = {str(i["n"]): i["elemento"] for i in m["items"]}
    for carpeta, it, campos in capas:
        if it not in valido:
            DESCARTES[f"capa {carpeta}: el ítem {it} no existe en la Matriz"].append(it)
            continue
        for a, la, lo in esri(carpeta):
            yield fila(it, valido[it], a, la, lo, "SENAPRED · visor GRD", campos)


def f_deportiva():
    """★ El catastro no distingue estadio de gimnasio: `categoria` dice «Base»
    en las 20.701. Lo único que separa es el NOMBRE, así que se usa eso y se
    declara: si el recinto se llama «Estadio», va al ítem de estadios."""
    p = (CRUDO / "infraestructura_deportiva" / "2026-08-25" /
         "infraestructura_deportiva.geojson")
    if not p.exists():
        return
    for f in leer_geojson(p):
        pt = punto_de(f.get("geometry"))
        if not pt or not dentro(pt[1], pt[0]):
            continue
        pr = f.get("properties", {})
        nom = str(pr.get("nombre") or "").strip()
        es_estadio = "estadio" in sin_tildes(nom)
        it, el = (("529", "Estadios Deportivos") if es_estadio
                  else ("567", "Gimnasios y Centros Deportivos"))
        yield {"item": it, "elemento": el, "nombre": nom[:70],
               "lat": round(pt[1], 6), "lon": round(pt[0], 6),
               "comuna": "", "region": pr.get("region", ""),
               "fuente": "MINDEP · Infraestructura Deportiva",
               "confianza": "consolidado"}


def f_parvularia():
    """★★ 11.951 JARDINES INFANTILES QUE NO TIENEN DÓNDE ENTRAR.

    El sector Educación de la Matriz enumera escuelas primarias, secundarias,
    universidades, institutos, centros de formación técnica… y **no tiene ítem
    para el nivel parvulario**. Son 11.951 establecimientos georreferenciados,
    con niños de 0 a 6 años, que el instrumento no puede ver.

    Es el mismo hallazgo que en Industrial pero más grave: ahí faltaba la
    categoría de una fábrica; aquí falta la de un jardín infantil.

    No se fuerza a ningún ítem. Queda contado y declarado.
    """
    p = CRUDO / "parvularia" / "2026-08-25" / "establecimientos_parvularia.geojson"
    if not p.exists():
        return
    n = sum(1 for _ in leer_geojson(p))
    DESCARTES["no hay ítem de jardín infantil / nivel parvulario en la Matriz"].append(
        f"MINEDUC · Educación Parvularia ({n:,} establecimientos con coordenada)")
    return
    yield  # noqa: la función es un generador por contrato


FUENTES += [
    ("salud", f_senapred_salud, "establecimientos de salud del visor GRD"),
    ("emergencia", f_senapred_simple, "bomberos, policía, cárceles, farmacias"),
    ("deportiva", f_deportiva, "infraestructura deportiva del MINDEP"),
    ("parvularia", f_parvularia, "educación parvularia"),
]


# ── LOTE 3 · shapefiles y catastros de superficie ───────────────────────────

def leer_shp(zip_path, patron="*.shp"):
    """Los shapefiles vienen dentro del ZIP tal como los publica el organismo.
    Se descomprimen a un temporal y se leen con pyshp — sin GDAL, que en este
    proyecto nunca hizo falta."""
    import shapefile
    import tempfile
    import zipfile
    z = Path(zip_path)
    if not z.exists():
        return
    with tempfile.TemporaryDirectory() as tmp:
        with zipfile.ZipFile(z) as zf:
            zf.extractall(tmp)
        shps = sorted(Path(tmp).rglob(patron))
        if not shps:
            return
        r = shapefile.Reader(str(shps[0]))
        campos = [f[0] for f in r.fields[1:]]
        for sr in r.iterShapeRecords():
            yield dict(zip(campos, sr.record)), sr.shape


def punto_shape(shape):
    pts = getattr(shape, "points", None)
    if not pts:
        return None
    x, y = pts[len(pts) // 2]
    return (y, x) if dentro(y, x) else None


def f_antenas_torres():
    """Las 18.007 antenas de la Ley de Torres. Van al ítem 183 (Torres de
    Telecomunicaciones), que ya está poblado: lo que duplique lo descarta la
    fusión por proximidad, y lo que quede es lo que esa fuente añade."""
    z = (CRUDO / "subtel_antenas_leydetorres" / "2026-08-25" /
         "antenas_servicio_ley_torres.zip")
    for a, shape in leer_shp(z):
        pt = punto_shape(shape)
        if not pt:
            continue
        yield {"item": "183", "elemento": "Torres de Telecomunicaciones (Celulares)",
               "nombre": f"{a.get('ALIAS','')} {a.get('TISO_DESCR','')}".strip()[:70],
               "lat": round(pt[0], 6), "lon": round(pt[1], 6),
               "comuna": a.get("COMUNA", ""), "region": a.get("REGION", ""),
               "fuente": "SUBTEL · Antenas en Servicio Ley de Torres",
               "confianza": "consolidado"}


def f_hidrografia():
    """★ 122.852 cursos de agua. Son LÍNEAS: se toma el vértice central."""
    z = CRUDO / "ide_hidrografia" / "2026-08-25" / "Hidrografia_V2.zip"
    for a, shape in leer_shp(z):
        pt = punto_shape(shape)
        if not pt:
            continue
        nom = ""
        for k in a:
            if "nom" in k.lower() and str(a[k]).strip():
                nom = str(a[k]).strip()
                break
        yield {"item": "3", "elemento": "Fuentes Naturales de Agua (Ríos o Lagos)",
               "nombre": nom[:70], "lat": round(pt[0], 6), "lon": round(pt[1], 6),
               "comuna": "", "region": "",
               "fuente": "IDE Chile · Hidrografía V2",
               "confianza": "consolidado"}


def f_geojson(carpeta, archivo, item, elemento, campos_nombre, fuente,
              campo_comuna="", campo_region=""):
    """Lector genérico para las capas donde toda la capa es un mismo ítem."""
    p = CRUDO / carpeta / "2026-08-25" / archivo
    if not p.exists():
        return
    for f in leer_geojson(p):
        pt = punto_de(f.get("geometry"))
        if not pt or not dentro(pt[1], pt[0]):
            continue
        pr = f.get("properties", {})
        nom = ""
        for k in campos_nombre:
            v = str(pr.get(k) or "").strip()
            if v and v.lower() not in ("none", "s/i"):
                nom = v
                break
        yield {"item": item, "elemento": elemento, "nombre": nom[:70],
               "lat": round(pt[1], 6), "lon": round(pt[0], 6),
               "comuna": pr.get(campo_comuna, "") if campo_comuna else "",
               "region": pr.get(campo_region, "") if campo_region else "",
               "fuente": fuente, "confianza": "consolidado"}


def f_bovinos():
    yield from f_geojson("ciren_rol_unico_pecuario",
                         "rol_unico_pecuario_bovinos_2025.geojson",
                         "399", "Granjas Ganaderas (Bovinos)", ("rup", "id"),
                         "SAG/CIREN · Rol Único Pecuario bovinos 2025", "nom_com")


def f_fruticola():
    yield from f_geojson("ciren_catastro_fruticola", "productores_fruticolas.geojson",
                         "398", "Campos de Cultivo (Frutas y Verduras)",
                         ("especie_01", "rolpredi"),
                         "CIREN · Catastro Frutícola", "desccomu")


def f_bocatomas():
    """Bocatomas de la CNR. Van al 403 (Sistemas de Riego Agrícola) y no al 844
    (Bocatoma de Agua Potable RURAL), que es otra cosa: éstas son de riego."""
    yield from f_geojson("cnr_infraestructura_riego", "bocatomas.geojson",
                         "403", "Sistemas de Riego Agrícola", ("nom_can", "cod_boc"),
                         "CNR · Bocatomas de riego")


def f_acuiferos():
    yield from f_geojson("dga_acuiferos",
                         "acuiferos_vegas_bofedales_protegidos.geojson",
                         "5", "Acuíferos Subterráneos", ("NOM_VEGA", "COD_ACUIFVF"),
                         "DGA · Acuíferos que alimentan vegas y bofedales",
                         "", "REGION")


def f_agricola():
    """★★ LOS 74.981 POLÍGONOS DE CONAF, AHORA SÍ.

    Estos polígonos no entraban porque la Matriz separaba «Campos de Cultivo
    (Granos)» de «(Frutas y Verduras)» y el catastro nacional no usa esa
    división: usa **Rotación Cultivo-Pradera** y **Terreno de Uso Agrícola**.

    En vez de forzar el catastro al ítem, se adaptó el ítem al catastro
    (`alinear_items_agricolas.py`), que es lo mismo que se hizo en Industrial
    con las divisiones CIIU:

        397  renombrado a «Terreno de Uso Agrícola»
        852  ítem nuevo · «Rotación Cultivo-Pradera»

    ⚠️ El catastro escribe la misma clase de tres formas —«Terreno de Uso
    Agrícola», «Terrenos de Uso Agricola» y «Terrenos de Uso Agrícola»—, con y
    sin tilde y en singular y plural. Se normalizan: si no, 2.327 polígonos se
    perderían por una tilde.
    """
    import gzip as _gz
    p = (CRUDO / "conaf_usos_tierra_agricola" / "2026-08-25" /
         "terrenos_agricolas.geojson.gz")
    if not p.exists():
        return
    m = {"rotacion cultivo-pradera": ("852", "Rotación Cultivo-Pradera"),
         "terreno de uso agricola": ("397", "Terreno de Uso Agrícola"),
         "terrenos de uso agricola": ("397", "Terreno de Uso Agrícola")}
    fs = json.loads(_gz.open(p, "rt", encoding="utf-8").read())["features"]
    for f in fs:
        it = m.get(sin_tildes(f["properties"].get("subuso")))
        pt = punto_de(f.get("geometry"))
        if not it:
            DESCARTES["clase de uso de tierra sin ítem"].append(
                str(f["properties"].get("subuso"))[:40])
            continue
        if not pt or not dentro(pt[1], pt[0]):
            continue
        pr = f["properties"]
        yield {"item": it[0], "elemento": it[1],
               "nombre": "", "lat": round(pt[1], 6), "lon": round(pt[0], 6),
               "comuna": pr.get("nom_com", ""), "region": pr.get("nom_reg", ""),
               "fuente": "CONAF/CIREN · Catastro de usos de la tierra",
               "confianza": "consolidado"}


FUENTES += [
    ("antenas", f_antenas_torres, "antenas de la Ley de Torres"),
    ("hidrografia", f_hidrografia, "cursos de agua del IDE"),
    ("bovinos", f_bovinos, "planteles bovinos del Rol Único Pecuario"),
    ("fruticola", f_fruticola, "predios del catastro frutícola"),
    ("bocatomas", f_bocatomas, "bocatomas de riego de la CNR"),
    ("acuiferos", f_acuiferos, "acuíferos protegidos de la DGA"),
    ("agricola", f_agricola, "usos de tierra agrícola de CONAF"),
]


if __name__ == "__main__":
    print("=" * 78)
    print("INTEGRAR EL BARRIDO · del crudo de los agentes al índice de activos")
    print("=" * 78)
    sys.exit(main())
