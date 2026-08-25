"""
SUBIR LAS SUB-MATRICES A SHAREPOINT
=====================================

Crea las listas de sub-matriz en el sitio RMD y sube sus 96.423 activos, con la
columna de búsqueda `MICR` ya vinculada a la Matriz — que es justamente lo que
una importación de Excel NO puede hacer.

★ SOBRE LAS CREDENCIALES, QUE ES LO PRIMERO
---------------------------------------------
**Este script NO pide ni acepta contraseña, y no hay ningún secreto escrito en
él.** Usa el flujo de **código de dispositivo** de Microsoft: al arrancar imprime
un código, se abre una página de Microsoft, se pega el código y se aprueba con la
cuenta de siempre —con su doble factor si lo tiene—. El script recibe entonces un
token temporal.

Ese token se guarda **fuera del proyecto**, en `~/.cache/rmd_submatrices/`, con
permisos de sólo-dueño. No entra al repositorio, no queda en ningún archivo del
proyecto, y se puede borrar en cualquier momento con `--cerrar-sesion`.

El identificador de aplicación que se usa es el público de Microsoft para
herramientas de línea de comandos sobre Graph. No es un secreto: es el mismo que
usa la propia CLI de Microsoft.

CÓMO CORRERLO
-------------
    # 1 · autenticar (una vez; el token se renueva solo mientras dure)
    ../.venv-esa/bin/python subir_submatrices_sharepoint.py --login

    # 2 · crear las listas vacías con su esquema completo  (~3 minutos)
    ../.venv-esa/bin/python subir_submatrices_sharepoint.py --crear

    # 3 · subir las filas  (horas; se puede cortar y retomar)
    ../.venv-esa/bin/python subir_submatrices_sharepoint.py --subir

    # en cualquier momento, desde otra terminal:
    ../.venv-esa/bin/python subir_submatrices_sharepoint.py --estado
    ../.venv-esa/bin/python subir_submatrices_sharepoint.py --cerrar-sesion

`--subir` es **reanudable**: lleva un registro de cuántas filas van subidas por
lista y, si se corta —o si Microsoft empieza a limitar la tasa y conviene parar—,
al volver a lanzarlo sigue donde iba. No duplica.

QUÉ HACE, EN ORDEN
------------------
1. Lee las 835 filas de la lista `mic` y arma la tabla que traduce «número de
   ítem de la Matriz» → «identificador interno de SharePoint». Sin eso no se
   puede rellenar la columna de búsqueda.
2. Por cada archivo `NNN Descriptor.xlsx`, crea la lista con sus trece columnas
   y el vínculo a la Matriz.
3. Sube las filas en lotes de 20, que es el máximo que admite Graph.
4. Respeta la limitación de tasa: cuando el servidor responde «espera», espera
   exactamente lo que pide en vez de insistir.

LO QUE NO HACE
--------------
No toca la lista `mic`, ni las listas `120` y `centrales` que ya existen. Si una
lista de sub-matriz ya existe, la salta y lo dice: no sobrescribe nada.
"""

import json
import os
import re
import sys
import time
from pathlib import Path

import requests

# ── constantes del sitio (leídas de SharePoint, no inventadas) ───────────────
SITIO = ("rstchilecom.sharepoint.com,d14670c1-197a-4976-ac78-92eb6d9ccff6,"
         "24a0c429-ccba-4544-bd83-e831a7a5a1a8")
LISTA_MIC = "315db0cf-1a0b-42b5-a140-c5f4bb396d46"
GRAPH = "https://graph.microsoft.com/v1.0"

# Identificador público de «Microsoft Graph Command Line Tools». No es secreto.
CLIENTE = "14d82eec-204b-4c2f-b7e8-296a70dab67e"
AUTORIDAD = "https://login.microsoftonline.com/organizations"
# ★ `Sites.Manage.All` y no `Sites.ReadWrite.All`: comprobado el 21-ago que
# crear una lista devuelve 403 con el segundo. Escribir FILAS sí basta con
# ReadWrite; CREAR LISTAS exige Manage. `offline_access` es lo que entrega el
# token de renovación, sin el cual la sesión vence en una hora — inservible para
# una subida de horas.
AMBITO = ("https://graph.microsoft.com/Sites.Manage.All "
          "https://graph.microsoft.com/Sites.ReadWrite.All offline_access")

AQUI = Path(__file__).resolve().parent
ORIGEN = AQUI / "submatrices_excel"
# ★ Fuera del proyecto y fuera del repositorio, a propósito.
CACHE = Path.home() / ".cache" / "rmd_submatrices"
TOKEN = CACHE / "token.json"
AVANCE = CACHE / "avance.json"
BITACORA = CACHE / "bitacora.log"

LOTE = 20            # máximo que admite el punto de entrada por lotes de Graph
PAUSA = 0.15         # respiro entre lotes, para no provocar la limitación


def anotar(msg):
    linea = f"{time.strftime('%H:%M:%S')}  {msg}"
    print(linea, flush=True)
    CACHE.mkdir(parents=True, exist_ok=True)
    with BITACORA.open("a", encoding="utf8") as fh:
        fh.write(linea + "\n")


# ── autenticación por código de dispositivo ─────────────────────────────────

def login():
    """Pide un código, espera a que la persona lo apruebe, guarda el token."""
    r = requests.post(f"{AUTORIDAD}/oauth2/v2.0/devicecode",
                      data={"client_id": CLIENTE, "scope": AMBITO}, timeout=30)
    if r.status_code != 200:
        anotar(f"✗ Microsoft rechazó la petición de código: {r.status_code} "
               f"{r.text[:300]}")
        return False
    d = r.json()
    # ★ `anotar` fuerza la escritura y además deja el código en la bitácora.
    # Sin eso, al correr en segundo plano el búfer se queda con todo y no se ve
    # nada — que fue exactamente lo que pasó el primer intento.
    anotar("=" * 66)
    anotar("ABRE ESTA PÁGINA Y PEGA EL CÓDIGO")
    anotar(f"   página : {d['verification_uri']}")
    anotar(f"   código : {d['user_code']}")
    anotar("La contraseña se pone en la página de Microsoft, no acá.")
    anotar("=" * 66)
    limite = time.time() + d["expires_in"]
    while time.time() < limite:
        time.sleep(d.get("interval", 5))
        t = requests.post(f"{AUTORIDAD}/oauth2/v2.0/token",
                          data={"client_id": CLIENTE,
                                "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
                                "device_code": d["device_code"]}, timeout=30)
        j = t.json()
        if "access_token" in j:
            guardar_token(j)
            anotar("sesión iniciada · token guardado en " + str(TOKEN))
            # ★ Comprobar aquí lo que se concedió, y no descubrirlo con un 403
            # a mitad de camino, que fue lo que pasó el primer intento.
            dados = j.get("scope", "")
            anotar("   permisos concedidos: " +
                   ", ".join(x.split("/")[-1] for x in dados.split()))
            if "Sites.Manage.All" not in dados:
                anotar("   ⚠ FALTA Sites.Manage.All → NO se podrán CREAR listas "
                       "(sí subir filas a listas que ya existan)")
            if "refresh_token" not in j:
                anotar("   ⚠ SIN token de renovación: la sesión vence en ~1 h y "
                       "habrá que repetir --login a mitad de la subida")
            else:
                anotar("   ✓ con token de renovación: la sesión se mantiene sola")
            return True
        if j.get("error") not in ("authorization_pending", "slow_down"):
            print("  ✗", j.get("error_description", j))
            return False
        if j.get("error") == "slow_down":
            d["interval"] = d.get("interval", 5) + 5
    print("  ✗ el código expiró sin aprobarse")
    return False


def guardar_token(j):
    CACHE.mkdir(parents=True, exist_ok=True)
    j["expira_en"] = time.time() + j.get("expires_in", 3600) - 120
    TOKEN.write_text(json.dumps(j))
    os.chmod(TOKEN, 0o600)          # sólo el dueño puede leerlo


def token():
    """Devuelve un token válido, renovándolo en silencio si hace falta."""
    if not TOKEN.exists():
        print("No hay sesión. Corre primero:  --login")
        raise SystemExit(1)
    j = json.loads(TOKEN.read_text())
    if time.time() < j.get("expira_en", 0):
        return j["access_token"]
    if "refresh_token" not in j:
        print("La sesión venció. Corre otra vez:  --login")
        raise SystemExit(1)
    r = requests.post(f"{AUTORIDAD}/oauth2/v2.0/token",
                      data={"client_id": CLIENTE, "grant_type": "refresh_token",
                            "refresh_token": j["refresh_token"], "scope": AMBITO},
                      timeout=30)
    n = r.json()
    if "access_token" not in n:
        print("No se pudo renovar la sesión. Corre otra vez:  --login")
        raise SystemExit(1)
    guardar_token(n)
    return n["access_token"]


# ── llamadas a Graph, con respeto a la limitación de tasa ────────────────────

def llamar(metodo, url, cuerpo=None, reintentos=6):
    """Una llamada. Si el servidor pide esperar, se espera lo que pide."""
    for intento in range(reintentos):
        r = requests.request(metodo, url if url.startswith("http") else GRAPH + url,
                             headers={"Authorization": "Bearer " + token(),
                                      "Content-Type": "application/json"},
                             data=json.dumps(cuerpo) if cuerpo else None, timeout=120)
        if r.status_code in (429, 503, 504):
            espera = int(r.headers.get("Retry-After", 2 ** intento * 5))
            anotar(f"   el servidor pide esperar {espera}s (intento {intento+1})")
            time.sleep(espera)
            continue
        return r
    return r


# ── esquema de las sub-matrices, copiado de las listas 120 y centrales ───────
# `name` es el nombre interno (sin tildes, para que no lo codifique) y
# `displayName` es el que se ve. Así las columnas quedan predecibles.
COLUMNAS = [
    ("Numero", "Número", "number"), ("Item", "Ítem", "number"),
    ("Sector", "Sector", "text"), ("Elemento", "Elemento", "text"),
    ("Region", "Región", "text"), ("Provincia", "Provincia", "text"),
    ("Comuna", "Comuna", "text"), ("Direccion", "Dirección", "text"),
    ("Latitud", "Latitud", "text"), ("Longitud", "Longitud", "text"),
    ("Responsable", "Responsable", "text"), ("Telefono", "Teléfono", "text"),
    ("LatitudDecimal", "Latitud decimal", "number"),
    ("LongitudDecimal", "Longitud decimal", "number"),
]


def definicion_columnas():
    cols = []
    for nombre, visible, tipo in COLUMNAS:
        c = {"name": nombre, "displayName": visible}
        c["number" if tipo == "number" else "text"] = {}
        cols.append(c)
    cols.append({"name": "MICR", "displayName": "MICR",
                 "description": "Vínculo a la Matriz de Infraestructura Crítica",
                 "lookup": {"listId": LISTA_MIC, "columnName": "Title"}})
    return cols


def listas_existentes():
    r = llamar("GET", f"/sites/{SITIO}/lists?$select=id,name,displayName&$top=200")
    return {x["displayName"]: x["id"] for x in r.json().get("value", [])}


def mapa_micr():
    """número de ítem de la Matriz → identificador interno de SharePoint."""
    mapa, url = {}, (f"/sites/{SITIO}/lists/{LISTA_MIC}/items"
                     "?$select=id&$expand=fields($select=N_x00b0_)&$top=500")
    while url:
        r = llamar("GET", url).json()
        for it in r.get("value", []):
            n = it["fields"].get("N_x00b0_")
            if n is not None:
                mapa[int(n)] = it["id"]
        url = r.get("@odata.nextLink")
    return mapa


def archivos():
    """Los .xlsx de sub-matriz, ordenados por número de ítem."""
    out = []
    for p in sorted(ORIGEN.glob("*.xlsx")):
        m = re.match(r"^(\d+)\s+(.+)\.xlsx$", p.name)
        if m:
            out.append((int(m.group(1)), m.group(2), p))
    return sorted(out)


# ── fase 1: crear las listas ────────────────────────────────────────────────

def crear():
    ya = listas_existentes()
    anotar(f"listas en el sitio: {len(ya)}")
    creadas = saltadas = 0
    for item, desc, p in archivos():
        nombre = f"{item} {desc}"
        if nombre in ya:
            anotar(f"  = ya existe, se salta: {nombre}")
            saltadas += 1
            continue
        cuerpo = {"displayName": nombre,
                  "description": f"Sub-matriz del ítem {item} de la Matriz de "
                                 f"Infraestructura Crítica",
                  "list": {"template": "genericList"},
                  "columns": definicion_columnas()}
        r = llamar("POST", f"/sites/{SITIO}/lists", cuerpo)
        if r.status_code in (200, 201):
            anotar(f"  ✓ creada: {nombre}")
            creadas += 1
        else:
            anotar(f"  ✗ FALLÓ {nombre}: {r.status_code} "
                   f"{r.json().get('error',{}).get('message','')[:160]}")
    anotar(f"listas creadas {creadas} · ya existían {saltadas}")


# ── fase 2: subir las filas ─────────────────────────────────────────────────

def leer_avance():
    return json.loads(AVANCE.read_text()) if AVANCE.exists() else {}


def guardar_avance(a):
    CACHE.mkdir(parents=True, exist_ok=True)
    AVANCE.write_text(json.dumps(a, indent=1))


def campos_de(f, micr):
    """Una fila del Excel → los campos de un elemento de la lista."""
    c = {"Title": str(f[0] or "")[:255], "Numero": f[1], "Item": f[2],
         "Sector": str(f[3] or "")[:255], "Elemento": str(f[4] or "")[:255],
         "Region": str(f[5] or "")[:255], "Provincia": str(f[6] or "")[:255],
         "Comuna": str(f[7] or "")[:255], "Direccion": str(f[8] or "")[:255],
         "Latitud": str(f[9] or ""), "Longitud": str(f[10] or ""),
         "Responsable": str(f[11] or "")[:255], "Telefono": str(f[12] or "")}
    if f[13] not in (None, ""):
        c["LatitudDecimal"] = f[13]
        c["LongitudDecimal"] = f[14]
    if f[2] in micr:
        c["MICRLookupId"] = micr[f[2]]
    return c


def subir():
    import openpyxl
    ya = listas_existentes()
    micr = mapa_micr()
    anotar(f"ítems de la Matriz mapeados: {len(micr)}")
    avance = leer_avance()

    for item, desc, p in archivos():
        nombre = f"{item} {desc}"
        if nombre not in ya:
            anotar(f"  ✗ la lista «{nombre}» no existe todavía — corre --crear")
            continue
        lid = ya[nombre]
        ws = openpyxl.load_workbook(p, read_only=True).active
        filas = list(ws.iter_rows(values_only=True))[1:]
        hechas = avance.get(nombre, 0)
        if hechas >= len(filas):
            anotar(f"  = completa: {nombre} ({len(filas):,})")
            continue
        # ★ Si la lista viene a medias, NO confiar en el registro: preguntarle a
        # SharePoint qué números tiene de verdad y saltarse ésos. Sin esto, un
        # corte a mitad de un lote deja filas creadas que el registro no anotó, y
        # al reanudar se escriben DUPLICADAS — pasó el 21-ago con 9 filas del
        # tramo 841-860 de «17 Agua Potable Rural».
        presentes = set()
        if hechas > 0:
            u = (f"/sites/{SITIO}/lists/{lid}/items"
                 "?$select=id&$expand=fields($select=Numero)&$top=500")
            while u:
                jr = llamar("GET", u).json()
                for it in jr.get("value", []):
                    v = it["fields"].get("Numero")
                    if v is not None:
                        presentes.add(int(v))
                u = jr.get("@odata.nextLink")
            anotar(f"     ya en SharePoint: {len(presentes):,} filas "
                   f"(el registro decía {hechas:,})")
        anotar(f"  ▶ {nombre}: {len(filas)-hechas:,} filas por subir "
               f"(van {hechas:,} de {len(filas):,})")

        t0, inicio = time.time(), hechas
        while hechas < len(filas):
            trozo = [f for f in filas[hechas:hechas + LOTE]
                     if int(f[1]) not in presentes]
            if not trozo:                      # todas las de este tramo ya están
                hechas += min(LOTE, len(filas) - hechas)
                avance[nombre] = hechas
                continue
            pendientes = list(range(len(trozo)))   # índices dentro del trozo
            vuelta = 0
            while pendientes and vuelta < 8:
                peticiones = [{"id": str(k), "method": "POST",
                               "url": f"/sites/{SITIO}/lists/{lid}/items",
                               "headers": {"Content-Type": "application/json"},
                               "body": {"fields": campos_de(trozo[k], micr)}}
                              for k in pendientes]
                r = llamar("POST", "/$batch", {"requests": peticiones})
                respuestas = {x["id"]: x for x in r.json().get("responses", [])}
                # ★ AQUÍ ESTABA EL FALLO: una llamada por lotes devuelve 200 en el
                # conjunto AUNQUE las peticiones de adentro devuelvan 429. Antes se
                # daba el trozo por subido igual, y las filas rechazadas se perdían
                # en silencio. Ahora sólo se dan por hechas las que de verdad se
                # crearon, y las demás se reintentan.
                fallaron, espera = [], 0
                for k in pendientes:
                    resp = respuestas.get(str(k), {})
                    if resp.get("status", 500) >= 300:
                        fallaron.append(k)
                        espera = max(espera, int(resp.get("headers", {})
                                                 .get("Retry-After", 10)))
                pendientes = fallaron
                if pendientes:
                    vuelta += 1
                    anotar(f"     {len(pendientes)} de {len(trozo)} rechazadas · "
                           f"espero {espera}s y reintento (vuelta {vuelta})")
                    time.sleep(espera)
                    # ★ el servidor está pidiendo aire: bajar el ritmo de verdad
                    globals()["PAUSA"] = min(PAUSA * 1.6 + 0.2, 8.0)
            if pendientes:
                anotar(f"     ✗ {len(pendientes)} filas no entraron tras 8 vueltas "
                       f"· quedan para --reparar")
            elif vuelta == 0:
                # ★ La pausa TIENE que saber bajar. En la corrida del 21-ago subió
                # a su tope de 8 s por una ráfaga de rechazos y se quedó ahí una
                # hora DESPUÉS de que la limitación cesara: el ritmo cayó de 11 a
                # 2,1 filas/s sin que el servidor lo pidiera. Un trinquete que
                # sólo sube no es prudencia, es lentitud gratis.
                globals()["PAUSA"] = max(PAUSA * 0.85, 0.15)
            hechas += min(LOTE, len(filas) - hechas)
            avance[nombre] = hechas
            guardar_avance(avance)
            if hechas % 500 < LOTE:
                # ★ velocidad de ESTA corrida: antes dividía el total acumulado
                # por el tiempo transcurrido y salían cifras infladas al reanudar
                v = (hechas - inicio) / max(time.time() - t0, 1)
                anotar(f"     {hechas:,}/{len(filas):,} · {v:.1f} filas/s · "
                       f"pausa {PAUSA:.2f}s · faltan ~"
                       f"{(len(filas)-hechas)/max(v,0.1)/60:.0f} min")
            time.sleep(PAUSA)
        anotar(f"  ✓ completa: {nombre} ({len(filas):,})")
    anotar("SUBIDA TERMINADA")


def reparar():
    """Reconcilia cada lista contra su Excel y sube SÓLO lo que falta.

    ★ Es la red de seguridad, y existe porque hizo falta: el 21-ago se
    detectaron 208 filas perdidas en silencio (una llamada por lotes devuelve
    200 aunque las peticiones de adentro devuelvan 429). Aunque ese fallo ya
    está corregido, este modo comprueba el resultado contra el origen en vez de
    confiar en el registro de avance — que es lo que había fallado.

    La comprobación es por la columna `Número`, que es un correlativo único
    dentro de cada archivo. Lo que esté en el Excel y no en la lista, se sube.
    """
    import openpyxl
    ya = listas_existentes()
    micr = mapa_micr()
    avance = leer_avance()
    total_faltantes = 0
    for item, desc, p in archivos():
        nombre = f"{item} {desc}"
        if nombre not in ya:
            continue
        lid = ya[nombre]
        filas = list(openpyxl.load_workbook(p, read_only=True).active
                     .iter_rows(values_only=True))[1:]
        if avance.get(nombre, 0) == 0:
            continue                      # ni empezada: la sube --subir
        # qué números están de verdad en SharePoint
        presentes, url = set(), (f"/sites/{SITIO}/lists/{lid}/items"
                                 "?$select=id&$expand=fields($select=Numero)&$top=500")
        while url:
            r = llamar("GET", url).json()
            for it in r.get("value", []):
                n = it["fields"].get("Numero")
                if n is not None:
                    presentes.add(int(n))
            url = r.get("@odata.nextLink")
        # ★ duplicados: mismo Número más de una vez. Se conserva el primero.
        from collections import defaultdict as _dd
        porn = _dd(list)
        u2 = (f"/sites/{SITIO}/lists/{lid}/items"
              "?$select=id&$expand=fields($select=Numero)&$top=500")
        while u2:
            jr = llamar("GET", u2).json()
            for it in jr.get("value", []):
                v = it["fields"].get("Numero")
                if v is not None:
                    porn[int(v)].append(int(it["id"]))
            u2 = jr.get("@odata.nextLink")
        sobran = [i for ids in porn.values() if len(ids) > 1
                  for i in sorted(ids)[1:]]
        if sobran:
            anotar(f"  ⚠ {nombre}: {len(sobran)} filas DUPLICADAS · borrando")
            for i in sobran:
                llamar("DELETE", f"/sites/{SITIO}/lists/{lid}/items/{i}")
        faltan = [f for f in filas[:avance.get(nombre, 0)]
                  if int(f[1]) not in presentes]
        if not faltan:
            anotar(f"  ✓ íntegra: {nombre} ({len(presentes):,} de "
                   f"{avance.get(nombre,0):,} revisadas)")
            continue
        anotar(f"  ▶ {nombre}: faltan {len(faltan):,} filas · reponiendo")
        total_faltantes += len(faltan)
        i = 0
        while i < len(faltan):
            trozo = faltan[i:i + LOTE]
            pendientes = list(range(len(trozo)))
            vuelta = 0
            while pendientes and vuelta < 8:
                peticiones = [{"id": str(k), "method": "POST",
                               "url": f"/sites/{SITIO}/lists/{lid}/items",
                               "headers": {"Content-Type": "application/json"},
                               "body": {"fields": campos_de(trozo[k], micr)}}
                              for k in pendientes]
                r = llamar("POST", "/$batch", {"requests": peticiones})
                resp = {x["id"]: x for x in r.json().get("responses", [])}
                fallaron, espera = [], 0
                for k in pendientes:
                    x = resp.get(str(k), {})
                    if x.get("status", 500) >= 300:
                        fallaron.append(k)
                        espera = max(espera, int(x.get("headers", {})
                                                 .get("Retry-After", 10)))
                pendientes = fallaron
                if pendientes:
                    vuelta += 1
                    time.sleep(espera)
            i += len(trozo)
            time.sleep(PAUSA)
        anotar(f"  ✓ repuestas: {nombre}")
    anotar(f"REPARACIÓN TERMINADA · filas repuestas: {total_faltantes:,}")


def estado():
    avance = leer_avance()
    print(f"{'lista':52s} {'subidas':>9s} {'total':>9s}  avance")
    print("-" * 92)
    import openpyxl
    th = tt = 0
    for item, desc, p in archivos():
        nombre = f"{item} {desc}"
        ws = openpyxl.load_workbook(p, read_only=True).active
        n = ws.max_row - 1
        h = avance.get(nombre, 0)
        th += h
        tt += n
        barra = "█" * int(20 * h / max(n, 1))
        print(f"{nombre:52s} {h:9,d} {n:9,d}  {barra:<20s} {100*h/max(n,1):5.1f} %")
    print("-" * 92)
    print(f"{'TOTAL':52s} {th:9,d} {tt:9,d}  {100*th/max(tt,1):5.1f} %")


def cerrar_sesion():
    for f in (TOKEN,):
        if f.exists():
            f.unlink()
            print("borrado:", f)
    print("sesión cerrada")


if __name__ == "__main__":
    acciones = {"--login": login, "--crear": crear, "--subir": subir,
                "--reparar": reparar, "--estado": estado,
                "--cerrar-sesion": cerrar_sesion}
    elegida = next((a for a in sys.argv[1:] if a in acciones), None)
    if not elegida:
        print(__doc__)
        raise SystemExit(1)
    acciones[elegida]()
