"""
Construye el Excel de Variables y Métricas RMD 2.0 acotado a ESTE proyecto:
INFRAESTRUCTURA CRÍTICA × CLIMA (Chile).

QUÉ HACE
--------
Toma el catálogo maestro de 318 variables y métricas del RMD 2.0, selecciona
únicamente las que sirven a este proyecto, y les agrega tres columnas propias
(rol en el proyecto, dato real que las alimenta, estado). Suma además tres
hojas con material que el catálogo de 318 NO contiene pero el proyecto
necesita: las columnas de la Matriz de Infraestructura Crítica (MICR), el
módulo de colapso sistémico (MCSGS/ICSGS), y la hoja de hallazgos.

PRECEDENCIA DE FUENTES (regla dada por Alexis, 15-ago-2026)
-----------------------------------------------------------
El documento Word `RMD_2_Variables_y_METRICAS_COMPLETAS-11-06-2026.docx` manda
sobre el Excel `Variables-y-Metricas-318-06-03-2026-Tabla.xlsx`. El Excel es
la versión operativa y es lo práctico, pero donde discrepan gana el Word.
Por eso cada fila lleva DOS columnas de fórmula: la del Word (canónica) y la
del Excel (operativa), para que la discrepancia se vea en vez de esconderse.

NO MODIFICA NADA de los archivos originales: sólo lee. Todo lo que produce
queda en esta carpeta.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_RMD = "/Users/alexis/Desktop/Go en Conflictos/RMD 2.0"
CATALOGO = f"{BASE_RMD}/Variables-y-Metricas-318-06-03-2026-Tabla.xlsx"
SALIDA = ("/Users/alexis/Desktop/RMD/Cosmolab/Cosmoclima/infraestructura/"
          "Variables_y_Metricas_Infraestructura_Critica_Clima.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# SELECCIÓN: qué del catálogo de 318 sirve a este proyecto, y por qué.
# La clave es el N° del catálogo maestro, para que la trazabilidad al original
# sea directa y verificable. El orden acá es el orden en que se leerán.
#   nucleo=True  → sin esta variable el proyecto no se sostiene
#   formula_word → fórmula canónica extraída de las ecuaciones OMML del Word
#                  (vacío = el Word no trae fórmula propia para esa fila)
# ─────────────────────────────────────────────────────────────────────────────
SELECCION = [
    # ── El traductor clima→RMD. Es literalmente el objeto del proyecto. ──
    (304, True, "ZBG_i = ⟨T_ref, P_ref, S_ref, R_eco, V_hid⟩",
     "Marco de normalización. Sin ZBG no hay anomalía, sólo milímetros sueltos: "
     "define qué es 'normal' en cada zona antes de medir nada.",
     "Cosmoclima trabaja en anomalías por zona desde el inicio. Falta declarar "
     "las ZBG de Chile (el instrumento hoy cubre una sola).", "A construir"),
    (305, True,
     "EstCicMacClim = (w_ENSO·ENSO_N + w_SAM·SAM_N + w_PDO·PDO_N + w_MJO·MJO_N)·F_Tele·F_Est·F_Conf",
     "Estado macro del Pacífico. Fija el 'viento de fondo' con el que se "
     "interpreta cualquier anomalía local.",
     "ONI ya integrado en Cosmoclima (bandas El Niño/La Niña, criterio oficial "
     "NOAA ≥5 temporadas). Faltan SAM, PDO y MJO.", "Parcial"),
    (306, False, "ANTermic = |(T_obs − T_ref(ZBG)) / σ_T(ZBG)| · F_Cic · F_Temp",
     "Anomalía térmica. Para infraestructura importa menos que la lluvia, pero "
     "entra en EstHidric y RIncFor.",
     "ERA5 y NASA POWER ya descargados para la zona del instrumento.",
     "Operativa (zona acotada)"),
    (307, True, "ANPrecip = |(P_obs − P_ref(ZBG)) / σ_P(ZBG)| · F_Cic · F_Est",
     "★ NÚCLEO. La lluvia anómala es el forzante que cortó los caminos en "
     "julio-agosto 2026. Es la entrada principal del proyecto.",
     "Serie 1966-2026 sin reanálisis, validada contra estación medida "
     "(ronda 17). Cobertura: Norte Chico. Ver HALLAZGO H-07 sobre el valor "
     "absoluto.", "Operativa (zona acotada)"),
    (308, True,
     "EstHidric = (w_P·ANPrecip + w_T·ANTermic + w_D·D_uso)·F_Cic·F_Res",
     "Estrés hídrico. Para infraestructura vial actúa por el lado del suelo: "
     "un suelo saturado responde distinto que uno seco a la misma lluvia.",
     "Humedad de suelo ESA CCI (satélite, 1988-2024, máximo mensual) bajada "
     "por OPeNDAP del CEDA. D_uso falta (DGA).", "Parcial"),
    (309, True,
     "InEvExtre = (Σ_j w_j·E_j)·FEX·FET_clim × 100",
     "★ NÚCLEO. Los subíndices EOP (lluvia intensa) y EAL (aluvión) son "
     "exactamente el evento que corta rutas. El propio RMD define EAL como "
     "'precipitación intensa + condición de suelo' — las dos las tenemos.",
     "Lluvia diaria + humedad de suelo. FEX (exposición) debe salir de la "
     "MICR local, que es el puente formal con la matriz.", "A construir"),
    (310, False,
     "RIncFor = (w_H·EstHidric + w_T·ANTermic + w_V·V_comb + w_W·V_viento)·F_Cic·F_Terr",
     "Incendio forestal. No es el evento de esta ronda (lluvias), pero comparte "
     "insumos y es el otro modo de falla de la misma infraestructura.",
     "EstHidric y ANTermic disponibles. Faltan V_comb y V_viento.", "Parcial"),
    (311, False,
     "PrCost = (w_M·M_ext + w_E·E_cost + w_I·I_litt + w_P·P_exp)·F_Cic·F_TerrCost",
     "Presión costera. Aplica a puertos y borde costero — que en la MICR son "
     "el elemento de Transporte con PF más alto (0,75).",
     "Sin datos aún (SHOA). Se declara N/A en territorio no costero, nunca "
     "cero.", "No iniciada"),
    (312, True,
     "InClimCo = w_T·f(ANTermic) + w_P·f(ANPrecip) + w_H·f(EstHidric) + "
     "w_E·f(InEvExtre) + w_F·f(RIncFor) + w_C·f(PrCost),  con Σw=1",
     "Síntesis climática en una cifra transportable. Es el insumo que entra "
     "al MACC.",
     "Se calcula a partir de las anteriores; no necesita dato nuevo.",
     "A construir"),
    (313, True, "MatCoefClim = { C_ij | i ∈ Variables/Métricas RMD, j ∈ Dimensiones Climáticas }",
     "★ NÚCLEO. El enchufe. Coeficientes 0,8–1,6 aplicados DESPUÉS de medir, "
     "sin reescribir ninguna fórmula del RMD. Acá se conecta el clima con la "
     "matriz de infraestructura.",
     "No es dato: es la tabla de reglas a construir. Ver HALLAZGO H-05: su "
     "Listado Blanco no cuadra con el catálogo.", "A construir"),

    # ── Infraestructura: lo que ya existe en el catálogo y aplica ──
    (109, True, "IDInf = (Fallos Sistémicos / Total Eventos Infraestructurales) × 100 × FDS",
     "★ NÚCLEO. Es el puente formal entre 'se cortó el camino' y 'la gente se "
     "enoja'. Sin esta variable, el proyecto mide daño pero no conflicto, y "
     "deja de ser RMD.",
     "Prensa + SENAPRED + reportes de Vialidad. FDS por encuestas/redes.",
     "A construir"),
    (234, True, "IVIC = (D_Inf · V_Sys) · (1 − R_Def + I_Crit)",
     "★ Vulnerabilidad de infraestructura crítica. Convive con el FVT de la "
     "MICR: hay que decidir cuál manda (ver HALLAZGO H-06).",
     "I_Crit sale de PF/IRMD de la MICR.", "A construir"),
    (246, True, "IDIn = (I_Crit · S_Dañ) · (P_Dep · V_Pob)",
     "★ Daño infraestructural consumado. Es la medición del hecho: cuánta "
     "infraestructura crítica quedó fuera de servicio.",
     "I_Crit basado en PF e IRMD de la MICR (enganche ya escrito en el canon).",
     "A construir"),
    (250, True, "IDCE = PF · (P_Afect · V_Pob) · W_IRMD",
     "★★ La pieza más importante del catálogo para este proyecto: es la ÚNICA "
     "fórmula del RMD que usa el PF de la MICR directamente. El enganche "
     "formal entre matriz y modelo ya está escrito acá.",
     "PF viene de la matriz de 835 ítems; P_Afect de TD e IVC.", "A construir"),
    (210, True, "IRL = Σ(w_i·S_i) · (1 − V_Inf + R_Sup − I_San)",
     "★ Resiliencia logística = capacidad de sostener el FLUJO. Es la bisagra "
     "con el MCSGS, donde el colapso no es destrucción sino interdicción.",
     "Inventario logístico + rutas alternativas. Ver HALLAZGO H-04 (el Excel "
     "transcribió mal esta fórmula).", "A construir"),
    (242, False, "TD = P_Afect · (%D_Viol + %D_Cris)",
     "Desplazamiento. Un pueblo aislado por un camino cortado desplaza "
     "población: es el efecto humano medible del corte.",
     "Censo + reportes de emergencia. Ver HALLAZGO H-04.", "A construir"),
    (241, False, "IVC = P_Exp · I_Acc · S_Cri · V_Pob",
     "Víctimas colaterales. Cota superior del daño humano del evento.",
     "SENAPRED, Salud.", "No iniciada"),
    (249, False, "IIE = I_Dañ · C_Econ · P_Afect · F_Pob",
     "Impacto económico del daño infraestructural.",
     "MOP (costo de reposición), actividad económica interrumpida.",
     "No iniciada"),
    (226, False, "IAIC = (N_Atq / T_Per) · Imp_Atq · C_ICS",
     "Control, no medición: sirve para NO confundir un camino cortado por "
     "aluvión con uno cortado por sabotaje. El RMD exige esa distinción.",
     "GTD, Fiscalía, prensa.", "Control"),
    (317, True, "ICSat_X = Demanda Servicios / Capacidad Servicios  (normalizar 0–1)",
     "★ Saturación de servicios. Es exactamente el mecanismo de SATURACIÓN que "
     "el MCSGS describe para los nodos latentes: la ruta alternativa recibe "
     "todo el flujo y se convierte en el nuevo cuello de botella.",
     "Capacidad vs demanda por servicio y comuna.", "A construir"),
    (318, True, "IDSE_X = (Interrupciones Servicios / Período Analizado) · Factor Severidad",
     "★ Disrupción de servicios esenciales. Es la medición directa del corte: "
     "cuántas veces y con qué severidad se quedó sin agua, luz o camino.",
     "SEC (electricidad), SISS (agua), Vialidad (rutas cortadas).",
     "A construir"),
    (315, False, "VT/FVT = 1 − (Demanda Servicios / Capacidad Servicios)  (normalizar 0–1)",
     "Capacidad de absorción. ATENCIÓN: comparte sigla con VT y FVT de la MICR "
     "y NO es lo mismo. Ver HALLAZGO H-03.",
     "Igual que ICSat_X (es su complemento).", "A construir"),
    (110, False,
     "Bext(t) = (Acumulación de Riesgo Temporal / Escala Temporal Total) × 100 × FET",
     "Riesgo de colapso acumulado en el tiempo. Sirve para la pregunta que "
     "importa: no un temporal, sino qué pasa cuando se repiten.",
     "Serie histórica de eventos; el instrumento ya corre 62 años día a día.",
     "A construir"),
    (19, False, "IRS = 100 − (Individuos Afectados Negativamente / Población Total) × 100 × FRS",
     "Resiliencia social: capacidad de recuperación de la comunidad afectada. "
     "Modula cuánto de un corte se convierte en conflicto.",
     "Encuestas, registros comunales.", "No iniciada"),
    (10, False, "ICS = f(IAMV, IAH, IT) / 100 · Factor Fragilidad Social",
     "Conflictividad social. Es la ÚNICA variable del Listado Blanco de MACC "
     "que existe en el catálogo con el mismo nombre (ver HALLAZGO H-05).",
     "Ya operativa en el RMD base.", "Operativa"),
    (60, False, "IIECo = (Costos Conflictos / PIB) · Factor Impacto",
     "Impacto económico del conflicto derivado del fallo de infraestructura.",
     "Banco Central, Hacienda.", "No iniciada"),
    (198, False,
     "FC = w1·FC_Histórico + w2·Σ FC_Solar,i + w3·FC_Planetario + w4·IIP",
     "Factor Cíclico de METECO. Convive con EstCicMacClim de MACLIMA: ambos "
     "son 'ciclos'. Hay que declarar cuál se usa para no contar dos veces.",
     "Series históricas.", "A revisar (posible solapamiento)"),
    (199, False, "IHS = Índice Cooperación Social · (1 − FC_Solar − ICE + IACN)",
     "Homeostasis social: si el sistema social vuelve a su punto de "
     "equilibrio después del golpe.",
     "Encuestas.", "No iniciada"),
    (206, False, "CAMO = Σ(w_i·A_i) · (1 − D_Log + E_Tec − F_San)",
     "Sólo como REFERENCIA: es la ponderación 0,25 que la MICR declara usar "
     "sobre FANC y VT. No se calcula en este proyecto.",
     "IISS Military Balance. No aplica a lluvias.", "Referencia"),
    (214, False, "PNT = (w_PNO·PNO + w_PNS·PNS) · (1 − F_Int + G_Geo)",
     "Sólo como REFERENCIA: ponderación 0,1 que la MICR declara usar sobre "
     "IB y PF. No se calcula en este proyecto.",
     "GFP Global Power Index. No aplica a lluvias.", "Referencia"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Columnas de la MICR. NO están en el catálogo de 318: la matriz vive aparte,
# en su propio Excel de 835 filas. Se documentan acá para que el proyecto tenga
# todo en un solo lugar.
# Fuente: «Matriz de Infraestructura Crítica - FINAL.docx» + sección 22 del Word.
# ─────────────────────────────────────────────────────────────────────────────
MICR_COLUMNAS = [
    ("FEN", "Fragilidad ante Eventos Naturales",
     "★ CUATRO niveles desde el 16-ago-2026: Muy Alta=4 / Alta=3 / "
     "Moderada=2 / Baja=1  (antes eran tres)",
     "Cualitativa 1-4",
     "★★ ES EL ÚNICO EJE CLIMÁTICO DE TODA LA MATRIZ, y hoy es una etiqueta "
     "estática: sin territorio y sin tiempo. Una autopista tiene FEN=Alta en "
     "Arica y en Aysén, en año seco y en año de temporal. Convertir FEN en "
     "función del lugar y del mes es la propuesta central del proyecto. "
     "SERNAGEOMIN ya publica exactamente esta escala, pero viva y por zona: es "
     "la prueba de que la propuesta no es especulativa.",
     "Sí — es el punto de inserción de MACLIMA"),
    ("FANC", "Fragilidad ante Ataques No Convencionales", "Alta=3 / Media=2 / Baja=1",
     "Cualitativa 1-3",
     "No aplica al forzante climático. Se conserva sin tocar para no romper "
     "FVT ni las prioridades Pev/Peh.", "No"),
    ("IB", "Importancia Base", "Asignada por criterio experto", "0 a 1",
     "Relevancia del elemento para la estabilidad. No la toca el clima: una "
     "carretera no es más importante porque llueva. Se conserva.", "No"),
    ("VT", "Vulnerabilidad Tecnológica", "Asignada por criterio experto", "0 a 1",
     "Dependencia tecnológica. OJO: comparte sigla con la variable 315 del "
     "catálogo, que es otra cosa (ver HALLAZGO H-03).", "No"),
    ("FVT", "Factor de Vulnerabilidad Total", "FVT = (FEN + FANC + VT) / 3, normalizado",
     "0 a 1",
     "★ Compuesto de vulnerabilidad. Si FEN pasa a ser dinámico, FVT lo hereda "
     "y con él PF, IRMD, Pev, Peh y Pen: toda la cadena se mueve sola. "
     "PERO ver HALLAZGO H-01: la fórmula publicada NO reproduce los datos.",
     "Sí — por herencia de FEN"),
    ("PF", "Ponderación Final", "PF = IB × FVT", "0 a 1",
     "Prioridad del elemento. Verificado contra las 835 filas: se cumple "
     "(error medio 0,0023, puro redondeo). Es el enganche con IDCE, la única "
     "fórmula del catálogo que usa PF directamente.",
     "Sí — por herencia de FVT"),
    ("IRMD", "Impacto en el RMD", "Alto si PF>0,5; Medio 0,3–0,5; Bajo <0,3",
     "Cualitativa",
     "Clasificación del impacto en la dinámica de conflicto. Ver HALLAZGO "
     "H-02: las bandas se solapan en los datos reales.",
     "Sí — por herencia de PF"),
    ("Pev", "Prioridad Estratégica · Conflicto Vertical y Regular",
     "PE_vertical = 0,5·IB + 0,3·FANC_num + 0,2·FVT ; normalizado por máx≈1,61",
     "Muy Alta / Alta / Media / Baja / Muy Baja",
     "Guerra convencional. No es el escenario de este proyecto, pero hereda "
     "FVT y por tanto se movería si FEN cambia. Peso de FEN vía FVT: 0,2.",
     "Indirecto"),
    ("Peh", "Prioridad Estratégica · Conflicto Horizontal e Irregular",
     "PE_horizontal = 0,5·FANC_num + 0,3·VT + 0,2·FVT ; normalizado por máx≈1,94",
     "Muy Alta / Alta / Media / Baja / Muy Baja",
     "Conflicto asimétrico/ciber. No es el escenario de este proyecto.",
     "Indirecto"),
    ("Pen", "Prioridad Estratégica · Desastres Naturales",
     "PE_natural = 0,5·FEN_num + 0,3·IB + 0,2·FVT ; normalizado por máx≈1,87",
     "Muy Alta / Alta / Media / Baja / Muy Baja",
     "★★ ES LA COLUMNA DEL PROYECTO. FEN entra con peso 0,5 directo MÁS 0,2 "
     "indirecto vía FVT: el 60% de Pen depende de FEN. Volver FEN dinámico "
     "convierte a Pen en un ranking que cambia con el mes y el lugar — que es "
     "exactamente lo que un planificador de emergencias necesita.",
     "Sí — ES EL OBJETIVO"),
]

# ─────────────────────────────────────────────────────────────────────────────
# MCSGS — Módulo de Colapso Sistémico Global Sincronizado (sección 23 del Word).
# Tampoco está en el catálogo de 318 (es posterior al Excel del 06-03-2026).
# Aporta lo que faltaba: por qué una ruta secundaria puede volverse crítica.
# ─────────────────────────────────────────────────────────────────────────────
MCSGS_ITEMS = [
    ("ICSGS", "Índice de Colapso Sistémico Global Sincronizado",
     "ICSGS = mín(100 ; √(FCN × FSS × FAS × FPI) × (1/FRC) × 100)", "0 a 100",
     "Métrica de colapso FUNCIONAL, no de destrucción. Mide si el sistema "
     "puede seguir funcionando, no cuánto se rompió. Escala: 1-10% tensión, "
     "11-25% disrupción localizada, 26-45% regional, 46-65% estrés alto, "
     "66-80% pre-colapso, 81-95% colapso funcional, 100% colapso global."),
    ("CSS", "Condición de Sincronización Sistémica", "Verificación previa (0 o >0)",
     "Binaria",
     "Compuerta obligatoria: si no hay dos o más nodos críticos "
     "interdependientes afectados de forma simultánea o encadenada, ICSGS = 0 "
     "y no se calcula nada. Evita inflar el índice con un solo corte."),
    ("US", "Unidad de Sistema", "Declaración del analista", "Nacional / Regional / Global",
     "Hay que declarar sobre qué sistema se mide. Para este proyecto: "
     "US-Regional (una región de Chile) o US-Nacional."),
    ("FCN", "Factor de Criticidad Nodal", "Derivado de PF + IRMD de la MICR", "0 a 1",
     "★ ENGANCHE DIRECTO CON LA MATRIZ: el canon dice explícitamente que FCN "
     "sale de PF e IRMD. La matriz de 835 ítems alimenta el módulo de colapso "
     "sin intermediarios."),
    ("FSS", "Factor de Sincronización Sistémica", "Simultaneidad del daño", "0 a 1",
     "★ Acá entra el clima de lleno: un temporal golpea MUCHOS nodos A LA VEZ. "
     "La sincronización que un atacante tendría que planificar, la lluvia la "
     "produce gratis."),
    ("FAS", "Factor de Acoplamiento Sistémico", "Dependencia entre sistemas", "0 a 1",
     "Cuánto depende un sistema de otro. Sin camino no llega el combustible; "
     "sin combustible no opera el generador; sin generador no hay agua."),
    ("FRC", "Factor de Resiliencia del Sistema", "Capacidad de absorción", "0 a 1",
     "Divide, no se suma: la resiliencia mitiga pero no compensa. Se alimenta "
     "de IRL (#210) y de ICSat_X (#317)."),
    ("FPI", "Factor de Propagación Inter-sistemas", "Velocidad de transmisión del shock",
     "0 a 1",
     "Qué tan rápido se propaga la falla. En un aluvión, cuestión de horas."),
    ("NGF", "Nodo Geoestratégico de Flujo", "Clasificación cualitativa", "Categórica",
     "Punto de paso obligado donde converge el flujo, sin alternativa barata en "
     "el corto plazo. Un puente único sobre un río es un NGF, aunque su PF sea "
     "modesto. La MICR clasifica por importancia; NGF clasifica por paso."),
    ("NGF-L", "Nodo Geoestratégico de Flujo Latente", "Activación por reconfiguración",
     "Categórica",
     "★★ LA PIEZA QUE FALTABA. Un nodo que no es crítico en condiciones "
     "normales y se vuelve crítico cuando cae el principal y el flujo se "
     "redirige hacia él. Es exactamente la ruta rural secundaria: PF=0,41 en "
     "la matriz (última del sector Transporte), pero cuando la autopista se "
     "corta por aluvión, pasa a ser el único acceso — y se satura. El RMD ya "
     "tenía el concepto; nadie lo había conectado con el clima."),
]

# ─────────────────────────────────────────────────────────────────────────────
# HALLAZGOS. Alexis pidió (15-ago-2026) dejarlos ANOTADOS, no corregidos:
# «sólo haremos modificaciones a la Matriz o al Módulo contra datos reales».
# Cada uno indica cómo fue comprobado, para que sea auditable.
# ─────────────────────────────────────────────────────────────────────────────
HALLAZGOS = [
    ("H-01", "MICR", "Alta",
     "La fórmula publicada de FVT no reproduce los datos de la matriz",
     "El canon dice FVT = (FEN + FANC + VT)/3 normalizado. Aplicada a las 835 "
     "filas da error medio 0,108 y máximo 0,27. El propio documento lo admite "
     "en su ejemplo: calcula 0,933 para Presas y la tabla dice 0,87. Un ajuste "
     "lineal por mínimos cuadrados baja el error a 0,027 pero tampoco calza "
     "exacto (máx 0,16).",
     "Se aplicó la fórmula a las 835 filas del Excel HOMOLOGADO y se comparó "
     "celda a celda con la columna FVT publicada.",
     "Si MACC va a multiplicar FVT por un coeficiente climático, la base tiene "
     "que ser reproducible: si no, el efecto del ajuste no se puede auditar ni "
     "distinguir del error de base.",
     "Pedir la regla real que generó la columna, o recalcular FVT desde cero "
     "con la fórmula canónica y ver qué se mueve. NO TOCAR hasta decidirlo."),
    ("H-02", "MICR", "Alta",
     "Las bandas de IRMD se solapan: el mismo PF recibe dos etiquetas distintas",
     "El canon fija Alto si PF>0,5; Medio entre 0,3 y 0,5. En los datos "
     "reales, 'Alto' arranca en PF=0,38 y 'Medio' llega hasta 0,49. Hay 278 "
     "elementos en la franja 0,38–0,49, y ahí conviven 216 marcados Alto y 62 "
     "marcados Medio, con el mismo PF.",
     "Se clasificaron las 835 filas según los umbrales del canon y se "
     "contrastaron con la columna IRMD publicada: 229 discrepancias.",
     "IRMD hoy no es una función de PF; hay criterio no escrito. Cualquier "
     "coeficiente climático que actúe vía PF llegará a IRMD de forma "
     "impredecible.",
     "Documentar el criterio real, o adoptar el umbral empírico (≈0,38). "
     "NO TOCAR hasta decidirlo."),
    ("H-03", "Catálogo", "Media",
     "Colisión de siglas: VT y FVT significan dos cosas distintas",
     "En la MICR, VT es 'Vulnerabilidad Tecnológica' y FVT 'Factor de "
     "Vulnerabilidad Total'. En el catálogo de 318, la fila 315 es "
     "'VT/FVT — Vulnerabilidad Técnica / Funcional (Capacidad de Absorción)', "
     "categoría Demográfica/Social, con fórmula 1 − Demanda/Capacidad. No "
     "tienen nada que ver.",
     "Comparación directa entre la fila 315 del catálogo y las columnas de la "
     "matriz.",
     "Riesgo real de que un analista o un script mezcle ambas y produzca un "
     "número sin sentido.",
     "En este proyecto se usan los nombres largos, nunca la sigla sola. "
     "Proponer renombre al RMD, sin aplicarlo."),
    ("H-04", "Catálogo", "Media",
     "El Excel operativo transcribió sumas como comas en al menos 11 fórmulas",
     "Donde el Word dice (1 − V_Inf + R_Sup − I_San), el Excel dice "
     "([1]-[V_Inf],[R_Sup]-[I_San]). El '+' se volvió coma. Afecta, entre "
     "otras, a IRL (#210), IVIC (#234), TD (#242), ICO (#243), TH (#244), "
     "ICMed (#219), IVD (#224), IPCIB (#228), IIC (#233), ICCNC (#236), "
     "ICINC (#237). Dos de ellas (IVIC y TD) están en este proyecto.",
     "Se extrajeron las ecuaciones OMML del Word y se compararon con la "
     "columna 'Fórmula SharePoint' del Excel de 318.",
     "Confirma la regla de precedencia: el Word manda. Una fórmula con coma "
     "en lugar de suma no calcula.",
     "Este Excel de proyecto trae AMBAS columnas (Word y Excel) para que la "
     "discrepancia quede a la vista. Corregir el maestro es decisión del RMD."),
    ("H-05", "MACLIMA", "Alta",
     "El Listado Blanco de MACC nombra variables que no existen o significan otra cosa",
     "De las 17 variables que MACC declara elegibles para ajuste climático, "
     "sólo 3 existen en el catálogo de 318 con ese nombre: ICS, ICR e IRDE. "
     "Cinco no tienen entrada alguna (IIEC, IVP, IDE, IOC, IPT). Nueve existen "
     "con la MISMA SIGLA pero son otra variable: IPS es 'Polarización Social' "
     "y no 'Presión Social'; IAH es 'Ánimo Hostil' y no 'Agotamiento Humano'; "
     "IVT es 'Vanguardias Transhistóricas' y no 'Vulnerabilidad Territorial'; "
     "IRT es 'Resonancia Transnacional' y no 'Riesgo Territorial'; IVD es "
     "'Vulnerabilidad Diplomática' y no 'Violencia Dispersa'; IVC es 'Víctimas "
     "Colaterales' y no 'Vulnerabilidad Colectiva'; IDCE es 'Daño Colateral "
     "Emergente' y no 'Daño Colectivo Ecosistémico'; IPE es 'Proyección "
     "Estratégica' y no 'Presión Económica'; TMS es 'Movilización Social' y no "
     "'Tensión de Movilización'.",
     "Cruce automático sigla por sigla entre el Listado Blanco de MACC (Word) "
     "y las 318 filas del catálogo, comparando nombre además de sigla.",
     "★ ES EL BLOQUEADOR PRINCIPAL DEL PUENTE. La Regla de Oro de MACC dice "
     "que sólo ajusta variables que YA EXISTEN. Hoy no puede cumplirla: no hay "
     "a qué aplicarle el coeficiente. Y el bloque peor afectado es justamente "
     "el Territorial/Infraestructura (IVT, IRT, IPT), que es el nuestro.",
     "Antes de construir la tabla MACC hay que resolver, con Alexis, si esas "
     "variables se crean, se renombran, o el Listado Blanco se reescribe "
     "apuntando a las siglas reales del catálogo."),
    ("H-06", "MICR vs Catálogo", "Media",
     "Dos medidas paralelas de lo mismo: FVT de la matriz e IVIC del catálogo",
     "FVT (MICR) e IVIC (#234, 'Índice de Vulnerabilidad de Infraestructura "
     "Crítica') miden ambos la vulnerabilidad de infraestructura crítica, con "
     "fórmulas distintas y sin ninguna regla que diga cuál manda ni cómo se "
     "relacionan.",
     "Lectura comparada de la sección 22 (MICR) y la métrica 234 (MACH) del "
     "mismo documento Word.",
     "Riesgo de doble conteo: aplicar el coeficiente climático a las dos sería "
     "contar el mismo efecto dos veces, que es justamente lo que MACC prohíbe.",
     "Declarar la relación antes de calcular nada. Propuesta a discutir: FVT "
     "es del ACTIVO (tipo de elemento), IVIC es del SISTEMA (conjunto "
     "desplegado). No aplicar coeficiente a ambos."),
    ("H-07", "MACLIMA", "Alta",
     "ANPrecip usa valor absoluto y con eso pierde el signo: no distingue sequía de inundación",
     "El rango declarado es [0, +∞) y el Excel operativo lo hace explícito con "
     "ABS(). El canon lo dice sin rodeos: '0,2–0,4 → anomalía leve (déficit o "
     "exceso)'. Déficit y exceso caen en el mismo número.",
     "Fórmula del Word (ecuación OMML) + rango declarado + fórmula operativa "
     "del Excel de 318, las tres coincidentes.",
     "★ PARA ESTE PROYECTO ES CRÍTICO. Una sequía no corta un camino; un "
     "temporal sí. Si ANPrecip entrega el mismo valor para −2σ y +2σ, el "
     "coeficiente climático subiría la fragilidad vial en plena sequía, que es "
     "lo contrario de lo que pasa.",
     "Proponer un ANPrecip con signo (o un par déficit/exceso separado) para "
     "el uso infraestructural, manteniendo el absoluto para los usos donde el "
     "canon ya lo aplica. Decisión de Alexis; NO aplicado."),
    ("H-08", "Catálogo", "Baja",
     "El catálogo de 318 está desactualizado respecto del Word",
     "El Word incluye dos módulos que el Excel del 06-03-2026 no tiene: MCSGS "
     "(sección 23, con ICSGS y los factores FCN/FSS/FAS/FRC/FPI) y MCIE "
     "(sección 24, con el CIE). Tampoco están las diez columnas de la MICR.",
     "Comparación del índice del Word (25 secciones) con las 318 filas del "
     "Excel.",
     "No es un error, es desfase de fechas. Pero significa que trabajar sólo "
     "con el Excel deja fuera justo lo que este proyecto más necesita.",
     "Este Excel de proyecto los incorpora en hojas propias. Actualizar el "
     "maestro es decisión del RMD."),
    ("H-09", "MICR", "Baja",
     "Los 19 sectores tienen exactamente 44 elementos cada uno",
     "835 filas = 19 sectores × 44 (uno tiene 43). Se verificó que NO es una "
     "grilla rellenada: las 19 listas de elementos son todas distintas y hay "
     "641 nombres únicos. Los repetidos entre sectores son genéricos legítimos "
     "(UPS, generadores diésel, infraestructura vulnerable a ransomware).",
     "Conteo y comparación de las listas de elementos por sector.",
     "El inventario es real, lo cual es una buena noticia. Pero el 44 exacto "
     "sugiere una restricción de diseño.",
     "Preguntar si hubo sectores truncados o completados para cuadrar en 44. "
     "No afecta el cálculo."),
    ("H-15", "MACLIMA y MICR", "★ RESUELTO",
     "Los niveles de peligro son CUATRO, no tres: falta «Muy Alta» — ADOPTADO 16-ago",
     "El diccionario de la propia capa de SERNAGEOMIN define "
     "MT-POSOC-01=Baja · 02=Moderada · 03=Alta · 04=MUY ALTA. Tanto el módulo "
     "MACLIMA como la Matriz de Infraestructura Crítica hablan de tres niveles "
     "(Alta/Media/Baja).",
     "Consulta al servicio ArcGIS de la Minuta Técnica el 15-ago-2026, leyendo "
     "el diccionario de dominios de la capa.",
     "★ Calibrar el FEN contra tres niveles cuando la fuente publica cuatro "
     "dejaría «Alta» y «Muy Alta» pegadas — que es justo la distinción que más "
     "importa para priorizar con recursos escasos. Si todo lo grave es «Alta», "
     "no hay a quién ir primero.",
     "★ ALEXIS LO APROBÓ EL 16-AGO-2026. El proyecto adopta la escala de cuatro "
     "niveles: FEN ∈ {Baja, Media/Moderada, Alta, Muy Alta}. Implementado como "
     "escala `peligro_4` en normalizar.py y es la que usan los adaptadores. "
     "La Matriz original del RMD sigue sin tocarse: la adopción rige para este "
     "proyecto, y su traslado al canon es decisión aparte."),
    ("H-16", "Sub-matriz", "Media",
     "Dos subestaciones tienen mal la provincia en la sub-matriz",
     "La sub-matriz trae región y provincia escritas a mano. Derivadas de la "
     "coordenada contra la capa oficial, coinciden en 34 de 39 (tres de las "
     "diferencias son sólo ortografía: «Coihaique»/«Coyhaique»). Dos son "
     "errores reales: **Nueva Pozo Almonte** figura en provincia Iquique y la "
     "coordenada da **Tamarugal**; **Escondida** figura en El Loa y la "
     "coordenada da **Antofagasta**.",
     "Consulta espacial punto a punto contra la capa COMUNAS_2020 de "
     "SERNAGEOMIN, comparada con lo que ya traía la sub-matriz.",
     "Son dos fuentes independientes en desacuerdo. Con la provincia mal, el "
     "activo se agrega al COGRID provincial equivocado.",
     "Decidir cuál manda: la coordenada o el dato escrito. NO se corrigió."),
    ("H-17", "Sub-matriz", "Media",
     "La coordenada de la Subestación Valparaíso no cae dentro de ninguna comuna",
     "De las 39 subestaciones, 38 se ubicaron correctamente contra la capa "
     "oficial. La de Valparaíso (-33,04 / -71,62) no cae dentro de ningún "
     "polígono comunal — con toda probabilidad está en la bahía.",
     "Consulta espacial contra COMUNAS_2020; el servicio devolvió cero "
     "polígonos para ese punto.",
     "Sin comuna no entra al nivel comunal, que es el que atiende el COGRID. Y "
     "sugiere que la coordenada está mal tomada.",
     "Verificar la coordenada real de la subestación. NO se le asignó comuna: "
     "quedó en blanco antes que inventada."),
    ("H-11", "MICR vs Ley 21.542", "Alta",
     "«Tiempo de recuperación» es criterio legal de impacto y no existe en ningún lado",
     "El proyecto de ley que desarrolla la Ley 21.542 (Boletín 16143-02) fija "
     "cinco criterios de IMPACTO: personas afectadas, impacto económico, "
     "impacto operativo, reputación del Estado y TIEMPO DE RECUPERACIÓN. Los "
     "cuatro primeros tienen dónde apoyarse en el RMD. El quinto no está ni en "
     "las 10 columnas de la MICR, ni en las 318 del catálogo, ni en el MCSGS.",
     "Cruce del articulado (vía Castillo y Saldaña 2024) contra el catálogo "
     "completo de 318 y las columnas de la matriz.",
     "★ Sin tiempo de recuperación, la matriz no puede cumplir el criterio de "
     "impacto que la ley exige. Y para clima es doblemente importante: un "
     "camino cortado dos días y uno cortado dos meses no son el mismo evento, "
     "aunque el daño físico sea parecido.",
     "Crear la métrica. Propuesta a discutir: T_rec por elemento, modulada por "
     "accesibilidad y por estación del año (en invierno se repara más lento). "
     "NO creada aún."),
    ("H-12", "MICR vs Ley 21.542", "Alta",
     "La MICR no tiene resiliencia ni interdependencia, que la ley pide como criterios de criticidad",
     "La ley define CRITICIDAD por cuatro sub-criterios: seguridad, "
     "resiliencia, vulnerabilidad e interdependencia. La MICR cubre seguridad "
     "(FANC) y vulnerabilidad (FVT/VT). No tiene resiliencia ni "
     "interdependencia: trata cada uno de sus 835 elementos como una isla.",
     "Comparación de los cuatro sub-criterios legales contra las diez columnas "
     "de la matriz.",
     "★ Es el argumento más fuerte para incorporar el MCSGS al proyecto: aporta "
     "resiliencia (FRC) e interdependencia (FAS y FPI), que son justamente los "
     "dos que faltan. El MCSGS no es un lujo teórico — es lo que vuelve la "
     "matriz compatible con la ley.",
     "Incorporar FRC y FAS como columnas derivadas de la MICR, calculadas por "
     "el MCSGS. NO incorporadas aún."),
    ("H-13", "Sub-matriz", "Media",
     "La sub-matriz de subestaciones no trae comuna, que es el nivel que exige el COGRID",
     "Trae Región y Provincia, pero no Comuna. El COGRID opera en nivel "
     "comunal, regional y nacional, y la instrucción del 15-ago fija los cuatro "
     "niveles (comunal, provincial, regional, nacional) como requisito de "
     "diseño. Falta justo el más fino, que es donde el corte se vive.",
     "Lectura de las 11 columnas del archivo «Matriz 120-Subestaciones.xlsx».",
     "Sin comuna, el instrumento no puede entregarle nada al COGRID comunal, "
     "que es el primero que responde.",
     "Derivar la comuna de la coordenada contra la capa oficial de límites "
     "comunales (BCN / IDE Chile). NO deducirla del nombre de la subestación: "
     "«Chungará», «Collahuasi», «Crucero», «Maitencillo», «Guindo» y "
     "«Chacabuco» son localidades o faenas, no comunas. Pendiente."),
    ("H-14", "MICR", "Baja",
     "El FEN=Alta de las subestaciones se justifica con un ejemplo de SABOTAJE, no de evento natural",
     "La descripción del ítem 120 dice: «Instalaciones que transforman y "
     "distribuyen electricidad, vulnerables a sabotaje (Península Ibérica)». El "
     "sabotaje es materia de FANC, no de FEN. La columna de fragilidad ante "
     "eventos NATURALES quedó en Alta sin que el documento diga ante qué "
     "evento natural.",
     "Lectura de la fila 120 del Excel HOMOLOGADO.",
     "Es un ejemplo concreto de lo que el proyecto viene a resolver: FEN está "
     "puesto sin evidencia climática y sin territorio. No invalida el valor "
     "—las subestaciones sí son frágiles a inundación y aluvión— pero muestra "
     "que la justificación no está escrita.",
     "Documentar el evento natural que respalda cada FEN=Alta, o dejarlo "
     "derivar del dato climático. NO modificado."),
    ("H-10", "Sub-matriz", "Media",
     "La sub-matriz de subestaciones es una muestra, no un inventario",
     "39 subestaciones para todo Chile, repartidas prolijamente en 2-3 por "
     "cada una de las 16 regiones, y 20 rurales contra 19 urbanas. El país "
     "tiene bastantes más. Es el único material georreferenciado disponible.",
     "Lectura completa del archivo «Matriz 120-Subestaciones.xlsx».",
     "Sirve perfecto como molde y como piloto (tiene coordenadas, operador y "
     "teléfono), pero no permite conclusiones de cobertura nacional.",
     "Usarla como piloto declarando que es muestra. Para vialidad hace falta "
     "conseguir el inventario de rutas georreferenciado del MOP/Vialidad, que "
     "hoy no existe en el proyecto."),
]

# ─────────────────────────────────────────────────────────────────────────────
# NIVELES ADMINISTRATIVOS Y MARCO NORMATIVO.
# Instrucción formal de Alexis (15-ago-2026): la matriz operacional debe modelar
# en los cuatro niveles jurídico-administrativos de Chile, acatar la Ley 21.542
# y servir a los propósitos del COGRID / SENAPRED.
# El canon ya tenía la mitad hecha: el MCSGS EXIGE declarar la «US» (Unidad de
# Sistema) antes de calcular nada, con US-Nacional / US-Regional / US-Global.
# Lo que falta es extender esa escalera hacia abajo. Es extensión, no invención.
# ─────────────────────────────────────────────────────────────────────────────
NIVELES = [
    ("US-Comunal", "Comuna", "346",
     "COGRID comunal · municipio · dirección comunal de emergencia",
     "El nivel donde el corte se vive: un pueblo aislado, un APR sin bomba, un "
     "puente caído. Es también el nivel donde la matriz actual es MÁS ciega, "
     "porque sus 835 elementos son tipos de activo sin ubicación.",
     "NUEVO — hay que crearlo. La sub-matriz de subestaciones no trae comuna "
     "(ver H-13); se deriva de coordenada contra la capa oficial de límites "
     "comunales, nunca a ojo por el nombre."),
    ("US-Provincial", "Provincia", "56",
     "Delegación Presidencial Provincial",
     "Nivel intermedio. Útil para cuencas y para corredores viales que cruzan "
     "varias comunas pero no la región entera.",
     "PARCIAL — la sub-matriz de subestaciones sí trae provincia."),
    ("US-Regional", "Región", "16",
     "COGRID regional · una de las 16 Direcciones Regionales de SENAPRED",
     "El nivel operativo de SENAPRED: hay exactamente una dirección regional "
     "por región. Es el nivel natural de entrega del instrumento.",
     "EXISTE en el canon (US-Regional) y en la sub-matriz."),
    ("US-Nacional", "País", "1",
     "COGRID nacional · SENAPRED · SINAPRED",
     "Nivel de catálogo nacional de infraestructura crítica y de decreto "
     "supremo del art. 32 N°21.",
     "EXISTE en el canon (US-Nacional)."),
]

# Criterios que la Ley 21.542 y su proyecto (Boletín 16143-02) exigen, y dónde
# están —o no— en el RMD. Esta tabla es el chequeo de cumplimiento normativo.
LEY_21542 = [
    ("Definición", "Infraestructura crítica (art. 32 N°21 CPR)",
     "«el conjunto de instalaciones, sistemas físicos o servicios esenciales y "
     "de utilidad pública, así como aquellos cuya afectación cause un grave daño "
     "a la salud o al abastecimiento de la población, a la actividad económica "
     "esencial, al medioambiente o a la seguridad del país»",
     "Compatible — y notablemente afín",
     "Es una definición FUNCIONAL: clasifica por el daño que causa la "
     "afectación, no por el tipo de activo. Es exactamente la lógica del MCSGS "
     "(colapso funcional, no destrucción física). El canon del RMD ya piensa "
     "como piensa la ley."),
    ("Criticidad", "Seguridad", "Sub-criterio de criticidad del catálogo nacional",
     "FANC (MICR)", "Cubierto."),
    ("Criticidad", "Vulnerabilidad", "Sub-criterio de criticidad",
     "FVT y VT (MICR)", "Cubierto — con la salvedad del H-01 (FVT no reproduce)."),
    ("Criticidad", "Resiliencia", "Sub-criterio de criticidad",
     "NO está en la MICR · sí en IRL #210 y en FRC del MCSGS",
     "★ HUECO EN LA MATRIZ. La ley lo exige y las diez columnas de la MICR no "
     "lo tienen. Lo aporta el MCSGS."),
    ("Criticidad", "Interdependencia", "Sub-criterio de criticidad",
     "NO está en la MICR · sí en FAS del MCSGS",
     "★ HUECO EN LA MATRIZ. La ley lo exige y la MICR trata cada elemento "
     "aislado. Lo aporta el MCSGS (acoplamiento y propagación)."),
    ("Impacto", "Cantidad de personas afectadas", "Criterio de impacto",
     "IVC #241, TD #242", "Cubierto."),
    ("Impacto", "Impacto económico", "Criterio de impacto",
     "IIE #249, IIECo #60", "Cubierto."),
    ("Impacto", "Impacto operativo", "Criterio de impacto",
     "IDInf_X #246", "Cubierto."),
    ("Impacto", "Impacto en la reputación del Estado", "Criterio de impacto",
     "IDInf #109, ILE", "Cubierto parcialmente."),
    ("Impacto", "Tiempo de recuperación", "Criterio de impacto",
     "NO EXISTE en ninguna parte",
     "★★ HUECO TOTAL. Ni la MICR ni las 318 del catálogo ni el MCSGS miden "
     "cuánto tarda un elemento en volver a operar. La ley lo exige "
     "explícitamente. Hay que crearlo (ver H-11)."),
    ("Fases", "Fase de normalidad", "Monitoreo, planificación, obligaciones de operadores",
     "Es la fase de ESTE proyecto",
     "★ Castillo y Saldaña (2024) señalan que la falla del estatuto chileno "
     "está justamente acá, no en la emergencia. Una matriz predictiva movida "
     "por clima opera exactamente en esta fase."),
    ("Fases", "Fase de emergencia / protección reforzada",
     "Decreto supremo fundado; intervención de las FF.AA.",
     "Pev y Peh de la MICR",
     "Fuera del alcance de este proyecto (es el lado ataque, no el lado "
     "desastre natural)."),
]

ENC_ORIG = ["Número", "Sigla", "Nombre", "Tipo", "Categoría",
            "Sub Categorías para Análisis y Gráficos", "Fórmula SharePoint",
            "Descripción Técnica", "Descripción No Técnica", "Tipo de Dato",
            "Dependencias", "Lista Externa", "Fuentes", "Fundamentación Teórica",
            "Metodología y Validez de Fuentes", "Rango de Factores de Ajuste",
            "Instrucciones de Uso y Notas", "Sugerencias de Visualización",
            "Ejemplos Históricos", "Relaciones con Otras Variables",
            "Limitaciones Conocidas"]

# ── estilos ──────────────────────────────────────────────────────────────────
AZUL = PatternFill("solid", fgColor="1F3864")
VERDE = PatternFill("solid", fgColor="2E6B4F")
GRIS = PatternFill("solid", fgColor="404040")
ROJO = PatternFill("solid", fgColor="7B2D26")
NUCLEO = PatternFill("solid", fgColor="FFF2CC")
BLANCO_NEG = Font(color="FFFFFF", bold=True, size=11)
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def encabezar(ws, encabezados, relleno, alto=42):
    ws.append(encabezados)
    for c in ws[1]:
        c.fill, c.font = relleno, BLANCO_NEG
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    ws.row_dimensions[1].height = alto
    ws.freeze_panes = "A2"


def anchos(ws, medidas):
    for i, w in enumerate(medidas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def envolver(ws, desde_fila=2):
    for fila in ws.iter_rows(min_row=desde_fila):
        for c in fila:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDE


def main():
    orig = load_workbook(CATALOGO, data_only=True)["Hoja1"]
    catalogo = {r[0]: r for r in orig.iter_rows(min_row=2, values_only=True)
                if r[0] is not None}

    wb = Workbook()

    # ── Hoja 1: LÉEME ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "LÉEME"
    anchos(ws, [26, 118])
    lineas = [
        ("VARIABLES Y MÉTRICAS — INFRAESTRUCTURA CRÍTICA × CLIMA", ""),
        ("", ""),
        ("Qué es esto",
         "El subconjunto del catálogo RMD 2.0 que sirve a este proyecto, y sólo ese. "
         "El catálogo maestro tiene 318 variables y métricas; acá quedan las que "
         "efectivamente participan en medir cómo el clima afecta a la infraestructura "
         "crítica de Chile."),
        ("Pregunta del proyecto",
         "Las lluvias de julio y agosto de 2026 mostraron que la infraestructura vial "
         "es el punto débil del país. La matriz de infraestructura crítica sabe QUÉ "
         "importa, pero no CUÁL, DÓNDE ni CUÁNDO está por fallar. Este proyecto busca "
         "darle esas tres cosas usando datos climáticos reales."),
        ("Fuentes y precedencia",
         "MANDA el Word «RMD_2_Variables_y_METRICAS_COMPLETAS-11-06-2026.docx». El Excel "
         "«Variables-y-Metricas-318-06-03-2026-Tabla.xlsx» es la versión operativa y es "
         "lo práctico, pero donde discrepan gana el Word. Por eso cada fila trae las DOS "
         "fórmulas, en columnas separadas: así la discrepancia se ve en lugar de "
         "esconderse. Se sumaron «Matriz de Infraestructura Crítica - FINAL.docx», "
         "«Matriz_Infraestructura_Critica_Prioridad_Estrategica_HOMOLOGADA.xlsx» (835 "
         "filas), «Matriz 120-Subestaciones.xlsx» (39 activos) y el módulo MACLIMA."),
        ("Cómo leer las hojas",
         "· «Variables y Métricas»: la selección del catálogo, con su N° original para "
         "poder volver siempre al maestro.\n"
         "· «MICR Columnas»: las 10 columnas de la matriz de infraestructura. NO están "
         "en el catálogo de 318 — la matriz vive en su propio archivo.\n"
         "· «MCSGS Colapso»: el módulo de colapso sistémico. Tampoco está en el "
         "catálogo (es posterior al Excel del 06-03-2026).\n"
         "· «Hallazgos»: 10 cosas que aparecieron al revisar y que NO se corrigieron."),
        ("Qué NO se hizo",
         "No se modificó ni un dato de la Matriz ni del Módulo. Alexis fijó la regla el "
         "15-ago-2026: sólo se modifica contra datos reales. Todo lo encontrado quedó "
         "anotado en la hoja «Hallazgos», con cómo se comprobó cada cosa, para que sea "
         "auditable y para decidir después."),
        ("El hallazgo que hay que mirar primero",
         "H-05. El Listado Blanco de MACC — la lista de variables a las que el módulo "
         "climático puede aplicar su coeficiente — nombra 17 variables, y sólo 3 existen "
         "en el catálogo con ese nombre. Cinco no existen, y nueve existen con la misma "
         "sigla pero significando otra cosa. Como la Regla de Oro de MACC dice que sólo "
         "ajusta variables que ya existen, hoy el módulo no tiene a qué aplicarse. Y el "
         "bloque peor afectado es el Territorial/Infraestructura, que es el nuestro."),
        ("La idea en una línea",
         "FEN (Fragilidad ante Eventos Naturales) es hoy una etiqueta de tres niveles, "
         "sin territorio y sin tiempo. La propuesta es volverla función del lugar y del "
         "mes: FEN_efectivo = FEN_base(tipo) × C_clim(lugar, mes), con C_clim saliendo "
         "de MACC. Como Pen depende de FEN en un 60% (0,5 directo + 0,2 vía FVT), la "
         "prioridad ante desastres pasaría a moverse con el clima real."),
        ("Generado", "15-ago-2026 · script «construir_variables_y_metricas_proyecto.py», "
                     "en esta misma carpeta. Es reproducible: se vuelve a correr y "
                     "reconstruye este archivo desde los originales, sin tocarlos."),
    ]
    for k, v in lineas:
        ws.append([k, v])
    ws["A1"].font = Font(bold=True, size=15, color="1F3864")
    for fila in ws.iter_rows(min_row=3):
        fila[0].font = Font(bold=True, color="1F3864")
        for c in fila:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Hoja 2: Variables y Métricas ─────────────────────────────────────────
    ws = wb.create_sheet("Variables y Métricas")
    enc = (["N° catálogo", "Sigla", "Nombre", "Tipo", "Categoría",
            "Subcategoría", "★ Núcleo", "Rol en ESTE proyecto",
            "Fórmula CANÓNICA (Word — manda)",
            "Fórmula OPERATIVA (Excel 318)",
            "Dato real que la alimenta", "Estado"]
           + ENC_ORIG[7:])
    encabezar(ws, enc, AZUL, 52)
    for num, nucleo, f_word, rol, dato, estado in SELECCION:
        o = catalogo.get(num)
        if o is None:
            print(f"  ¡OJO! el N° {num} no está en el catálogo — fila omitida")
            continue
        ws.append([o[0], o[1], o[2], o[3], o[4], o[5],
                   "★" if nucleo else "", rol, f_word, o[6], dato, estado]
                  + list(o[7:21]))
    anchos(ws, [11, 15, 34, 13, 14, 26, 8, 60, 58, 44, 42, 20]
           + [46, 40, 14, 30, 22, 34, 40, 36, 34, 34, 36, 34])
    envolver(ws)
    for fila in ws.iter_rows(min_row=2):
        if fila[6].value == "★":
            for c in fila[:12]:
                c.fill = NUCLEO

    # ── Hoja 3: MICR Columnas ────────────────────────────────────────────────
    ws = wb.create_sheet("MICR Columnas")
    encabezar(ws, ["Columna", "Nombre completo", "Cálculo", "Escala",
                   "Rol en ESTE proyecto", "¿La toca el clima?"], VERDE, 46)
    for f in MICR_COLUMNAS:
        ws.append(list(f))
    anchos(ws, [12, 46, 62, 34, 82, 26])
    envolver(ws)
    for fila in ws.iter_rows(min_row=2):
        if str(fila[5].value).startswith("Sí"):
            fila[0].fill = NUCLEO

    # ── Hoja 4: MCSGS ────────────────────────────────────────────────────────
    ws = wb.create_sheet("MCSGS Colapso")
    encabezar(ws, ["Sigla", "Nombre", "Fórmula / criterio", "Escala",
                   "Rol en ESTE proyecto"], GRIS, 40)
    for f in MCSGS_ITEMS:
        ws.append(list(f))
    anchos(ws, [10, 48, 58, 34, 96])
    envolver(ws)

    # ── Hoja 5: Niveles y Normativa ──────────────────────────────────────────
    ws = wb.create_sheet("Niveles y Normativa")
    ws.append(["NIVELES DE MODELACIÓN — requisito de diseño (instrucción del 15-ago-2026)"])
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.append([])
    fila_enc = ws.max_row + 1
    ws.append(["Unidad de Sistema (US)", "Nivel administrativo", "N° en Chile",
               "A quién sirve", "Por qué importa acá", "Estado"])
    for c in ws[fila_enc]:
        c.fill, c.font = AZUL, BLANCO_NEG
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    for f in NIVELES:
        ws.append(list(f))
    ws.append([])
    ws.append(["LEY 21.542 — chequeo de cumplimiento normativo"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=13, color="7B2D26")
    ws.append([])
    fila_enc2 = ws.max_row + 1
    ws.append(["Bloque", "Criterio que exige la ley", "Qué dice la norma",
               "Dónde está en el RMD", "Comentario"])
    for c in ws[fila_enc2]:
        c.fill, c.font = ROJO, BLANCO_NEG
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")
    for f in LEY_21542:
        ws.append(list(f))
    anchos(ws, [22, 24, 13, 46, 76, 60])
    for fila in ws.iter_rows(min_row=2):
        for c in fila:
            if c.value and c.row not in (fila_enc, fila_enc2):
                c.alignment = Alignment(wrap_text=True, vertical="top")
    # resalto los tres huecos normativos
    for fila in ws.iter_rows(min_row=fila_enc2 + 1):
        if fila[4].value and str(fila[4].value).startswith("★"):
            fila[3].fill = NUCLEO
            fila[4].font = Font(bold=True, color="C00000")

    # ── Hoja 6: Hallazgos ────────────────────────────────────────────────────
    ws = wb.create_sheet("Hallazgos")
    encabezar(ws, ["ID", "Dónde", "Gravedad", "Hallazgo", "Detalle",
                   "Cómo se comprobó", "Por qué importa acá",
                   "Qué hacer (NO aplicado)"], ROJO, 42)
    for f in sorted(HALLAZGOS, key=lambda h: h[0]):   # por ID, no por orden de hallazgo
        ws.append(list(f))
    anchos(ws, [8, 18, 11, 54, 78, 52, 62, 62])
    envolver(ws)
    for fila in ws.iter_rows(min_row=2):
        if fila[2].value == "Alta":
            fila[2].font = Font(bold=True, color="C00000")

    wb.save(SALIDA)
    print(f"Escrito: {SALIDA}")
    print(f"  · {len(SELECCION)} variables/métricas del catálogo de 318")
    print(f"  · {len(MICR_COLUMNAS)} columnas MICR")
    print(f"  · {len(MCSGS_ITEMS)} ítems MCSGS")
    print(f"  · {len(NIVELES)} niveles administrativos + {len(LEY_21542)} "
          f"criterios de la Ley 21.542")
    print(f"  · {len(HALLAZGOS)} hallazgos anotados")


if __name__ == "__main__":
    main()
