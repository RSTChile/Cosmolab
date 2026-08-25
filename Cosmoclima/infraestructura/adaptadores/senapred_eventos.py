"""
Adaptador SENAPRED — la capa de validación: qué falló de verdad, dónde y cuándo.

QUÉ ES
------
El «Consolidado de Eventos de Emergencia en Chile 2015-2024» que SENAPRED
publicó en su biblioteca digital (BiblioGRD) en septiembre de 2025: **50.457
eventos**, con fecha, comuna, tipo, organismos que respondieron y personas
afectadas. Es la agregación de los informes ALFA/DELTA de las Unidades de Alerta
Temprana regionales.

POR QUÉ ES LA PIEZA MÁS IMPORTANTE DEL PROYECTO
-----------------------------------------------
Todo lo demás dice qué PUEDE pasar. Esto dice qué PASÓ. Sin esta serie, el
modelo sería elegante e infalsable: podría producir prioridades preciosas y no
habría forma de saber si sirven. Con ella se puede preguntar lo único que
importa: cuando el modelo dijo «peligro alto acá», ¿falló algo?

Y trae dos cosas que no esperábamos:

1. **`Total Aislados`** — el aislamiento, cuantificado oficialmente por comuna y
   fecha. El glosario define aislamiento como «acceso interrumpido y sin acceso
   alternativo», que es la definición de nodo de flujo del MCSGS. Acá está
   medido en personas.
2. La confirmación empírica de la premisa del proyecto: de los 289 eventos con
   personas aisladas (140.303 personas en total), las causas dominantes son
   inundación, sistema frontal, lluvia, nevadas y remoción en masa. **El
   aislamiento en Chile lo produce el clima**, que es justo lo que este modelo
   pretende anticipar.

PRIVACIDAD — regla dura
-----------------------
La columna `Antecedentes Observaciones` es texto libre y contiene datos
personales (se detectaron RUT y descripciones de personas fallecidas). **NUNCA
se lee, no se copia y no se guarda.** El adaptador trabaja sólo con conteos
agregados por comuna, que es lo que el proyecto necesita. La columna
`Sub Evento 1` también se descarta, por venir corrupta.

CONDICIONES DE USO
------------------
BiblioGRD permite el acceso automatizado a `/bitstream/` (robots.txt verificado
el 15-ago-2026). El archivo se baja una vez y queda en `datos/crudo/`; no hace
falta volver a pedirlo.
"""

import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

AQUI = Path(__file__).parent.parent
sys.path.insert(0, str(AQUI))
import normalizar  # noqa: E402

XLSX = (AQUI / "datos" / "crudo" / "senapred" / "2026-08-15" /
        "Eventos_Emergencia_2015_2024.xlsx")
HOJA = "Eventos_de_Emergencia_2015_2024"
ID_FUENTE = "senapred_eventos_2015_2024"

# Columnas que NO se leen nunca. No es una preferencia: es la regla de privacidad.
COLUMNAS_PROHIBIDAS = {43}          # Antecedentes Observaciones (texto libre)
COLUMNAS_DESCARTADAS = {8}          # Sub Evento 1 (corrupta en origen)

FUENTE = dict(
    id=ID_FUENTE,
    organismo="SENAPRED (Unidades de Alerta Temprana Regionales)",
    producto="Consolidado de Eventos de Emergencia en Chile 2015-2024",
    url="https://bibliogrd.senapred.gob.cl/handle/1671/7120",
    formato="xlsx",
    familia="ESTADO",
    acceso="anonimo",
    acceso_verificado=1,
    condiciones_uso="BiblioGRD permite /bitstream/ en robots.txt (verificado "
                    "15-ago-2026). Publicado sep-2025.",
    permite_automatizacion="si",
    granularidad="comuna",
    historia_desde="2015-01-01",
    frecuencia="consolidado anual (última edición cubre hasta 2024-12-31)",
    confianza_base=0.90,
    notas="Partes operativos de las UAT regionales, no prensa. 430 cadenas de "
          "comuna para 346 comunas reales: requiere normalización de nombres. "
          "La columna de observaciones tiene datos personales y NO se lee.",
)

# Qué cuenta como «falló un servicio». Se arma por expresión regular sobre el
# tipo de evento porque el mismo hecho aparece escrito de varias formas
# («Interrupción Suministro Eléctrico», «Interrupción de suministro de
# electricidad», «Alteración Suministro Eléctrico»...). Diez años de captura
# manual en dieciséis regiones se ven en la ortografía.
FAMILIAS_FALLA = {
    "electricidad": re.compile(r"(interrup|alterac).*(el[eé]ctric|electricidad)", re.I),
    "agua": re.compile(r"(interrup|alterac).*(agua potable)", re.I),
    "telecomunicaciones": re.compile(r"(interrup|alterac).*(telecomunicac)", re.I),
    "conectividad_vial": re.compile(r"(conectividad|^vial$)", re.I),
    "gas": re.compile(r"(interrup|alterac).*gas", re.I),
    "alcantarillado": re.compile(r"(interrup|alterac).*alcantarill", re.I),
    "suministros_basicos": re.compile(r"suministros b[aá]sicos", re.I),
}


def normalizar_nombre(texto):
    """Nombre de comuna a forma comparable: sin tildes, sin dobles espacios,
    en minúsculas. El archivo trae 430 cadenas distintas para 346 comunas
    reales, casi todo por tildes y espacios."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFD", str(texto).strip().lower())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t)


def clasificar(tipo_evento, clase_evento):
    """Devuelve la familia de falla, o None si el evento no es una falla de
    servicio (un incendio forestal, por ejemplo, no lo es)."""
    texto = f"{tipo_evento} {clase_evento}"
    for familia, patron in FAMILIAS_FALLA.items():
        if patron.search(texto):
            return familia
    return None


def traer(desde=None, hasta=None):
    """Interfaz común: devuelve (observaciones, motivo_si_falla).

    Emite dos tipos de observación por comuna y mes:
      · `falla_<familia>` — cuántas fallas de ese servicio hubo
      · `personas_aisladas` — cuántas personas quedaron aisladas
    Se agrega por MES y no por evento porque es la resolución a la que el
    modelo produce peligro, y comparar a resoluciones distintas sería trampa.
    """
    if not XLSX.exists():
        return [], f"falta el archivo {XLSX.name}"

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    it = wb[HOJA].iter_rows(values_only=True)
    encabezado = next(it)

    # control: que las columnas estén donde creemos que están. Si el archivo
    # cambia de forma, mejor parar que producir números sin sentido.
    esperado = {0: "Fecha Inicio", 2: "Región", 4: "Comuna", 6: "Clase Evento",
                7: "Tipo Evento", 11: "Total Afectados", 21: "Total Aislados"}
    for i, nombre in esperado.items():
        if encabezado[i] != nombre:
            return [], (f"la columna {i} debería ser «{nombre}» y es "
                        f"«{encabezado[i]}» — el archivo cambió de forma")

    # (comuna, region, anio, mes) → conteos
    acumulado = defaultdict(lambda: defaultdict(float))

    for fila in it:
        if fila[0] is None:
            continue
        fecha = str(fila[0])[:10]
        if len(fecha) < 10:
            continue
        anio, mes = int(fecha[:4]), int(fecha[5:7])
        if desde and (anio, mes) < desde:
            continue
        if hasta and (anio, mes) > hasta:
            continue

        comuna = normalizar_nombre(fila[4])
        region = str(fila[2] or "").strip()
        if not comuna:
            continue

        clave = (comuna, region, anio, mes)
        familia = clasificar(fila[7], fila[6])
        if familia:
            acumulado[clave][f"falla_{familia}"] += 1
        aislados = fila[21]
        if isinstance(aislados, (int, float)) and aislados > 0:
            acumulado[clave]["personas_aisladas"] += aislados
        afectados = fila[11]
        if isinstance(afectados, (int, float)) and afectados > 0:
            acumulado[clave]["personas_afectadas"] += afectados

    hoy = date.today().isoformat()
    observaciones = []
    for (comuna, region, anio, mes), conteos in acumulado.items():
        ultimo = 31 if mes in (1, 3, 5, 7, 8, 10, 12) else (30 if mes != 2 else 28)
        if mes == 2 and (anio % 4 == 0 and (anio % 100 != 0 or anio % 400 == 0)):
            ultimo = 29
        base = dict(
            id_fuente=ID_FUENTE, familia="ESTADO",
            vigencia_inicio=f"{anio:04d}-{mes:02d}-01",
            vigencia_fin=f"{anio:04d}-{mes:02d}-{ultimo:02d}",
            territorio_tipo="comuna", territorio_id=comuna,
            comuna=comuna, region=region,
            confianza=FUENTE["confianza_base"],
            fecha_descarga=hoy, url_exacta=FUENTE["url"],
            ruta_crudo=str(XLSX.relative_to(AQUI)),
        )
        for variable, valor in conteos.items():
            # Es un CONTEO de hechos ocurridos, no una medición de intensidad:
            # no se normaliza a 0-1. Normalizar un conteo de fallas lo volvería
            # incomparable con el próximo mes, y acá lo que importa es
            # justamente comparar entre meses.
            observaciones.append(dict(
                base, variable=variable,
                valor_original=f"{valor:.0f}",
                unidad_original="eventos" if variable.startswith("falla_")
                                else "personas",
                notas="conteo mensual agregado por comuna"))

    return observaciones, None


if __name__ == "__main__":
    obs, problema = traer()
    if problema:
        print("SIN DATO:", problema)
        raise SystemExit(1)
    conteo = defaultdict(int)
    suma = defaultdict(float)
    for o in obs:
        conteo[o["variable"]] += 1
        suma[o["variable"]] += float(o["valor_original"])
    print(f"{len(obs):,} observaciones de {ID_FUENTE}\n")
    print(f"{'variable':28s} {'meses-comuna':>13s} {'total':>12s}")
    for v in sorted(conteo, key=lambda x: -suma[x]):
        print(f"{v:28s} {conteo[v]:13,d} {suma[v]:12,.0f}")
