#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ingerir_excel_estaciones.py -- suma a la tabla consolidada los datos diarios
reales de cualquier tanda de estaciones en formato "Informe Anual de Agua
Caída" (DMC/FDF/INIA).

Nació el 13-ago-2026 para la tanda de COQUIMBO ('datos/Excel Estaciones/
Nuevos/', 38 archivos, 84.116 filas) y se generalizó al día siguiente porque
van a llegar más tandas: ahora la carpeta se pasa por argumento y el guardia
de duplicados mira TODO el árbol ya ingerido, no una sola carpeta.

    python3 ingerir_excel_estaciones.py --carpeta "<ruta>" [--etiqueta "..."] [--dry-run]

QUÉ MIRAR EN CADA TANDA NUEVA. Lo más valioso no es que haya más estaciones
sino que alguna CRUCE EL CORTE DE 2018: todas las históricas mueren el
2017-05-31 (fin de la compilación CR2) y todas las nuevas empiezan en 2019,
así que el supuesto "la estación nueva mide como la histórica co-localizada"
quedó sin validar, con un escalón medido de 1,00x a 1,44x. Una sola estación
con registro continuo 2016-2020 lo resuelve. El informe de este script avisa
si aparece alguna.

POR QUÉ IMPORTAN ESTAS Y NO SON "MÁS DE LO MISMO". Las 17 anteriores eran de
Atacama, al norte del punto-reloj. Éstas son de Coquimbo, que es donde el
punto-reloj ESTÁ (Huintil, -31.5669/-70.9817). Y sobre todo: cubren
2019-2026, que es exactamente el tramo donde el instrumento se quedó sin
estación real y hubo que empalmar con reanálisis ERA5 corregido por sesgo
(ver empalmar_lluvia_calibrada.py). Once de estas estaciones caen dentro de
60 km del punto-reloj, y la más cercana -- Vivero Conaf, Illapel -- a 17,7 km.
Si los valores son coherentes, el empalme con reanálisis puede reemplazarse
por medición real y la serie 1966-2026 queda enteramente de estación.
Ese análisis NO lo hace este script: acá sólo se ingiere el dato crudo.

FORMATO: idéntico al de Atacama ("Informe Anual de Agua Caída" DMC/FDF/INIA),
así que se reusa el mismo parser, verificado contra estos archivos:
  - hoja de metadata con Código Nacional / Nombre / Propietario / Latitud /
    Longitud (×100000 sin punto) / Altura(Mts.), a veces con 'Código OMM'
    de séptima columna y con una fila en blanco de por medio -- por eso la
    metadata se busca por "primera fila cuya primera celda es numérica",
    no por posición fija.
  - una hoja por año con grilla Día×Mes de 13 columnas.
  - una hoja "AGOSTO-2026"/"MARZO-2026" con el año en curso hasta el mes
    más reciente (mismo layout de grilla).
  - "s/p" = SIN PRECIPITACIÓN, dato real = 0.0mm. "." = SIN OBSERVACIÓN,
    se descarta. Nunca se inventa un dato.

DOS DIFERENCIAS DELIBERADAS CON EL SCRIPT DE ATACAMA:

1. La hoja de metadata se DETECTA, no se declara a mano. El script de
   Atacama traía un diccionario archivo->hoja escrito a mano tras revisión
   manual; con 38 archivos nuevos eso es una lista larga de oportunidades de
   equivocarse. Acá se busca la hoja que contiene la fila "Código Nacional".

2. Se SALTAN los archivos byte a byte idénticos a uno ya ingerido. En la
   carpeta venía 'ALTO.DEL.CARMEN_LOS SAUCES-ATACAMA-SOLO2026 (1).xlsx',
   copia exacta (mismo SHA-256) del que ya está en el nivel de arriba y ya
   fue ingerido con la tanda de Atacama. Sin este guardia se insertaría dos
   veces la misma estación bajo la misma localidad.

Uso:
    python3 agregar_excel_estaciones_coquimbo_2026_a_consolidada.py --dry-run
    python3 agregar_excel_estaciones_coquimbo_2026_a_consolidada.py
"""
import csv
import glob
import hashlib
import math
import os
import re
import sqlite3
import sys
from datetime import date

import openpyxl

RAIZ = "/Users/alexis/Desktop/RMD/Cosmolab/Cosmoclima"
CARPETA_EXCEL_RAIZ = os.path.join(RAIZ, "datos/Excel Estaciones")
DB = os.path.join(RAIZ, "investigacion/fuentes/pluviosidad_diaria_consolidada.sqlite")
# el CSV de trazabilidad se nombra segun la etiqueta de la tanda

HUINTIL_LAT, HUINTIL_LON = -31.5669, -70.9817

MESES_COLUMNA = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
                 "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def km_a_huintil(lat, lon):
    return math.hypot((lat - HUINTIL_LAT) * 111.0,
                      (lon - HUINTIL_LON) * 111.0 * math.cos(math.radians(HUINTIL_LAT)))


def sha256(ruta):
    h = hashlib.sha256()
    with open(ruta, "rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return h.hexdigest()


def hallar_hoja_metadata(wb):
    """La hoja de metadata es la que contiene la fila con 'Código Nacional'."""
    for nombre in wb.sheetnames:
        for fila in wb[nombre].iter_rows(min_row=1, max_row=15, values_only=True):
            if fila and fila[0] is not None and "Código Nacional" in str(fila[0]):
                return nombre
    return None


def leer_metadata(wb, hoja):
    """Primera fila cuya primera celda es numérica: tolera la fila en blanco
    que algunos archivos meten entre el encabezado y el dato."""
    for fila in wb[hoja].iter_rows(min_row=1, max_row=20, values_only=True):
        primera = fila[0]
        if primera is None:
            continue
        try:
            codigo = int(str(primera).strip())
        except (TypeError, ValueError):
            continue
        nombre, propietario, lat_raw, lon_raw, altura = fila[1], fila[2], fila[3], fila[4], fila[5]
        lat = float(str(lat_raw)) / 100000
        lon = float(str(lon_raw)) / 100000
        return codigo, str(nombre).strip(), str(propietario).strip(), lat, lon, float(altura)
    return None


def anio_de_hoja(nombre_hoja, hoja_metadata):
    if nombre_hoja == hoja_metadata:
        return None
    if re.fullmatch(r"(19|20)\d{2}", nombre_hoja):
        return int(nombre_hoja)
    m = re.fullmatch(r"AÑO(\d{4})", nombre_hoja)
    if m:
        return int(m.group(1))
    m = re.fullmatch(r"[A-ZÁÉÍÓÚÑ]+[-\s](\d{4})", nombre_hoja.strip())
    if m:
        return int(m.group(1))
    return None


def parsear_grilla_anio(ws, anio, etiqueta):
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 3:
        return [], f"{etiqueta}: hoja demasiado corta"
    header = list(filas[1][1:13])
    if header != MESES_COLUMNA:
        return [], f"{etiqueta}: header de meses inesperado {header}"
    dias = []
    for fila in filas[2:]:
        dia = fila[0]
        try:
            dia = int(str(dia).strip())
        except (TypeError, ValueError):
            continue
        if not (1 <= dia <= 31):
            continue
        for mes_idx, valor in enumerate(fila[1:13], start=1):
            if valor is None:
                continue
            txt = str(valor).strip()
            if txt in (".", "", "-"):
                continue          # sin observación: se descarta, no se inventa
            if txt.lower() == "s/p":
                mm = 0.0          # sin precipitación: dato real confirmado
            else:
                try:
                    mm = float(txt.replace(",", "."))
                except ValueError:
                    continue
            try:
                date(anio, mes_idx, dia)
            except ValueError:
                continue          # 31 de febrero y similares de plantilla
            dias.append((f"{anio:04d}-{mes_idx:02d}-{dia:02d}", mm))
    return dias, None


def arg(nombre, defecto=None):
    if nombre in sys.argv:
        i = sys.argv.index(nombre)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return defecto


def main():
    dry = "--dry-run" in sys.argv
    carpeta = arg("--carpeta")
    if not carpeta or not os.path.isdir(carpeta):
        sys.exit("Falta --carpeta <ruta a la carpeta con los .xlsx de la tanda>\n"
                 "Ej: --carpeta 'datos/Excel Estaciones/Nuevos'")
    etiqueta = arg("--etiqueta", os.path.basename(carpeta.rstrip("/")))
    csv_out = os.path.join(RAIZ, "investigacion/fuentes",
                           f"pluviosidad_diaria_estaciones_{etiqueta.lower().replace(' ', '_')}.csv")
    print(f"Tanda: {etiqueta}\nCarpeta: {carpeta}\n")

    # Guardia de duplicados: se compara contra TODO el arbol ya ingerido, no
    # contra una sola carpeta -- en la tanda de Coquimbo venia una copia byte a
    # byte de un archivo de Atacama y sin esto se habria insertado dos veces.
    ya = {}
    for ruta in glob.glob(os.path.join(CARPETA_EXCEL_RAIZ, "**", "*.xlsx"), recursive=True):
        if os.path.abspath(os.path.dirname(ruta)) == os.path.abspath(carpeta):
            continue
        ya[sha256(ruta)] = os.path.relpath(ruta, CARPETA_EXCEL_RAIZ)

    filas_totales, metadatas, avisos, saltados = [], [], [], []

    for ruta in sorted(glob.glob(os.path.join(carpeta, "*.xlsx"))):
        archivo = os.path.basename(ruta)
        h = sha256(ruta)
        if h in ya:
            saltados.append((archivo, ya[h]))
            continue

        wb = openpyxl.load_workbook(ruta, data_only=True)
        hoja_meta = hallar_hoja_metadata(wb)
        if hoja_meta is None:
            avisos.append(f"{archivo}: no se encontró hoja con 'Código Nacional' -- NO ingerido")
            continue
        meta = leer_metadata(wb, hoja_meta)
        if meta is None:
            avisos.append(f"{archivo}: hoja {hoja_meta!r} sin fila de metadata legible -- NO ingerido")
            continue
        codigo, nombre, propietario, lat, lon, altura = meta

        dias_estacion = []
        for nombre_hoja in wb.sheetnames:
            anio = anio_de_hoja(nombre_hoja, hoja_meta)
            if anio is None:
                continue
            dias, err = parsear_grilla_anio(wb[nombre_hoja], anio, f"{archivo}/{nombre_hoja}")
            if err:
                avisos.append(err)
                continue
            dias_estacion.extend(dias)

        if not dias_estacion:
            avisos.append(f"{archivo}: ninguna hoja aportó días reales -- NO ingerido")
            continue

        localidad = f"{nombre} (DMC/INIA {codigo})"
        fuente = (f"Excel estación real {etiqueta} (Alexis), "
                  f"código {codigo}, propietario {propietario}")
        for fecha, mm in dias_estacion:
            filas_totales.append((fecha, localidad, lat, lon, mm, fuente, codigo))

        fechas = sorted(f for f, _ in dias_estacion)
        d = km_a_huintil(lat, lon)
        metadatas.append((archivo, codigo, nombre, propietario, lat, lon, altura,
                          len(dias_estacion), fechas[0], fechas[-1], d))

    metadatas.sort(key=lambda m: m[10])
    print(f"{'estación':<40}{'km':>7}{'días':>8}  rango")
    for (_a, cod, nom, _p, _la, _lo, _alt, n, f0, f1, d) in metadatas:
        print(f"{(nom[:37] + ' (' + str(cod) + ')')[:39]:<40}{d:>7.1f}{n:>8}  {f0} a {f1}")

    cerca = [m for m in metadatas if m[10] <= 60]
    print(f"\nEstaciones dentro de 60 km del punto-reloj: {len(cerca)} de {len(metadatas)}")
    dias_cerca = sum(m[7] for m in cerca)
    print(f"Días reales que aportan esas {len(cerca)}: {dias_cerca:,}")

    # El puente REAL exige dato a ambos lados: antes de que muera el registro
    # historico (2017-05-31, fin de la compilacion CR2) y despues de que
    # empiecen las estaciones nuevas (2019-01). Una estacion que solo cubra
    # 2018-2019 NO sirve: no solapa con ninguna historica, y de hecho la unica
    # que hay asi (La Canela reten) fue la que hizo fallar la validacion de
    # cadena en +105%. Criterio flojo = luz verde falsa.
    cruzan = [m for m in metadatas if m[8] < "2017-06" and m[9] > "2019-01"]
    if cruzan:
        print("\n*** ESTACIONES QUE CRUZAN EL CORTE DE 2018 ***")
        print("Esto cierra el supuesto que quedó sin validar en la ronda 17:")
        for (_a, cod, nom, _p, _la, _lo, _alt, n, f0, f1, d) in cruzan:
            print(f"  {nom} ({cod})  {f0} a {f1}  ·  {d:.1f} km del punto-reloj")
        print("Conviene recalibrar la reconstrucción de lluvia con esto.")
    else:
        print("\n(ninguna estación de esta tanda cruza el corte de 2018 --")
        print(" el escalón 1,00x-1,44x sigue sin poder validarse)")

    if saltados:
        print(f"\nSaltados por ser copia exacta de un archivo ya ingerido ({len(saltados)}):")
        for a, orig in saltados:
            print(f"  {a}\n     idéntico (SHA-256) a {orig}")
    if avisos:
        print(f"\nAVISOS ({len(avisos)}):")
        for a in avisos:
            print("  " + a)

    print(f"\nFilas a insertar: {len(filas_totales):,} de {len(metadatas)} estaciones")
    if dry:
        print("\n--dry-run: no se tocó la base ni se escribió el CSV.")
        return

    with open(csv_out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["fecha", "localidad", "lat", "lon", "lluvia_mm", "fuente"])
        w.writerows((fe, l, la, lo, mm, fu) for (fe, l, la, lo, mm, fu, _c) in filas_totales)
    print(f"CSV de trazabilidad: {csv_out}")

    # Inserción idempotente: por localidad se borra el rango exacto que este
    # script está por insertar y se reinserta -- correr dos veces deja la base
    # igual que correrlo una. Mismo patrón que el script de Atacama.
    con = sqlite3.connect(DB)
    antes = con.execute("SELECT COUNT(*) FROM pluviosidad_diaria").fetchone()[0]
    por_localidad = {}
    for fila in filas_totales:
        por_localidad.setdefault(fila[1], []).append(fila)
    for localidad, filas in por_localidad.items():
        con.execute("DELETE FROM pluviosidad_diaria WHERE localidad = ? AND fecha BETWEEN ? AND ?",
                    (localidad, min(f[0] for f in filas), max(f[0] for f in filas)))
    con.executemany(
        "INSERT INTO pluviosidad_diaria (fecha, localidad, lat, lon, lluvia_mm, fuente, tipo_fuente) "
        "VALUES (?,?,?,?,?,?,?)",
        [(fe, l, la, lo, mm, fu, "estacion_real") for (fe, l, la, lo, mm, fu, _c) in filas_totales])
    con.commit()
    despues = con.execute("SELECT COUNT(*) FROM pluviosidad_diaria").fetchone()[0]
    con.close()
    print(f"Tabla consolidada: {antes:,} -> {despues:,} filas (+{despues - antes:,})")


if __name__ == "__main__":
    main()
